from functools import wraps
from datetime import date, datetime, timedelta, timezone
from decimal import Decimal, InvalidOperation
from io import BytesIO, StringIO
import base64
import csv
import json
import os
import re
import unicodedata
from uuid import uuid4

import click
import qrcode
import reportlab
from flask import (
    abort,
    current_app,
    Flask,
    flash,
    jsonify,
    redirect,
    render_template,
    request,
    send_file,
    send_from_directory,
    session,
    url_for,
)
from flask_migrate import Migrate
from flask_sqlalchemy import SQLAlchemy
from openpyxl import Workbook, load_workbook
from PIL import Image as PILImage, UnidentifiedImageError
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import (
    Image,
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)
from sqlalchemy import and_, or_
from werkzeug.utils import secure_filename
from werkzeug.security import check_password_hash, generate_password_hash

from config import Config

db = SQLAlchemy()
migrate = Migrate()

STATUS_LABELS = {
    "IN_STOCK": "Raktáron",
    "RESERVED": "Előjegyezve",
    "ISSUED": "Kiadva",
    "INSTALLED": "Telepítve",
    "RETURNED": "Visszavéve",
    "IN_SERVICE": "Szervizben",
    "SCRAPPED": "Selejtezve",
}

MOVEMENT_TYPE_LABELS = {
    "INBOUND": "Bevételezés",
    "RESERVE": "Előjegyzés",
    "ISSUE": "Kiadás",
    "INSTALL": "Telepítés",
    "RETURN": "Visszavétel",
    "SERVICE": "Szervizbe küldés",
    "SCRAP": "Selejtezés",
    "TRANSFER": "Áthelyezés",
    "RELEASE": "Foglalás feloldása",
    "REVERSAL": "Mozgás visszavonása",
}

CATEGORY_LABELS = {
    "EV charger": "EV töltő",
    "Parking controller": "Parkolásvezérlő",
    "Barrier gate": "Sorompó",
    "Sensor": "Szenzor",
    "Energy meter": "Fogyasztásmérő",
    "Network device": "Hálózati eszköz",
    "Cabinet": "Szekrény",
    "Sticker": "Matrica",
    "Camera": "Kamera",
    "Kiosk": "Kioszk",
    "Opener": "Nyitó eszköz",
    "Router": "Router",
    "Parkl box": "Parkl box",
    "Other": "Egyéb",
}

LOCATION_TYPE_LABELS = {
    "warehouse": "Raktár",
    "service_vehicle": "Szervizautó",
    "service": "Szerviz / javítás",
    "supplier": "Beszállító",
    "subcontractor_warehouse": "Alvállalkozó raktára",
}
LOGISTIC_LOCATION_TYPES = frozenset(LOCATION_TYPE_LABELS)
LEGACY_LOCATION_TYPE_LABELS = {
    "project_site": "Projekt helyszín (archív)",
    "installed": "Telepített helyszín (archív)",
}

PROJECT_STATUS_LABELS = {
    "planned": "Tervezett",
    "active": "Aktív",
    "handover": "Átadás alatt",
    "completed": "Lezárt",
}

ASSIGNMENT_STATUS_LABELS = {
    "unassigned": "Nincs hozzárendelve",
    "assigned": "Hozzárendelve",
    "ignored": "Nem releváns",
}

IMPORT_STATUS_LABELS = {
    "running": "Fut",
    "completed": "Kész",
    "rolled_back": "Visszavonva",
    "partial_rollback": "Részben visszavonva",
}

DEVICE_CURRENCIES = {"HUF", "EUR"}
DEVICE_QR_MODE_LABELS = {
    "none": "Nincs QR",
    "group": "Csoport QR",
    "individual": "Egyedi QR példányonként",
}

PHYSICAL_LOCATION_STATUSES = {"IN_STOCK", "RESERVED", "RETURNED", "IN_SERVICE"}
FREE_STOCK_STATUSES = {"IN_STOCK", "RETURNED"}
PROJECT_ACTIVE_STATUSES = {"RESERVED", "ISSUED", "INSTALLED"}

TEMPLATE_PROJECT_HEADERS = [
    "project_code", "project_name", "customer_name", "site_name", "address",
    "city", "country", "latitude", "longitude", "google_maps_url",
    "site_notes", "status", "notes",
]
OPTIONAL_TEMPLATE_PROJECT_HEADERS = {
    "city",
    "country",
    "latitude",
    "longitude",
    "google_maps_url",
    "site_notes",
}
TEMPLATE_DEVICE_HEADERS = [
    "project_code", "category", "product_name", "manufacturer", "model", "serial_number",
    "asset_tag", "quantity", "currency", "unit_net_price", "total_net_price", "vat_rate",
    "unit_gross_price", "total_gross_price", "location_name", "status", "tracking_mode",
    "unit_generation", "unit_code_prefix", "notes",
]
OPTIONAL_TEMPLATE_DEVICE_HEADERS = {
    "tracking_mode",
    "unit_generation",
    "unit_code_prefix",
}
TEMPLATE_LOCATION_HEADERS = ["location_name", "location_type", "address", "notes"]

INVENTORY_SHEETS = {
    "tolto",
    "toltok",
    "bmw tolto",
    "kioszk",
    "kamera",
    "egyeb",
    "nyito",
    "matricak",
}
ORPHAN_INVOICE_SHEET = "gazdatlanul"
IGNORED_IMPORT_SHEETS = {"workflow", "dashboard", "seged", "onkoltseg", "sheet1"}
UPLOAD_SUBDIR = "uploads"
DRAWING_UPLOAD_SUBDIR = "drawings"
WORK_ORDER_UPLOAD_SUBDIR = "work_orders"
M2M_UPLOAD_SUBDIR = "m2m"

M2M_STATUS_LABELS = {
    "active": "Aktív",
    "suspended": "Felfüggesztve",
    "inactive": "Inaktív",
    "cancelled": "Megszüntetve",
}

M2M_USAGE_SOURCE_LABELS = {
    "manual": "Kézi rögzítés",
    "import": "Import",
    "teltonika_api": "Teltonika API",
}
ALLOWED_DRAWING_EXTENSIONS = {"png", "jpg", "jpeg", "webp", "pdf"}
ALLOWED_PHOTO_EXTENSIONS = {"png", "jpg", "jpeg", "webp"}

WORK_ORDER_TYPE_LABELS = {
    "maintenance": "Karbantartás",
    "troubleshooting": "Hibaelhárítás",
    "cable_replacement": "Kábelcsere",
    "installation": "Telepítés",
    "site_visit": "Helyszíni kiszállás",
    "inspection": "Felülvizsgálat",
    "other": "Egyéb",
}

WORK_ORDER_STATUS_LABELS = {
    "draft": "Tervezet",
    "in_progress": "Folyamatban",
    "closed": "Lezárt",
    "pdf_generated": "PDF generálva",
}

WORK_ORDER_PHOTO_CATEGORY_LABELS = {
    "before": "Hiba előtti állapot",
    "during": "Munka közbeni állapot",
    "after": "Javítás utáni állapot",
}

PDF_FONT_REGULAR = "ParklSans"
PDF_FONT_BOLD = "ParklSans-Bold"

DRAWING_ICON_CATEGORIES = {
    "Parking/access": [
        ("barrier", "Sorompó"),
        ("entry_barrier", "Bejárati sorompó"),
        ("exit_barrier", "Kijárati sorompó"),
        ("loop_detector", "Hurokdetektor"),
        ("rfid_reader", "RFID olvasó"),
        ("parking_space", "Parkolóhely"),
        ("direction_arrow", "Irány nyíl"),
    ],
    "Cameras": [
        ("entry_anpr_camera", "Belépő ANPR kamera"),
        ("exit_anpr_camera", "Kilépő ANPR kamera"),
        ("overview_camera", "Áttekintő kamera"),
    ],
    "Charging": [
        ("ac_charger", "AC töltő"),
        ("dc_charger", "DC töltő"),
        ("charger_pedestal", "Töltőoszlop"),
        ("dlm_controller", "DLM vezérlő"),
        ("energy_meter", "Fogyasztásmérő"),
        ("ct", "Áramváltó"),
    ],
    "Network/IT": [
        ("rack", "Rack"),
        ("switch", "Switch"),
        ("poe_switch", "PoE switch"),
        ("router", "Router"),
        ("teltonika", "Teltonika"),
        ("parkl_box", "Raspberry Pi / Parkl box"),
        ("patch_panel", "Patch panel"),
    ],
    "Electrical": [
        ("distribution_board", "Elosztószekrény"),
        ("power_supply", "Tápegység"),
        ("breaker", "Kismegszakító"),
        ("busbar", "Sín / trunking"),
        ("cable_tray", "Kábeltálca"),
        ("wall_penetration", "Falfúrás"),
        ("floor_penetration", "Födémáttörés"),
        ("junction_box", "Kötődoboz"),
    ],
}

DRAWING_LINE_TYPES = [
    ("cat5e", "CAT5e / UTP", "#2563eb"),
    ("power", "Erősáramú kábel", "#dc2626"),
    ("barrier_control", "Sorompó vezérlés", "#f59e0b"),
    ("camera_network", "Kamera hálózat", "#7c3aed"),
    ("dlm", "DLM kommunikáció", "#0891b2"),
    ("main_supply", "Fő betáp", "#111827"),
    ("spare_conduit", "Tartalék védőcső", "#64748b"),
]


def create_app(config_class=Config):
    app = Flask(__name__)
    app.config.from_object(config_class)
    os.makedirs(app.instance_path, exist_ok=True)

    db.init_app(app)
    migrate.init_app(app, db)

    from models import (
        DEVICE_CATEGORIES,
        DEVICE_STATUSES,
        MOVEMENT_TYPES,
        TRACKING_MODES,
        USER_ROLES,
        USER_ROLE_LABELS,
        AuditLog,
        AuthRateLimit,
        BulkStockBalance,
        Device,
        DeviceUnit,
        ImportBatch,
        Location,
        M2MMonthlyUsage,
        M2MPackageHistory,
        M2MSubscription,
        Project,
        ProjectDrawing,
        StockMovement,
        UnassignedInvoiceItem,
        User,
        WorkOrder,
        WorkOrderMaterial,
        WorkOrderMeasurement,
        WorkOrderPhoto,
        WorkOrderTemplate,
    )

    def get_current_user():
        user_id = session.get("user_id")
        if not user_id:
            return None
        return db.session.get(User, user_id)

    def client_ip_address():
        forwarded_for = request.headers.get("X-Forwarded-For", "")
        if forwarded_for:
            return forwarded_for.split(",", 1)[0].strip()
        return request.remote_addr

    def aware_utc(value):
        if value and value.tzinfo is None:
            return value.replace(tzinfo=timezone.utc)
        return value

    def log_audit_event(event_type, user=None, username=None, success=None, details=None):
        db.session.add(
            AuditLog(
                user_id=user.id if user else None,
                event_type=event_type,
                username=username or (user.username if user else None),
                ip_address=client_ip_address(),
                user_agent=(request.user_agent.string or "")[:255],
                success=success,
                details=details,
            )
        )

    def lockout_identifier(username):
        return (username or "").strip().lower()

    def get_or_create_rate_limit(identifier):
        record = AuthRateLimit.query.filter_by(identifier=identifier).first()
        if record is None:
            record = AuthRateLimit(identifier=identifier)
            db.session.add(record)
            db.session.flush()
        return record

    def is_login_locked(user, rate_limit_record, now):
        locked_until_values = []
        if user and user.locked_until:
            locked_until_values.append(user.locked_until)
        if rate_limit_record and rate_limit_record.locked_until:
            locked_until_values.append(rate_limit_record.locked_until)
        return any(aware_utc(value) and aware_utc(value) > now for value in locked_until_values)

    def register_failed_login(user, rate_limit_record, now):
        max_attempts = app.config["LOGIN_MAX_FAILED_ATTEMPTS"]
        lockout_until = None
        if rate_limit_record:
            rate_limit_record.failed_count += 1
            rate_limit_record.last_failed_at = now
            if rate_limit_record.failed_count >= max_attempts:
                rate_limit_record.locked_until = now + timedelta(
                    minutes=app.config["LOGIN_LOCKOUT_MINUTES"]
                )
                lockout_until = rate_limit_record.locked_until
        if user:
            user.failed_login_count += 1
            if user.failed_login_count >= max_attempts:
                user.locked_until = now + timedelta(
                    minutes=app.config["LOGIN_LOCKOUT_MINUTES"]
                )
                lockout_until = user.locked_until
        return lockout_until

    def register_successful_login(user, rate_limit_record, now):
        user.failed_login_count = 0
        user.locked_until = None
        user.last_login_at = now
        user.last_seen_at = now
        if rate_limit_record:
            rate_limit_record.failed_count = 0
            rate_limit_record.locked_until = None
        session.clear()
        session.permanent = True
        session["user_id"] = user.id
        session["last_activity"] = now.isoformat()

    def user_can(*roles):
        user = get_current_user()
        return bool(user and user.is_active and user.has_role(*roles))

    @app.before_request
    def enforce_session_security():
        endpoint = request.endpoint or ""
        if endpoint in {"static", "login", "logout"}:
            return None
        user = get_current_user()
        if not user:
            return None

        now = datetime.now(timezone.utc)
        last_activity_raw = session.get("last_activity")
        if last_activity_raw:
            try:
                last_activity = datetime.fromisoformat(last_activity_raw)
            except ValueError:
                last_activity = None
            if last_activity and now - last_activity > app.config["PERMANENT_SESSION_LIFETIME"]:
                log_audit_event("session_timeout", user=user, success=True)
                db.session.commit()
                session.clear()
                flash("A munkamenet lejárt 8 óra inaktivitás után. Jelentkezz be újra.", "warning")
                return redirect(url_for("login"))

        session["last_activity"] = now.isoformat()
        user.last_seen_at = now
        db.session.commit()
        if user.force_password_change and endpoint != "change_password":
            flash("Az első belépéshez jelszócsere szükséges.", "warning")
            return redirect(url_for("change_password"))
        return None

    @app.context_processor
    def inject_current_user():
        user = get_current_user()
        return {
            "current_user": user,
            "can_write": user_can("admin", "manager"),
            "can_manage_work_orders": user_can("admin", "manager", "technician"),
            "can_export": user_can("admin", "manager"),
            "can_view_finance": user_can("admin", "manager"),
            "can_manage_users": user_can("admin"),
            "can_manage_m2m": user_can("admin", "manager"),
            "user_role_label": lambda value: USER_ROLE_LABELS.get(value, value or "–"),
            "status_label": status_label,
            "movement_type_label": movement_type_label,
            "category_label": category_label,
            "location_type_label": location_type_label,
            "project_status_label": project_status_label,
            "assignment_status_label": assignment_status_label,
            "import_status_label": import_status_label,
            "yes_no_label": yes_no_label,
            "format_number": format_number,
            "format_vat_rate": format_vat_rate,
            "line_net_amount": line_net_amount,
            "invoice_item_value": invoice_item_value,
            "device_display_label": device_display_label,
            "device_primary_label": device_primary_label,
            "device_active_project_codes": device_active_project_codes,
            "device_money_text": device_money_text,
            "device_qr_mode_label": device_qr_mode_label,
            "tracking_mode_label": tracking_mode_label,
            "bulk_balance_summary": bulk_balance_summary,
            "device_inventory_values": device_inventory_values,
            "status_badge_class": status_badge_class,
            "movement_badge_class": movement_badge_class,
            "work_order_type_label": work_order_type_label,
            "work_order_status_label": work_order_status_label,
            "work_order_photo_category_label": work_order_photo_category_label,
            "format_duration": format_duration,
            "template_json_rows": template_json_rows,
            "available_device_movements": available_device_movements,
            "current_date": date.today().isoformat(),
            "m2m_status_label": m2m_status_label,
            "m2m_usage_source_label": m2m_usage_source_label,
            "m2m_package_limit_mb": m2m_package_limit_mb,
            "m2m_usage_state": m2m_usage_state,
        }

    def login_required(view):
        @wraps(view)
        def wrapped_view(*args, **kwargs):
            user = get_current_user()
            if user is None:
                session.clear()
                flash("A folytatáshoz jelentkezz be.", "warning")
                return redirect(url_for("login"))
            if not user.is_active:
                session.clear()
                flash("A felhasználói fiók inaktív. Fordulj egy adminisztrátorhoz.", "danger")
                return redirect(url_for("login"))
            return view(*args, **kwargs)

        return wrapped_view

    def role_required(*roles):
        def decorator(view):
            @wraps(view)
            @login_required
            def wrapped_view(*args, **kwargs):
                if not user_can(*roles):
                    abort(403)
                return view(*args, **kwargs)

            return wrapped_view

        return decorator

    def method_role_required(*roles):
        def decorator(view):
            @wraps(view)
            @login_required
            def wrapped_view(*args, **kwargs):
                if request.method not in {"GET", "HEAD", "OPTIONS"} and not user_can(*roles):
                    abort(403)
                return view(*args, **kwargs)

            return wrapped_view

        return decorator

    def admin_required(view):
        return role_required("admin")(view)

    def write_required(view):
        return method_role_required("admin", "manager")(view)

    def export_required(view):
        return role_required("admin", "manager")(view)

    def finance_required(view):
        return role_required("admin", "manager")(view)

    def m2m_required(view):
        return role_required("admin", "manager")(view)

    def work_order_write_required(view):
        return method_role_required("admin", "manager", "technician")(view)

    def work_order_edit_required(view):
        return role_required("admin", "manager", "technician")(view)

    def manager_write_required(view):
        return role_required("admin", "manager")(view)

    @app.errorhandler(403)
    def forbidden(_error):
        if session.get("user_id"):
            flash("Ehhez a művelethez nincs jogosultságod.", "danger")
            return redirect(url_for("dashboard"))
        return redirect(url_for("login"))

    @app.cli.command("seed-role-users")
    def seed_role_users():
        """Create local role test users. Do not use in production."""
        if os.environ.get("FLASK_ENV", "").lower() == "production":
            raise click.ClickException(
                "A seed-role-users parancs production környezetben nem futtatható."
            )
        demo_users = {
            "admin": ("admin", "AdminDemo123!"),
            "manager": ("manager", "ManagerDemo123!"),
            "technician": ("technician", "TechnicianDemo123!"),
            "viewer": ("viewer", "ViewerDemo123!"),
        }
        for username, (role, password) in demo_users.items():
            user = User.query.filter_by(username=username).first()
            if user is None:
                user = User(username=username)
                db.session.add(user)
            user.password_hash = generate_password_hash(password)
            user.role = role
            user.is_admin = role == "admin"
            user.is_active = True
            user.force_password_change = False
            user.failed_login_count = 0
            user.locked_until = None
        db.session.commit()
        click.echo("Szerepkör tesztfelhasználók létrehozva/frissítve.")
        for username, (role, password) in demo_users.items():
            click.echo(f"- {username} / {password} ({role})")

    def validate_user_role(role):
        if role not in USER_ROLES:
            abort(400)
        return role

    def sync_admin_flag(user):
        user.is_admin = user.role == "admin"

    def prevent_last_active_admin_change(user, new_role=None, new_active=None):
        removes_admin = new_role is not None and new_role != "admin"
        deactivates_admin = new_active is False
        if user.effective_role != "admin" or not user.is_active:
            return
        if not removes_admin and not deactivates_admin:
            return
        active_admin_count = User.query.filter_by(role="admin", is_active=True).count()
        if active_admin_count <= 1:
            flash("Az utolsó aktív adminisztrátor nem módosítható vagy deaktiválható.", "danger")
            return False
        return True

    def get_user_or_404(user_id):
        user = db.session.get(User, user_id)
        if user is None:
            abort(404)
        return user

    def set_user_role(user, role):
        user.role = validate_user_role(role)
        sync_admin_flag(user)

    def set_user_active(user, is_active):
        user.is_active = is_active

    def current_user_is(user):
        current = get_current_user()
        return bool(current and current.id == user.id)

    def protect_current_user_deactivation(user, is_active):
        if current_user_is(user) and not is_active:
            flash("A saját felhasználói fiókodat nem deaktiválhatod.", "danger")
            return False
        return True

    def protect_current_user_role_change(user, role):
        if current_user_is(user) and role != "admin":
            flash("A saját admin szerepkörödet nem módosíthatod.", "danger")
            return False
        return True

    def apply_user_management_change(user, role=None, is_active=None):
        if role is not None:
            if not protect_current_user_role_change(user, role):
                return False
            if prevent_last_active_admin_change(user, new_role=role) is False:
                return False
            set_user_role(user, role)
        if is_active is not None:
            if not protect_current_user_deactivation(user, is_active):
                return False
            if prevent_last_active_admin_change(user, new_active=is_active) is False:
                return False
            set_user_active(user, is_active)
        return True

    def user_search_query(search):
        query = User.query
        if search:
            query = query.filter(User.username.ilike(f"%{search}%"))
        return query.order_by(User.username.asc())

    def recent_login_audit_query(limit=50):
        login_events = (
            "login_success",
            "login_failure",
            "login_locked",
            "login_inactive",
            "logout",
            "session_timeout",
            "password_changed",
            "password_change_failure",
        )
        return (
            AuditLog.query.filter(AuditLog.event_type.in_(login_events))
            .order_by(AuditLog.created_at.desc())
            .limit(limit)
        )

    def reject_inactive_login(user):
        if user and not user.is_active:
            flash("A felhasználói fiók inaktív. Fordulj egy adminisztrátorhoz.", "danger")
            return True
        return False

    def validate_new_password(password, confirmation=None):
        if len(password or "") < 12:
            return "Az új jelszó legalább 12 karakter legyen."
        if confirmation is not None and password != confirmation:
            return "A két új jelszó nem egyezik."
        return None

    @app.cli.command("seed-admin")
    @click.option("--password", hide_input=True, required=False)
    def seed_admin(password):
        """Create or update the default admin user."""
        username = app.config["ADMIN_USERNAME"]
        password = app.config["ADMIN_PASSWORD"] or password
        if not password:
            password = click.prompt(
                "Admin jelszó",
                hide_input=True,
                confirmation_prompt=True,
            )
        user = User.query.filter_by(username=username).first()
        if user is None:
            user = User(
                username=username,
                password_hash=generate_password_hash(password),
                is_admin=True,
                role="admin",
                is_active=True,
                force_password_change=True,
            )
            db.session.add(user)
            action = "Létrehozva"
        else:
            user.password_hash = generate_password_hash(password)
            user.is_admin = True
            user.role = "admin"
            user.is_active = True
            user.force_password_change = True
            user.failed_login_count = 0
            user.locked_until = None
            action = "Frissítve"
        for location in Location.query.all():
            if location.name == "Main Warehouse":
                location.name = "Fő raktár"
            elif location.name.startswith("Stock Room"):
                location.name = location.name.replace("Stock Room", "Raktár", 1)
            elif location.name.startswith("Site"):
                location.name = location.name.replace("Site", "Helyszín", 1)
        seed_work_order_templates(WorkOrderTemplate)
        db.session.commit()
        print(f"{action}: '{username}' admin felhasználó.")

    @app.cli.command("reset-demo-data")
    @click.option("--yes", is_flag=True, help="Megerősítés bekérésének kihagyása.")
    def reset_demo_data_command(yes):
        """Reset local business data and create a small Parkl demo dataset."""
        if os.environ.get("FLASK_ENV", "").lower() == "production":
            raise click.ClickException(
                "A reset-demo-data parancs production környezetben nem futtatható."
            )
        if not yes and not click.confirm(
            "Ez törli a helyi projekt-, eszköz-, lokáció-, import- és mozgásadatokat. Folytatod?"
        ):
            click.echo("Megszakítva.")
            return

        summary = reset_demo_dataset(
            app,
            User,
            Project,
            Location,
            Device,
            DeviceUnit,
            BulkStockBalance,
            StockMovement,
            UnassignedInvoiceItem,
            ImportBatch,
            ProjectDrawing,
            WorkOrder,
            WorkOrderMaterial,
            WorkOrderMeasurement,
            WorkOrderPhoto,
            WorkOrderTemplate,
        )
        click.echo(
            "Demo adatbázis újraépítve: "
            f"{summary['projects']} projekt, {summary['locations']} készlethely, "
            f"{summary['devices']} terméktétel, {summary['device_units']} egyedi példány, "
            f"{summary['bulk_physical']} db bulk készlet "
            f"({summary['bulk_reserved']} db foglalt), "
            f"{summary['unit_reserved']} foglalt példány, "
            f"{summary['movements']} mozgás, "
            f"{summary['invoice_items']} gazdátlan számlasor."
        )

    @app.route("/")
    def index():
        if session.get("user_id"):
            return redirect(url_for("dashboard"))
        return redirect(url_for("login"))

    @app.route("/login", methods=["GET", "POST"])
    def login():
        if request.method == "POST":
            username = request.form.get("username", "").strip()
            password = request.form.get("password", "")
            identifier = lockout_identifier(username) or f"ip:{client_ip_address() or 'unknown'}"
            now = datetime.now(timezone.utc)
            rate_limit = get_or_create_rate_limit(identifier)
            user = User.query.filter_by(username=username).first()
            if reject_inactive_login(user):
                log_audit_event("login_inactive", user=user, username=username, success=False)
                db.session.commit()
                return render_template("login.html")
            if is_login_locked(user, rate_limit, now):
                log_audit_event("login_locked", user=user, username=username, success=False)
                db.session.commit()
                flash("Túl sok hibás belépési kísérlet. Próbáld újra 15 perc múlva.", "danger")
                return render_template("login.html")
            if user and check_password_hash(user.password_hash, password):
                if user.is_admin and user.role != "admin":
                    user.role = "admin"
                register_successful_login(user, rate_limit, now)
                log_audit_event("login_success", user=user, username=username, success=True)
                db.session.commit()
                flash("Sikeres bejelentkezés.", "success")
                if user.force_password_change:
                    return redirect(url_for("change_password"))
                return redirect(url_for("dashboard"))
            lockout_until = register_failed_login(user, rate_limit, now)
            log_audit_event(
                "login_failure",
                user=user,
                username=username,
                success=False,
                details="locked" if lockout_until else None,
            )
            db.session.commit()
            if lockout_until:
                flash("Túl sok hibás belépési kísérlet. A fiók 15 percre tiltva.", "danger")
            else:
                remaining = app.config["LOGIN_MAX_FAILED_ATTEMPTS"]
                if rate_limit:
                    remaining = max(0, remaining - rate_limit.failed_count)
                flash(f"Hibás felhasználónév vagy jelszó. Hátralévő próbálkozás: {remaining}.", "danger")
        return render_template("login.html")

    @app.route("/change-password", methods=["GET", "POST"])
    @login_required
    def change_password():
        user = get_current_user()
        if request.method == "POST":
            current_password = request.form.get("current_password", "")
            new_password = request.form.get("new_password", "")
            confirmation = request.form.get("new_password_confirm", "")
            if not check_password_hash(user.password_hash, current_password):
                log_audit_event(
                    "password_change_failure",
                    user=user,
                    success=False,
                    details="bad_current_password",
                )
                db.session.commit()
                flash("A jelenlegi jelszó hibás.", "danger")
            else:
                error = validate_new_password(new_password, confirmation)
                if error:
                    flash(error, "danger")
                elif check_password_hash(user.password_hash, new_password):
                    flash("Az új jelszó nem egyezhet meg a jelenlegi jelszóval.", "danger")
                else:
                    user.password_hash = generate_password_hash(new_password)
                    user.force_password_change = False
                    user.failed_login_count = 0
                    user.locked_until = None
                    log_audit_event("password_changed", user=user, success=True)
                    db.session.commit()
                    flash("A jelszó módosítva.", "success")
                    return redirect(url_for("dashboard"))
        return render_template("change_password.html", forced=user.force_password_change)

    @app.route("/logout", methods=["POST"])
    @login_required
    def logout():
        user = get_current_user()
        if user:
            log_audit_event("logout", user=user, success=True)
            db.session.commit()
        session.clear()
        flash("Kijelentkeztél.", "info")
        return redirect(url_for("login"))

    @app.route("/dashboard")
    @login_required
    def dashboard():
        active_devices = Device.query.filter(Device.archived_at.is_(None)).all()
        active_units = (
            DeviceUnit.query.filter(DeviceUnit.archived_at.is_(None))
            .join(Device)
            .filter(Device.archived_at.is_(None), Device.tracking_mode == "unit")
            .all()
        )
        finance_visible = user_can("admin", "manager")
        active_invoice_items = (
            UnassignedInvoiceItem.query.filter(UnassignedInvoiceItem.archived_at.is_(None)).all()
            if finance_visible
            else []
        )
        attention_items = build_attention_items(
            active_devices, active_invoice_items, include_finance=finance_visible
        )
        stats = {
            "projects": Project.query.filter(Project.archived_at.is_(None)).count(),
            "locations": Location.query.filter(
                Location.archived_at.is_(None),
                Location.location_type.in_(LOGISTIC_LOCATION_TYPES),
            ).count(),
            "movements": StockMovement.query.count(),
            "in_stock": inventory_status_quantity(active_devices, active_units, "IN_STOCK"),
            "reserved": inventory_status_quantity(active_devices, active_units, "RESERVED"),
            "issued": inventory_status_quantity(active_devices, active_units, "ISSUED"),
            "installed": inventory_status_quantity(active_devices, active_units, "INSTALLED"),
            "awaiting_arrival": sum(1 for device in active_devices if is_awaiting_arrival(device)),
            "unassigned_invoices": sum(
                1
                for item in active_invoice_items
                if item.assignment_status == "unassigned"
            ),
            "financial_open": sum(1 for device in active_devices if is_financially_open(device)),
            "attention": len(attention_items),
        }
        recent_movements = (
            StockMovement.query.order_by(StockMovement.created_at.desc()).limit(6).all()
        )
        return render_template(
            "dashboard.html",
            stats=stats,
            recent_movements=recent_movements,
            attention_items=attention_items[:8],
        )

    @app.route("/attention")
    @login_required
    def attention():
        devices = Device.query.filter(Device.archived_at.is_(None)).all()
        finance_visible = user_can("admin", "manager")
        invoice_items = (
            UnassignedInvoiceItem.query.filter(UnassignedInvoiceItem.archived_at.is_(None)).all()
            if finance_visible
            else []
        )
        attention_items = build_attention_items(
            devices, invoice_items, include_finance=finance_visible
        )
        return render_template("attention.html", attention_items=attention_items)

    @app.route("/labels")
    @login_required
    def labels():
        active_devices = Device.query.filter(Device.archived_at.is_(None)).count()
        active_units = DeviceUnit.query.filter(DeviceUnit.archived_at.is_(None)).count()
        return render_template(
            "labels.html",
            active_devices=active_devices,
            active_units=active_units,
        )

    @app.route("/m2m")
    @m2m_required
    def m2m_subscriptions():
        search = request.args.get("q", "").strip()
        selected_status = request.args.get("status", "").strip()
        selected_location = request.args.get("location", "").strip()
        selected_package = request.args.get("package", "").strip()
        selected_usage = request.args.get("usage", "").strip()
        selected_view = request.args.get("view", "table").strip()
        if selected_view not in {"table", "cards"}:
            selected_view = "table"

        query = M2MSubscription.query
        if search:
            term = f"%{search}%"
            query = query.filter(
                or_(
                    M2MSubscription.subscriber_name.ilike(term),
                    M2MSubscription.account_number.ilike(term),
                    M2MSubscription.contract_number.ilike(term),
                    M2MSubscription.phone_number.ilike(term),
                    M2MSubscription.device_number.ilike(term),
                    M2MSubscription.device_identifier.ilike(term),
                    M2MSubscription.sim_number.ilike(term),
                    M2MSubscription.location_name.ilike(term),
                )
            )
        if selected_status in M2M_STATUS_LABELS:
            query = query.filter(M2MSubscription.status == selected_status)
        if selected_location:
            query = query.filter(M2MSubscription.location_name == selected_location)
        if selected_package:
            query = query.filter(M2MSubscription.current_package == selected_package)

        subscriptions = query.order_by(
            M2MSubscription.location_name.asc(),
            M2MSubscription.phone_number.asc(),
        ).all()
        rows = []
        for subscription in subscriptions:
            current_usage = m2m_current_usage(subscription)
            state = m2m_subscription_usage_state(subscription, current_usage)
            if selected_usage and state["key"] != selected_usage:
                continue
            rows.append(
                {
                    "subscription": subscription,
                    "current_usage": current_usage,
                    "usage_state": state,
                }
            )

        all_subscriptions = M2MSubscription.query.all()
        locations = [
            value[0]
            for value in db.session.query(M2MSubscription.location_name)
            .filter(
                M2MSubscription.location_name.is_not(None),
                M2MSubscription.location_name != "",
            )
            .distinct()
            .order_by(M2MSubscription.location_name.asc())
            .all()
        ]
        packages = [
            value[0]
            for value in db.session.query(M2MSubscription.current_package)
            .filter(
                M2MSubscription.current_package.is_not(None),
                M2MSubscription.current_package != "",
            )
            .distinct()
            .order_by(M2MSubscription.current_package.asc())
            .all()
        ]
        usage_states = [
            m2m_subscription_usage_state(item, m2m_current_usage(item))["key"]
            for item in all_subscriptions
        ]
        return render_template(
            "m2m_subscriptions.html",
            rows=rows,
            statuses=M2M_STATUS_LABELS,
            locations=locations,
            packages=packages,
            search=search,
            selected_status=selected_status,
            selected_location=selected_location,
            selected_package=selected_package,
            selected_usage=selected_usage,
            selected_view=selected_view,
            summary={
                "total": len(all_subscriptions),
                "active": sum(1 for item in all_subscriptions if item.status == "active"),
                "warning": usage_states.count("warning"),
                "exceeded": usage_states.count("exceeded"),
            },
        )

    @app.route("/m2m/new", methods=["GET", "POST"])
    @m2m_required
    def m2m_subscription_new():
        subscription = M2MSubscription(status="active")
        if request.method == "POST":
            error = update_m2m_subscription_from_form(
                subscription, request.form, Device
            )
            if error:
                flash(error, "danger")
            else:
                db.session.add(subscription)
                db.session.flush()
                if subscription.current_package:
                    db.session.add(
                        M2MPackageHistory(
                            subscription_id=subscription.id,
                            package_name=subscription.current_package,
                            monthly_fee=subscription.current_monthly_fee,
                            valid_from=subscription.registration_date or date.today(),
                            notes="Kezdő csomag a SIM létrehozásakor.",
                        )
                    )
                log_audit_event(
                    "m2m_subscription_created",
                    user=get_current_user(),
                    success=True,
                    details=f"subscription_id={subscription.id}",
                )
                db.session.commit()
                flash("Az M2M SIM előfizetés létrejött.", "success")
                return redirect(
                    url_for("m2m_subscription_detail", subscription_id=subscription.id)
                )
        return render_template(
            "m2m_subscription_form.html",
            subscription=subscription,
            statuses=M2M_STATUS_LABELS,
            devices=m2m_teltonika_devices(Device),
            form_title="Új M2M SIM",
            submit_label="Előfizetés létrehozása",
        )

    @app.route("/m2m/<int:subscription_id>")
    @m2m_required
    def m2m_subscription_detail(subscription_id):
        subscription = M2MSubscription.query.get_or_404(subscription_id)
        usage_history = m2m_effective_usage_history(subscription)
        current_usage = m2m_current_usage(subscription)
        return render_template(
            "m2m_subscription_detail.html",
            subscription=subscription,
            statuses=M2M_STATUS_LABELS,
            usage_sources=M2M_USAGE_SOURCE_LABELS,
            usage_history=usage_history,
            current_usage=current_usage,
            usage_state=m2m_subscription_usage_state(
                subscription, current_usage
            ),
            chart_labels=[
                f"{item['year']}-{item['month']:02d}" for item in usage_history[-12:]
            ],
            chart_values=[
                float(item["usage_mb"]) for item in usage_history[-12:]
            ],
        )

    @app.route("/m2m/<int:subscription_id>/edit", methods=["GET", "POST"])
    @m2m_required
    def m2m_subscription_edit(subscription_id):
        subscription = M2MSubscription.query.get_or_404(subscription_id)
        if request.method == "POST":
            error = update_m2m_subscription_from_form(
                subscription, request.form, Device
            )
            if error:
                flash(error, "danger")
            else:
                log_audit_event(
                    "m2m_subscription_updated",
                    user=get_current_user(),
                    success=True,
                    details=f"subscription_id={subscription.id}",
                )
                db.session.commit()
                flash("Az M2M előfizetés adatai frissültek.", "success")
                return redirect(
                    url_for("m2m_subscription_detail", subscription_id=subscription.id)
                )
        return render_template(
            "m2m_subscription_form.html",
            subscription=subscription,
            statuses=M2M_STATUS_LABELS,
            devices=m2m_teltonika_devices(Device),
            form_title="M2M SIM szerkesztése",
            submit_label="Módosítások mentése",
        )

    @app.route("/m2m/<int:subscription_id>/status", methods=["POST"])
    @m2m_required
    def m2m_subscription_status(subscription_id):
        subscription = M2MSubscription.query.get_or_404(subscription_id)
        status = request.form.get("status", "").strip()
        if status not in M2M_STATUS_LABELS:
            flash("Érvénytelen M2M státusz.", "danger")
        else:
            subscription.status = status
            log_audit_event(
                "m2m_status_changed",
                user=get_current_user(),
                success=True,
                details=f"subscription_id={subscription.id};status={status}",
            )
            db.session.commit()
            flash("Az előfizetés státusza frissült.", "success")
        return redirect(
            url_for("m2m_subscription_detail", subscription_id=subscription.id)
        )

    @app.route("/m2m/<int:subscription_id>/usage", methods=["POST"])
    @m2m_required
    def m2m_subscription_usage(subscription_id):
        subscription = M2MSubscription.query.get_or_404(subscription_id)
        year = optional_int(request.form.get("year"))
        month = optional_int(request.form.get("month"))
        usage_mb = optional_decimal(request.form.get("usage_mb"))
        source = request.form.get("source", "manual").strip()
        if not year or not 2000 <= year <= 2200:
            flash("Az év értéke nem érvényes.", "danger")
        elif not month or not 1 <= month <= 12:
            flash("A hónap értéke nem érvényes.", "danger")
        elif usage_mb is None or usage_mb < 0:
            flash("Az adatforgalom nem lehet üres vagy negatív.", "danger")
        elif source not in M2M_USAGE_SOURCE_LABELS:
            flash("Érvénytelen adatforrás.", "danger")
        else:
            usage = M2MMonthlyUsage.query.filter_by(
                subscription_id=subscription.id,
                year=year,
                month=month,
                source=source,
            ).first()
            if usage is None:
                usage = M2MMonthlyUsage(
                    subscription_id=subscription.id,
                    year=year,
                    month=month,
                    source=source,
                )
                db.session.add(usage)
            usage.usage_mb = usage_mb
            log_audit_event(
                "m2m_usage_saved",
                user=get_current_user(),
                success=True,
                details=(
                    f"subscription_id={subscription.id};"
                    f"period={year}-{month:02d};source={source}"
                ),
            )
            db.session.commit()
            flash("A havi adatforgalom mentve.", "success")
        return redirect(
            url_for("m2m_subscription_detail", subscription_id=subscription.id)
        )

    @app.route("/m2m/<int:subscription_id>/package", methods=["POST"])
    @m2m_required
    def m2m_subscription_package(subscription_id):
        subscription = M2MSubscription.query.get_or_404(subscription_id)
        package_name = request.form.get("package_name", "").strip()
        monthly_fee = optional_decimal(request.form.get("monthly_fee"))
        valid_from = optional_date(request.form.get("valid_from"))
        if not package_name:
            flash("A csomag neve kötelező.", "danger")
        elif not valid_from:
            flash("Az érvényesség kezdete kötelező.", "danger")
        elif monthly_fee is not None and monthly_fee < 0:
            flash("A havidíj nem lehet negatív.", "danger")
        else:
            active_history = (
                M2MPackageHistory.query.filter_by(
                    subscription_id=subscription.id, valid_to=None
                )
                .order_by(M2MPackageHistory.valid_from.desc())
                .first()
            )
            if active_history and active_history.valid_from < valid_from:
                active_history.valid_to = valid_from - timedelta(days=1)
            history = M2MPackageHistory(
                subscription_id=subscription.id,
                package_name=package_name,
                monthly_fee=monthly_fee,
                valid_from=valid_from,
                notes=request.form.get("notes", "").strip() or None,
            )
            subscription.current_package = package_name
            subscription.current_monthly_fee = monthly_fee
            db.session.add(history)
            log_audit_event(
                "m2m_package_changed",
                user=get_current_user(),
                success=True,
                details=f"subscription_id={subscription.id};package={package_name}",
            )
            db.session.commit()
            flash("A csomagváltás rögzítve.", "success")
        return redirect(
            url_for("m2m_subscription_detail", subscription_id=subscription.id)
        )

    @app.route("/m2m/import", methods=["GET", "POST"])
    @m2m_required
    def m2m_import():
        result = None
        if request.method == "POST":
            upload = request.files.get("m2m_file")
            if not upload or not upload.filename:
                flash("Válassz CSV vagy XLSX fájlt.", "danger")
            elif not upload.filename.lower().endswith((".csv", ".xlsx")):
                flash("Csak .csv vagy .xlsx fájl tölthető fel.", "danger")
            else:
                upload_dir = os.path.join(app.instance_path, M2M_UPLOAD_SUBDIR)
                os.makedirs(upload_dir, exist_ok=True)
                safe_name = secure_filename(upload.filename)
                upload_path = os.path.join(
                    upload_dir, f"m2m_{uuid4().hex}_{safe_name}"
                )
                upload.save(upload_path)
                try:
                    parsed_rows = parse_m2m_import_file(upload_path)
                    result = import_m2m_rows(
                        parsed_rows,
                        M2MSubscription,
                        M2MMonthlyUsage,
                        M2MPackageHistory,
                        Device,
                    )
                    log_audit_event(
                        "m2m_import",
                        user=get_current_user(),
                        success=not result["errors"],
                        details=(
                            f"filename={safe_name};created={result['created']};"
                            f"updated={result['updated']};usages={result['usages']};"
                            f"errors={len(result['errors'])}"
                        ),
                    )
                    db.session.commit()
                    flash(
                        "M2M import kész: "
                        f"{result['created']} új, {result['updated']} frissített SIM, "
                        f"{result['usages']} havi forgalmi rekord.",
                        "success" if not result["errors"] else "warning",
                    )
                except (ValueError, OSError) as exc:
                    db.session.rollback()
                    flash(f"Az M2M import nem sikerült: {exc}", "danger")
                finally:
                    if os.path.exists(upload_path):
                        os.remove(upload_path)
        return render_template("m2m_import.html", result=result)

    @app.route("/m2m/import/template")
    @m2m_required
    def m2m_import_template():
        return send_file(
            build_m2m_import_template(),
            mimetype=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
            as_attachment=True,
            download_name="Parkl_M2M_import_sablon.xlsx",
        )

    @app.route("/m2m/rms-sync", methods=["POST"])
    @m2m_required
    def m2m_rms_sync():
        from services.teltonika_rms import (
            TeltonikaRMSError,
            list_rms_devices,
            sync_rms_devices_to_m2m,
            sync_rms_usage_to_m2m,
        )

        try:
            raw_devices = list_rms_devices()
            device_result = sync_rms_devices_to_m2m(raw_devices)
            usage_result = sync_rms_usage_to_m2m(raw_devices)
            log_audit_event(
                "m2m_rms_sync",
                user=get_current_user(),
                success=True,
                details=(
                    f"devices={device_result['rms_devices']};"
                    f"linked={device_result['linked_by_iccid']};"
                    f"mobile={device_result['mobile_updated']};"
                    f"skipped={device_result['skipped_wired_unknown']};"
                    f"usage_requested={usage_result['usage_requested']};"
                    f"usage_chunks={usage_result['usage_chunk_requests']};"
                    f"usage_daily={usage_result['usage_daily_records']};"
                    f"usage_created={usage_result['usage_created']};"
                    f"usage_updated={usage_result['usage_updated']};"
                    f"errors={device_result['errors'] + usage_result['errors']}"
                ),
            )
            db.session.commit()
            scope_warning = (
                " Valószínűleg hiányzik a "
                "company_device_statistics:read RMS scope."
                if usage_result["scope_errors"]
                else ""
            )
            flash(
                "RMS szinkron kész: "
                f"{device_result['rms_devices']} RMS eszköz, "
                f"{device_result['linked_by_iccid']} ICCID kapcsolat, "
                f"{device_result['mobile_updated']} mobile frissítés, "
                f"{device_result['skipped_wired_unknown']} wired/unknown kihagyva, "
                f"{device_result['errors'] + usage_result['errors']} hibás rekord. "
                f"Havi data usage lekérés: {usage_result['usage_requested']} SIM, "
                f"{usage_result['usage_chunk_requests']} chunk, "
                f"{usage_result['usage_daily_records']} napi rekord, "
                f"{format_number(usage_result['usage_total_mb'])} MB összesen, "
                f"{usage_result['usage_created']} új, "
                f"{usage_result['usage_updated']} frissített, "
                f"{usage_result['usage_no_data']} adat nélkül, "
                f"{usage_result['errors']} hibás."
                f"{scope_warning}",
                "warning" if usage_result["errors"] else "success",
            )
        except TeltonikaRMSError as exc:
            db.session.rollback()
            current_app.logger.exception(
                "A Teltonika RMS szinkron ismert RMS hibával leállt. "
                "exception_type=%s",
                type(exc).__name__,
            )
            log_audit_event(
                "m2m_rms_sync",
                user=get_current_user(),
                success=False,
                details=exc.__class__.__name__,
            )
            db.session.commit()
            flash(f"Az RMS szinkron nem sikerült: {exc}", "danger")
        except Exception as exc:
            db.session.rollback()
            current_app.logger.exception(
                "A Teltonika RMS szinkron váratlan hibával leállt. "
                "exception_type=%s",
                type(exc).__name__,
            )
            log_audit_event(
                "m2m_rms_sync",
                user=get_current_user(),
                success=False,
                details="UnexpectedError",
            )
            db.session.commit()
            flash(
                "Az RMS szinkron váratlan hiba miatt nem sikerült. "
                "A részletek a szerver naplójában láthatók.",
                "danger",
            )
        return redirect(url_for("m2m_subscriptions"))

    @app.route("/drawings")
    @login_required
    def drawings():
        drawing_list = (
            ProjectDrawing.query.join(Project)
            .filter(Project.archived_at.is_(None))
            .order_by(ProjectDrawing.updated_at.desc())
            .all()
        )
        projects_with_drawings = Project.query.filter(
            Project.archived_at.is_(None),
            Project.drawings.any(),
        ).order_by(Project.name.asc()).all()
        return render_template(
            "drawings.html",
            drawings=drawing_list,
            projects=projects_with_drawings,
        )

    @app.route("/import-export", methods=["GET", "POST"])
    @export_required
    def import_export():
        pending_import = session.get("pending_template_import")
        preview = None
        if request.method == "POST":
            action = request.form.get("action", "dry_run")
            if action == "confirm":
                if request.form.get("execute_import") != "on":
                    flash("Az importálás végrehajtásához jelöld be a megerősítést.", "danger")
                    return redirect(url_for("import_export"))
                if not pending_import or not os.path.exists(pending_import["path"]):
                    flash("Nincs érvényes előnézeti import. Töltsd fel újra a sablont.", "danger")
                    return redirect(url_for("import_export"))
                preview = parse_template_workbook(
                    pending_import["path"], Project, Device, Location
                )
                if preview["critical_error_count"]:
                    flash("Az import nem véglegesíthető, mert kritikus hibák vannak.", "danger")
                    return redirect(url_for("import_export"))
                result = import_template_workbook(
                    preview, Project, Device, Location, session["user_id"]
                )
                db.session.commit()
                session.pop("pending_template_import", None)
                flash(
                    "Sablonimport kész: "
                    f"{result['projects_created']} projekt, "
                    f"{result['locations_created']} készlethely és "
                    f"{result['devices_created']} eszköztétel, "
                    f"{result['units_created']} egyedi példány létrehozva.",
                    "success",
                )
                return redirect(url_for("import_export"))

            upload = request.files.get("template_file")
            if not upload or upload.filename == "":
                flash("Válassz ki egy .xlsx sablonfájlt.", "danger")
                return redirect(url_for("import_export"))
            if not upload.filename.lower().endswith(".xlsx"):
                flash("Csak .xlsx fájl tölthető fel.", "danger")
                return redirect(url_for("import_export"))
            upload_dir = os.path.join(app.instance_path, UPLOAD_SUBDIR)
            os.makedirs(upload_dir, exist_ok=True)
            safe_name = secure_filename(upload.filename)
            upload_path = os.path.join(upload_dir, f"template_{uuid4().hex}_{safe_name}")
            upload.save(upload_path)
            preview = parse_template_workbook(upload_path, Project, Device, Location)
            session["pending_template_import"] = {"path": upload_path, "filename": safe_name}
            flash("Sablonimport előnézet elkészült. Az adatbázis még nem módosult.", "info")

        if preview is None and pending_import and os.path.exists(pending_import["path"]):
            preview = parse_template_workbook(
                pending_import["path"], Project, Device, Location
            )
        return render_template(
            "import_export.html",
            preview=preview,
            pending_import=session.get("pending_template_import"),
        )

    @app.route("/import-export/template")
    @export_required
    def import_template_download():
        return send_file(
            build_import_template_workbook(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name="Parkl_Infra_Manager_import_sablon.xlsx",
        )

    @app.route("/import-export/export/<export_type>")
    @export_required
    def data_export(export_type):
        if export_type not in {"devices", "projects", "locations"}:
            abort(404)
        return send_file(
            build_data_export_workbook(export_type, Project, Device, Location),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=f"Parkl_{export_type}_export.xlsx",
        )

    @app.route("/help")
    @login_required
    def help_page():
        return render_template("help.html")

    @app.route("/documents")
    @export_required
    def documents():
        return render_template("documents.html")

    @app.route("/admin")
    @admin_required
    def admin_tools():
        return render_template("admin_tools.html")

    @app.route("/admin/users")
    @admin_required
    def admin_users():
        search = request.args.get("q", "").strip()
        users = user_search_query(search).all()
        return render_template(
            "admin_users.html",
            users=users,
            search=search,
            roles=USER_ROLES,
            role_labels=USER_ROLE_LABELS,
            login_audit_logs=recent_login_audit_query().all(),
        )

    @app.route("/admin/users/new", methods=["GET", "POST"])
    @admin_required
    def admin_user_new():
        if request.method == "POST":
            username = request.form.get("username", "").strip()
            role = validate_user_role(request.form.get("role", "viewer"))
            password = request.form.get("temporary_password", "")
            if not username:
                flash("A felhasználónév kötelező.", "danger")
            elif User.query.filter_by(username=username).first():
                flash("Ezzel a felhasználónévvel már létezik fiók.", "danger")
            else:
                error = validate_new_password(password)
                if error:
                    flash(error, "danger")
                else:
                    user = User(
                        username=username,
                        password_hash=generate_password_hash(password),
                        role=role,
                        is_admin=role == "admin",
                        is_active=True,
                        force_password_change=True,
                    )
                    db.session.add(user)
                    log_audit_event(
                        "user_created",
                        user=get_current_user(),
                        username=username,
                        success=True,
                        details=f"role={role}",
                    )
                    db.session.commit()
                    flash("A felhasználó létrejött. Első belépéskor jelszót kell cserélnie.", "success")
                    return redirect(url_for("admin_users"))
        return render_template(
            "admin_user_form.html",
            roles=USER_ROLES,
            role_labels=USER_ROLE_LABELS,
        )

    @app.route("/admin/users/<int:user_id>/role", methods=["POST"])
    @admin_required
    def admin_user_role(user_id):
        user = get_user_or_404(user_id)
        role = validate_user_role(request.form.get("role", ""))
        if not apply_user_management_change(user, role=role):
            return redirect(url_for("admin_users"))
        log_audit_event(
            "user_role_changed",
            user=get_current_user(),
            username=user.username,
            success=True,
            details=f"role={role}",
        )
        db.session.commit()
        flash(f"{user.username} szerepköre módosítva.", "success")
        return redirect(url_for("admin_users"))

    @app.route("/admin/users/<int:user_id>/toggle-active", methods=["POST"])
    @admin_required
    def admin_user_toggle_active(user_id):
        user = get_user_or_404(user_id)
        new_active = not user.is_active
        if not apply_user_management_change(user, is_active=new_active):
            return redirect(url_for("admin_users"))
        log_audit_event(
            "user_activation_changed",
            user=get_current_user(),
            username=user.username,
            success=True,
            details=f"is_active={new_active}",
        )
        db.session.commit()
        state = "aktiválva" if user.is_active else "deaktiválva"
        flash(f"{user.username} felhasználó {state}.", "success")
        return redirect(url_for("admin_users"))

    @app.route("/admin/users/<int:user_id>/force-password-change", methods=["POST"])
    @admin_required
    def admin_user_force_password_change(user_id):
        user = get_user_or_404(user_id)
        user.force_password_change = True
        log_audit_event(
            "user_force_password_change",
            user=get_current_user(),
            username=user.username,
            success=True,
        )
        db.session.commit()
        flash(f"{user.username} következő belépéskor jelszót cserél.", "success")
        return redirect(url_for("admin_users"))

    @app.route("/legacy")
    @admin_required
    def legacy():
        recent_batches = (
            ImportBatch.query.filter(ImportBatch.archived_at.is_(None))
            .order_by(ImportBatch.created_at.desc())
            .limit(8)
            .all()
        )
        return render_template("legacy.html", recent_batches=recent_batches)

    @app.route("/work-orders")
    @login_required
    def work_orders():
        query = WorkOrder.query.filter(WorkOrder.archived_at.is_(None))
        search = request.args.get("q", "").strip()
        work_type = request.args.get("work_type", "").strip()
        status = request.args.get("status", "").strip()
        date_from = optional_date(request.args.get("date_from"))
        date_to = optional_date(request.args.get("date_to"))
        if search:
            term = f"%{search}%"
            query = query.filter(
                or_(
                    WorkOrder.number.ilike(term),
                    WorkOrder.customer_name.ilike(term),
                    WorkOrder.site_name.ilike(term),
                    WorkOrder.technician_name.ilike(term),
                )
            )
        if work_type in WORK_ORDER_TYPE_LABELS:
            query = query.filter(WorkOrder.work_type == work_type)
        if status in WORK_ORDER_STATUS_LABELS:
            query = query.filter(WorkOrder.status == status)
        if date_from:
            query = query.filter(WorkOrder.work_date >= date_from)
        if date_to:
            query = query.filter(WorkOrder.work_date <= date_to)
        return render_template(
            "work_orders.html",
            work_orders=query.order_by(WorkOrder.created_at.desc()).all(),
            work_order_types=WORK_ORDER_TYPE_LABELS,
            work_order_statuses=WORK_ORDER_STATUS_LABELS,
            search=search,
            selected_work_type=work_type,
            selected_status=status,
            date_from=date_from,
            date_to=date_to,
        )

    @app.route("/work-orders/new", methods=["GET", "POST"])
    @work_order_edit_required
    def work_order_new():
        template = None
        template_id = optional_int(request.args.get("template_id"))
        if template_id:
            template = WorkOrderTemplate.query.get_or_404(template_id)
        if request.method == "POST":
            work_order = WorkOrder(
                number=request.form.get("number", "").strip(),
                work_type=request.form.get("work_type", "").strip(),
                created_date=optional_date(request.form.get("created_date")) or date.today(),
                created_by_id=session["user_id"],
            )
            error = update_work_order_from_form(work_order, request.form)
            if error:
                flash(error, "danger")
            elif WorkOrder.query.filter_by(number=work_order.number).first():
                flash("Ezzel a munkalapszámmal már létezik munkalap.", "danger")
            else:
                db.session.add(work_order)
                db.session.flush()
                replace_work_order_rows(work_order, request.form, WorkOrderMaterial, WorkOrderMeasurement)
                save_work_order_uploads(app, work_order, request.files, request.form, WorkOrderPhoto)
                db.session.commit()
                flash("A munkalap létrejött.", "success")
                return redirect(url_for("work_order_detail", work_order_id=work_order.id))
        return render_template(
            "work_order_form.html",
            work_order=None,
            template=template,
            template_materials=template_json_rows(template, "materials_json"),
            template_measurements=template_json_rows(template, "measurements_json"),
            suggested_number=next_work_order_number(WorkOrder),
            work_order_types=WORK_ORDER_TYPE_LABELS,
            work_order_statuses=WORK_ORDER_STATUS_LABELS,
            photo_categories=WORK_ORDER_PHOTO_CATEGORY_LABELS,
        )

    @app.route("/work-orders/<int:work_order_id>")
    @login_required
    def work_order_detail(work_order_id):
        work_order = WorkOrder.query.get_or_404(work_order_id)
        return render_template(
            "work_order_detail.html",
            work_order=work_order,
            photo_categories=WORK_ORDER_PHOTO_CATEGORY_LABELS,
        )

    @app.route("/work-orders/<int:work_order_id>/edit", methods=["GET", "POST"])
    @work_order_edit_required
    def work_order_edit(work_order_id):
        work_order = WorkOrder.query.get_or_404(work_order_id)
        if request.method == "POST":
            old_number = work_order.number
            error = update_work_order_from_form(work_order, request.form)
            duplicate = (
                WorkOrder.query.filter(WorkOrder.id != work_order.id)
                .filter(WorkOrder.number == work_order.number)
                .first()
            )
            if error:
                work_order.number = old_number
                flash(error, "danger")
            elif duplicate:
                work_order.number = old_number
                flash("Ezzel a munkalapszámmal már létezik másik munkalap.", "danger")
            else:
                replace_work_order_rows(work_order, request.form, WorkOrderMaterial, WorkOrderMeasurement)
                save_work_order_uploads(app, work_order, request.files, request.form, WorkOrderPhoto)
                db.session.commit()
                flash("A munkalap módosítva.", "success")
                return redirect(url_for("work_order_detail", work_order_id=work_order.id))
        return render_template(
            "work_order_form.html",
            work_order=work_order,
            template=None,
            template_materials=[],
            template_measurements=[],
            work_order_types=WORK_ORDER_TYPE_LABELS,
            work_order_statuses=WORK_ORDER_STATUS_LABELS,
            photo_categories=WORK_ORDER_PHOTO_CATEGORY_LABELS,
        )

    @app.route("/work-orders/<int:work_order_id>/copy", methods=["POST"])
    @work_order_edit_required
    def work_order_copy(work_order_id):
        source = WorkOrder.query.get_or_404(work_order_id)
        copied = copy_work_order(source, session["user_id"], WorkOrder, WorkOrderMaterial, WorkOrderMeasurement)
        db.session.add(copied)
        db.session.commit()
        flash("A munkalap másolata létrejött tervezetként.", "success")
        return redirect(url_for("work_order_edit", work_order_id=copied.id))

    @app.route("/work-orders/<int:work_order_id>/archive", methods=["POST"])
    @work_order_edit_required
    def work_order_archive(work_order_id):
        work_order = WorkOrder.query.get_or_404(work_order_id)
        work_order.archived_at = now_utc()
        db.session.commit()
        flash("A munkalap archiválva.", "info")
        return redirect(url_for("work_orders"))

    @app.route("/work-orders/<int:work_order_id>/pdf")
    @export_required
    def work_order_pdf(work_order_id):
        work_order = WorkOrder.query.get_or_404(work_order_id)
        pdf_buffer = build_work_order_pdf(app, work_order)
        work_order.status = "pdf_generated"
        work_order.pdf_generated_at = now_utc()
        db.session.commit()
        return send_file(
            pdf_buffer,
            mimetype="application/pdf",
            as_attachment=True,
            download_name=secure_filename(f"MUNKALAP_{work_order.number}.pdf"),
        )

    @app.route("/work-orders/files/<path:filename>")
    @login_required
    def work_order_file(filename):
        return send_from_directory(
            os.path.join(app.instance_path, WORK_ORDER_UPLOAD_SUBDIR),
            filename,
        )

    @app.route("/work-order-templates")
    @login_required
    def work_order_templates():
        templates = (
            WorkOrderTemplate.query.filter(WorkOrderTemplate.archived_at.is_(None))
            .order_by(WorkOrderTemplate.name.asc())
            .all()
        )
        return render_template(
            "work_order_templates.html",
            templates=templates,
            work_order_types=WORK_ORDER_TYPE_LABELS,
        )

    @app.route("/work-order-templates/new", methods=["GET", "POST"])
    @manager_write_required
    def work_order_template_new():
        template = WorkOrderTemplate()
        if request.method == "POST":
            error = update_work_order_template_from_form(template, request.form)
            if error:
                flash(error, "danger")
            elif WorkOrderTemplate.query.filter_by(name=template.name).first():
                flash("Ezzel a névvel már létezik sablon.", "danger")
            else:
                db.session.add(template)
                db.session.commit()
                flash("A munkalap sablon létrejött.", "success")
                return redirect(url_for("work_order_templates"))
        return render_template(
            "work_order_template_form.html",
            template=None,
            work_order_types=WORK_ORDER_TYPE_LABELS,
        )

    @app.route("/work-order-templates/<int:template_id>/edit", methods=["GET", "POST"])
    @manager_write_required
    def work_order_template_edit(template_id):
        template = WorkOrderTemplate.query.get_or_404(template_id)
        if request.method == "POST":
            error = update_work_order_template_from_form(template, request.form)
            duplicate = (
                WorkOrderTemplate.query.filter(WorkOrderTemplate.id != template.id)
                .filter(WorkOrderTemplate.name == template.name)
                .first()
            )
            if error:
                flash(error, "danger")
            elif duplicate:
                flash("Ezzel a névvel már létezik másik sablon.", "danger")
            else:
                db.session.commit()
                flash("A munkalap sablon módosítva.", "success")
                return redirect(url_for("work_order_templates"))
        return render_template(
            "work_order_template_form.html",
            template=template,
            work_order_types=WORK_ORDER_TYPE_LABELS,
        )

    @app.route("/work-order-templates/<int:template_id>/archive", methods=["POST"])
    @manager_write_required
    def work_order_template_archive(template_id):
        template = WorkOrderTemplate.query.get_or_404(template_id)
        template.archived_at = now_utc()
        db.session.commit()
        flash("A munkalap sablon archiválva.", "info")
        return redirect(url_for("work_order_templates"))

    @app.route("/projects")
    @write_required
    def projects():
        search = request.args.get("q", "").strip()
        selected_status = request.args.get("status", "").strip()
        project_query = Project.query.filter(Project.archived_at.is_(None))
        if search:
            term = f"%{search}%"
            project_query = project_query.filter(
                or_(
                    Project.code.ilike(term),
                    Project.name.ilike(term),
                    Project.customer.ilike(term),
                    Project.site_name.ilike(term),
                    Project.address.ilike(term),
                    Project.city.ilike(term),
                )
            )
        if selected_status in PROJECT_STATUS_LABELS:
            project_query = project_query.filter(Project.status == selected_status)

        project_list = project_query.order_by(Project.created_at.desc()).all()
        project_inventory = {
            project.id: project_inventory_summary(project) for project in project_list
        }
        return render_template(
            "projects.html",
            projects=project_list,
            project_inventory=project_inventory,
            project_statuses=PROJECT_STATUS_LABELS,
            search=search,
            selected_status=selected_status,
        )

    @app.route("/projects/new", methods=["GET", "POST"])
    @manager_write_required
    def project_new():
        if request.method == "POST":
            project_data, location_error = project_form_data(request.form)
            name = project_data["name"]
            code = project_data["code"]
            if not name or not code:
                flash("A projekt neve és kódja kötelező.", "danger")
            elif location_error:
                flash(location_error, "danger")
            elif Project.query.filter_by(code=code).first():
                flash("Ezzel a kóddal már létezik projekt.", "danger")
            else:
                project = Project(**project_data)
                db.session.add(project)
                db.session.commit()
                flash("A projekt létrejött.", "success")
                return redirect(url_for("project_detail", project_id=project.id))
        return render_template(
            "project_form.html",
            project=None,
            project_statuses=PROJECT_STATUS_LABELS,
            form_title="Új projekt",
            submit_label="Projekt létrehozása",
            cancel_url=url_for("projects"),
        )

    @app.route("/projects/<int:project_id>")
    @login_required
    def project_detail(project_id):
        project = Project.query.get_or_404(project_id)
        devices = (
            Device.query.join(BulkStockBalance)
            .filter(
                BulkStockBalance.project_id == project.id,
                BulkStockBalance.quantity > 0,
                BulkStockBalance.status.in_(PROJECT_ACTIVE_STATUSES),
                Device.archived_at.is_(None),
                Device.tracking_mode == "bulk",
            )
            .distinct()
            .order_by(Device.asset_tag.asc())
            .all()
        )
        bulk_balances = (
            BulkStockBalance.query.filter(
                BulkStockBalance.project_id == project.id,
                BulkStockBalance.quantity > 0,
                BulkStockBalance.status.in_(PROJECT_ACTIVE_STATUSES),
            )
            .join(Device)
            .filter(Device.archived_at.is_(None), Device.tracking_mode == "bulk")
            .order_by(BulkStockBalance.id.asc())
            .all()
        )
        units = (
            DeviceUnit.query.filter_by(project_id=project.id)
            .filter(
                DeviceUnit.archived_at.is_(None),
                DeviceUnit.status.in_(PROJECT_ACTIVE_STATUSES),
            )
            .join(Device)
            .filter(Device.archived_at.is_(None), Device.tracking_mode == "unit")
            .order_by(DeviceUnit.unit_code.asc())
            .all()
        )
        reserved_units = [unit for unit in units if unit.status == "RESERVED"]
        issued_units = [unit for unit in units if unit.status == "ISSUED"]
        installed_units = [unit for unit in units if unit.status == "INSTALLED"]
        inventory_rows = [
            {
                "device": balance.device,
                "unit": None,
                "quantity": balance.quantity,
                "status": balance.status,
                "location": balance.location,
            }
            for balance in bulk_balances
        ] + [
            {
                "device": unit.device,
                "unit": unit,
                "quantity": 1,
                "status": unit.status,
                "location": unit.location,
            }
            for unit in units
        ]
        project_devices = list(
            {row["device"].id: row["device"] for row in inventory_rows}.values()
        )
        movements = (
            StockMovement.query.filter(
                or_(
                    StockMovement.project_id == project.id,
                    StockMovement.from_project_id == project.id,
                    StockMovement.to_project_id == project.id,
                )
            )
            .order_by(StockMovement.created_at.desc())
            .limit(50)
            .all()
        )
        drawings = (
            ProjectDrawing.query.filter_by(project_id=project.id)
            .order_by(ProjectDrawing.updated_at.desc())
            .all()
        )
        finance_summary = {
            "device_count": sum(balance.quantity for balance in bulk_balances) + len(units),
            "quantity": sum(balance.quantity for balance in bulk_balances) + len(units),
            "bulk_quantity": sum(balance.quantity for balance in bulk_balances),
            **inventory_rows_currency_totals(inventory_rows),
            "invoice_value": sum(device.invoice_value or 0 for device in project_devices),
            "ordered": sum(1 for device in project_devices if device.is_ordered),
            "arrived": sum(1 for device in project_devices if device.has_arrived),
            "issued": sum(balance.quantity for balance in bulk_balances if balance.status == "ISSUED")
            + sum(1 for unit in units if unit.status == "ISSUED"),
            "installed": sum(balance.quantity for balance in bulk_balances if balance.status == "INSTALLED")
            + sum(1 for unit in units if unit.status == "INSTALLED"),
            "returned": sum(balance.quantity for balance in bulk_balances if balance.status == "RETURNED")
            + sum(1 for unit in units if unit.status == "RETURNED"),
            "unpaid_supplier_invoice_count": sum(
                1
                for device in project_devices
                if device.supplier_invoice_number and device.supplier_invoice_paid is not True
            ),
            "awaiting_arrival_count": sum(
                1 for device in project_devices if is_awaiting_arrival(device)
            ),
        }
        finance_visible = user_can("admin", "manager")
        attention_items = [
            {
                "device": device,
                "reasons": device_attention_reasons(device, include_finance=finance_visible),
            }
            for device in project_devices
            if device_attention_reasons(device, include_finance=finance_visible)
        ]
        warehouses = (
            Location.query.filter(
                Location.archived_at.is_(None),
                Location.location_type == "warehouse",
            )
            .order_by(Location.name.asc())
            .all()
        )
        return render_template(
            "project_detail.html",
            project=project,
            devices=project_devices,
            bulk_balances=bulk_balances,
            units=units,
            reserved_units=reserved_units,
            issued_units=issued_units,
            installed_units=installed_units,
            movements=movements,
            drawings=drawings,
            finance_summary=finance_summary,
            attention_items=attention_items,
            warehouses=warehouses,
            archive_blockers=project_archive_blockers(project),
            reversed_movement_ids=reversed_movement_ids(movements),
            reversible_movement_ids=reversible_movement_ids(movements),
        )

    @app.route(
        "/projects/<int:project_id>/bulk-balances/<int:balance_id>/return",
        methods=["POST"],
    )
    @manager_write_required
    def project_bulk_return(project_id, balance_id):
        project = Project.query.get_or_404(project_id)
        balance = (
            BulkStockBalance.query.filter_by(id=balance_id, project_id=project.id)
            .with_for_update()
            .first_or_404()
        )
        quantity = optional_float(request.form.get("quantity"))
        to_location_id = optional_int(request.form.get("to_location_id"))
        warehouse = db.session.get(Location, to_location_id) if to_location_id else None
        if warehouse is None or warehouse.archived_at is not None:
            flash("Visszavételhez aktív célraktár kiválasztása kötelező.", "danger")
            return redirect(url_for("project_detail", project_id=project.id))
        if warehouse.location_type != "warehouse":
            flash("Projektből történő visszavétel célja csak raktár lehet.", "danger")
            return redirect(url_for("project_detail", project_id=project.id))

        error = validate_movement(
            balance.device,
            "RETURN",
            to_location_id=warehouse.id,
            project_id=None,
            quantity=quantity,
            source_balance=balance,
        )
        if error:
            flash(error, "danger")
            return redirect(url_for("project_detail", project_id=project.id))
        try:
            create_movement(
                device=balance.device,
                movement_type="RETURN",
                quantity=quantity,
                to_location_id=warehouse.id,
                source_balance=balance,
                notes=request.form.get("notes", "").strip()
                or f"Részleges visszavétel a(z) {project.code} projektből.",
                user_id=session["user_id"],
            )
            apply_device_state(
                balance.device,
                "RETURN",
                warehouse.id,
                None,
                quantity=quantity,
                source_balance=balance,
            )
            db.session.commit()
        except ValueError as error:
            db.session.rollback()
            flash(str(error), "danger")
            return redirect(url_for("project_detail", project_id=project.id))
        flash(
            f"{format_number(quantity)} db visszavétele rögzítve a(z) {warehouse.name} raktárba.",
            "success",
        )
        return redirect(url_for("project_detail", project_id=project.id))

    @app.route(
        "/projects/<int:project_id>/device-units/<int:unit_id>/return",
        methods=["POST"],
    )
    @manager_write_required
    def project_unit_return(project_id, unit_id):
        project = Project.query.get_or_404(project_id)
        unit = (
            DeviceUnit.query.filter_by(id=unit_id, project_id=project.id)
            .with_for_update()
            .first_or_404()
        )
        to_location_id = optional_int(request.form.get("to_location_id"))
        warehouse = db.session.get(Location, to_location_id) if to_location_id else None
        if warehouse is None or warehouse.archived_at is not None:
            flash("Visszavételhez aktív célraktár kiválasztása kötelező.", "danger")
            return redirect(url_for("project_detail", project_id=project.id))
        if warehouse.location_type != "warehouse":
            flash("Projektből történő visszavétel célja csak raktár lehet.", "danger")
            return redirect(url_for("project_detail", project_id=project.id))

        error = validate_movement(
            unit.device,
            "RETURN",
            to_location_id=warehouse.id,
            project_id=None,
            quantity=1,
            unit=unit,
        )
        if error:
            flash(error, "danger")
            return redirect(url_for("project_detail", project_id=project.id))
        create_movement(
            device=unit.device,
            unit=unit,
            movement_type="RETURN",
            quantity=1,
            to_location_id=warehouse.id,
            notes=request.form.get("notes", "").strip()
            or f"Visszavétel a(z) {project.code} projektből.",
            user_id=session["user_id"],
        )
        apply_device_state(unit.device, "RETURN", warehouse.id, None, unit=unit)
        db.session.commit()
        flash(f"A(z) {unit.unit_code} példány visszavétele rögzítve.", "success")
        return redirect(url_for("project_detail", project_id=project.id))

    @app.route("/projects/<int:project_id>/pdf/<pdf_type>")
    @export_required
    def project_pdf(project_id, pdf_type):
        project = Project.query.get_or_404(project_id)
        if pdf_type not in {"equipment", "issue", "installation", "finance"}:
            abort(404)

        pdf_buffer = build_project_pdf(
            project,
            project_inventory_rows(project),
            pdf_type,
        )
        filenames = {
            "equipment": "projekt-eszkozlista",
            "issue": "kiadasi-lista",
            "installation": "telepitesi-lista",
            "finance": "penzugyi-osszesito",
        }
        filename = f"{project.code}-{filenames[pdf_type]}.pdf"
        return send_file(
            pdf_buffer,
            mimetype="application/pdf",
            as_attachment=True,
            download_name=filename,
        )

    @app.route("/projects/<int:project_id>/qr")
    @login_required
    def project_qr(project_id):
        project = Project.query.get_or_404(project_id)
        return qr_png_response(
            url_for("project_detail", project_id=project.id, _external=True),
            f"{project.code}-qr.png",
        )

    @app.route("/projects/<int:project_id>/drawings", methods=["POST"])
    @manager_write_required
    def project_drawing_create(project_id):
        project = Project.query.get_or_404(project_id)
        name = request.form.get("name", "").strip() or "Helyszíni rajz"
        upload = request.files.get("background_image")
        background_filename = None
        if upload and upload.filename:
            if not allowed_drawing_file(upload.filename):
                flash(
                    "Csak PDF, PNG, JPG, JPEG vagy WEBP alaprajz tölthető fel.",
                    "danger",
                )
                return redirect(url_for("project_detail", project_id=project.id) + "#drawings")
            background_filename = save_drawing_background(app, upload, project.id)

        drawing = ProjectDrawing(
            project_id=project.id,
            name=name,
            background_filename=background_filename,
            canvas_json=json.dumps({"version": "5.3.0", "objects": []}),
        )
        db.session.add(drawing)
        db.session.commit()
        flash("A rajz létrejött.", "success")
        return redirect(
            url_for("project_drawing_editor", project_id=project.id, drawing_id=drawing.id)
        )

    @app.route("/projects/<int:project_id>/drawings/<int:drawing_id>")
    @manager_write_required
    def project_drawing_editor(project_id, drawing_id):
        project = Project.query.get_or_404(project_id)
        drawing = ProjectDrawing.query.filter_by(
            id=drawing_id, project_id=project.id
        ).first_or_404()
        background_url = None
        if drawing.background_filename:
            background_url = url_for(
                "drawing_background", filename=drawing.background_filename
            )
        drawing_units = (
            DeviceUnit.query.join(Device)
            .filter(
                DeviceUnit.project_id == project.id,
                DeviceUnit.archived_at.is_(None),
                Device.archived_at.is_(None),
            )
            .order_by(Device.product_name.asc(), DeviceUnit.unit_code.asc())
            .all()
        )
        drawing_bulk_balances = (
            BulkStockBalance.query.join(Device)
            .filter(
                BulkStockBalance.project_id == project.id,
                BulkStockBalance.quantity > 0,
                Device.archived_at.is_(None),
            )
            .order_by(Device.product_name.asc())
            .all()
        )
        project_drawing_items = [
            {
                "kind": "unit",
                "unit_id": unit.id,
                "device_id": unit.device_id,
                "code": unit.unit_code or unit.asset_tag or unit.serial_number,
                "label": unit.device.product_name or unit.device.name,
                "category": unit.device.device_type,
                "status": unit.status,
                "quantity": 1,
            }
            for unit in drawing_units
        ] + [
            {
                "kind": "bulk",
                "balance_id": balance.id,
                "device_id": balance.device_id,
                "code": balance.device.asset_tag,
                "label": balance.device.product_name or balance.device.name,
                "category": balance.device.device_type,
                "status": balance.status,
                "quantity": balance.quantity,
            }
            for balance in drawing_bulk_balances
        ]
        return render_template(
            "drawing_editor.html",
            project=project,
            drawing=drawing,
            background_url=background_url,
            icon_categories=DRAWING_ICON_CATEGORIES,
            line_types=DRAWING_LINE_TYPES,
            project_drawing_items=project_drawing_items,
        )

    @app.route("/projects/<int:project_id>/drawings/<int:drawing_id>/save", methods=["POST"])
    @manager_write_required
    def project_drawing_save(project_id, drawing_id):
        drawing = ProjectDrawing.query.filter_by(
            id=drawing_id, project_id=project_id
        ).first_or_404()
        payload = request.get_json(silent=True) or {}
        canvas_json = payload.get("canvas_json")
        if not isinstance(canvas_json, str):
            return jsonify({"ok": False, "error": "Hiányzó rajz JSON."}), 400
        drawing.canvas_json = canvas_json
        drawing.updated_at = now_utc()
        db.session.commit()
        return jsonify({"ok": True, "updated_at": drawing.updated_at.isoformat()})

    @app.route("/drawing-backgrounds/<path:filename>")
    @login_required
    def drawing_background(filename):
        upload_dir = os.path.join(app.instance_path, DRAWING_UPLOAD_SUBDIR)
        return send_from_directory(upload_dir, filename)

    @app.route("/projects/<int:project_id>/edit", methods=["GET", "POST"])
    @manager_write_required
    def project_edit(project_id):
        project = Project.query.get_or_404(project_id)
        if request.method == "POST":
            project_data, location_error = project_form_data(request.form)
            name = project_data["name"]
            code = project_data["code"]
            if not name or not code:
                flash("A projekt neve és kódja kötelező.", "danger")
            elif location_error:
                flash(location_error, "danger")
            elif (
                Project.query.filter(Project.id != project.id)
                .filter(Project.code == code)
                .first()
            ):
                flash("Ezzel a kóddal már létezik másik projekt.", "danger")
            else:
                for field, value in project_data.items():
                    setattr(project, field, value)
                db.session.commit()
                flash("A projekt módosítva.", "success")
                return redirect(url_for("project_detail", project_id=project.id))
        return render_template(
            "project_form.html",
            project=project,
            project_statuses=PROJECT_STATUS_LABELS,
            form_title="Projekt szerkesztése",
            submit_label="Mentés",
            cancel_url=url_for("project_detail", project_id=project.id),
        )

    @app.route("/projects/<int:project_id>/archive", methods=["POST"])
    @manager_write_required
    def project_archive(project_id):
        project = Project.query.get_or_404(project_id)
        blockers = project_archive_blockers(project)
        if blockers:
            flash_archive_blockers("A projekt nem archiválható.", blockers)
            return redirect(url_for("project_detail", project_id=project.id))
        project.archived_at = now_utc()
        db.session.commit()
        flash("A projekt archiválva.", "info")
        return redirect(url_for("projects"))

    @app.route("/devices")
    @write_required
    def devices():
        projects = Project.query.filter(Project.archived_at.is_(None)).order_by(Project.name.asc()).all()
        locations = active_logistic_locations()
        source_sheets = [
            row[0]
            for row in db.session.query(Device.source_sheet)
            .filter(Device.source_sheet.isnot(None))
            .distinct()
            .order_by(Device.source_sheet.asc())
            .all()
        ]
        selected_status = request.args.get("status", "").strip()
        selected_category = request.args.get("category", "").strip()
        selected_source_sheet = request.args.get("source_sheet", "").strip()
        selected_project_id = optional_int(request.args.get("project_id"))
        selected_location_id = optional_int(request.args.get("location_id"))
        search = request.args.get("q", "").strip()
        selected_view = request.args.get("view", "all").strip() or "all"

        device_query = Device.query.filter(Device.archived_at.is_(None))
        if search:
            term = f"%{search}%"
            device_query = device_query.filter(
                or_(
                    Device.product_name.ilike(term),
                    Device.asset_tag.ilike(term),
                    Device.serial_number.ilike(term),
                    Device.supplier_manufacturer.ilike(term),
                    Device.units.any(
                        and_(
                            DeviceUnit.archived_at.is_(None),
                            or_(
                                DeviceUnit.unit_code.ilike(term),
                                DeviceUnit.asset_tag.ilike(term),
                                DeviceUnit.serial_number.ilike(term),
                                DeviceUnit.project.has(Project.code.ilike(term)),
                            ),
                        )
                    ),
                )
            )
        if selected_status in DEVICE_STATUSES:
            device_query = device_query.filter(
                or_(
                    and_(
                        Device.tracking_mode == "bulk",
                        Device.bulk_balances.any(
                            and_(
                                BulkStockBalance.quantity > 0,
                                BulkStockBalance.status == selected_status,
                            )
                        ),
                    ),
                    and_(
                        Device.tracking_mode == "unit",
                        Device.units.any(
                            and_(
                                DeviceUnit.archived_at.is_(None),
                                DeviceUnit.status == selected_status,
                            )
                        ),
                    ),
                )
            )
        if selected_category in DEVICE_CATEGORIES:
            device_query = device_query.filter(Device.device_type == selected_category)
        if selected_source_sheet:
            device_query = device_query.filter(Device.source_sheet == selected_source_sheet)
        if selected_project_id:
            device_query = device_query.filter(
                or_(
                    and_(
                        Device.tracking_mode == "bulk",
                        Device.bulk_balances.any(
                            and_(
                                BulkStockBalance.quantity > 0,
                                BulkStockBalance.project_id == selected_project_id,
                                BulkStockBalance.status.in_(PROJECT_ACTIVE_STATUSES),
                            )
                        ),
                    ),
                    and_(
                        Device.tracking_mode == "unit",
                        Device.units.any(
                            and_(
                                DeviceUnit.archived_at.is_(None),
                                DeviceUnit.project_id == selected_project_id,
                                DeviceUnit.status.in_(PROJECT_ACTIVE_STATUSES),
                            )
                        ),
                    ),
                )
            )
        if selected_location_id:
            device_query = device_query.filter(
                or_(
                    and_(
                        Device.tracking_mode == "bulk",
                        Device.bulk_balances.any(
                            and_(
                                BulkStockBalance.quantity > 0,
                                BulkStockBalance.location_id == selected_location_id,
                                BulkStockBalance.status.in_(PHYSICAL_LOCATION_STATUSES),
                            )
                        ),
                    ),
                    and_(
                        Device.tracking_mode == "unit",
                        Device.units.any(
                            and_(
                                DeviceUnit.archived_at.is_(None),
                                DeviceUnit.location_id == selected_location_id,
                                DeviceUnit.status.in_(PHYSICAL_LOCATION_STATUSES),
                            )
                        ),
                    ),
                )
            )

        device_list = device_query.order_by(Device.created_at.desc()).all()
        quick_filter = request.args.get("quick_filter", "").strip()
        workflow_filter = quick_filter or selected_view
        if workflow_filter == "financial_open" and not user_can("admin", "manager"):
            abort(403)
        if workflow_filter == "in_stock":
            device_list = [device for device in device_list if device_has_status(device, "IN_STOCK")]
        elif workflow_filter == "assigned":
            device_list = [device for device in device_list if device_has_project(device)]
        elif workflow_filter == "issued":
            device_list = [device for device in device_list if device_has_status(device, "ISSUED")]
        elif workflow_filter == "installed":
            device_list = [device for device in device_list if device_has_status(device, "INSTALLED")]
        elif workflow_filter == "attention":
            device_list = [
                device
                for device in device_list
                if device_attention_reasons(device, include_finance=user_can("admin", "manager"))
            ]
        elif workflow_filter == "financial_open":
            device_list = [device for device in device_list if is_financially_open(device)]
        elif workflow_filter == "awaiting_arrival":
            device_list = [device for device in device_list if is_awaiting_arrival(device)]
        elif workflow_filter == "arrived_unassigned":
            device_list = [device for device in device_list if is_arrived_unassigned(device)]
        visible_summary = {
            "count": len(device_list),
            **device_currency_totals(device_list),
            "unpaid_invoice_count": (
                sum(1 for device in device_list if is_financially_open(device))
                if user_can("admin", "manager")
                else 0
            ),
            "awaiting_arrival_count": sum(1 for device in device_list if is_awaiting_arrival(device)),
        }
        return render_template(
            "devices.html",
            devices=device_list,
            projects=projects,
            locations=locations,
            statuses=DEVICE_STATUSES,
            categories=DEVICE_CATEGORIES,
            selected_status=selected_status,
            selected_category=selected_category,
            source_sheets=source_sheets,
            selected_source_sheet=selected_source_sheet,
            selected_project_id=selected_project_id,
            selected_location_id=selected_location_id,
            search=search,
            quick_filter=quick_filter,
            selected_view=selected_view,
            visible_summary=visible_summary,
        )

    @app.route("/devices/new", methods=["GET", "POST"])
    @manager_write_required
    def device_new():
        projects = Project.query.filter(Project.archived_at.is_(None)).order_by(Project.name.asc()).all()
        locations = active_logistic_locations()
        if request.method == "POST":
            data = device_form_data(request.form)
            inventory_kind = request.form.get("inventory_kind", "").strip()
            initial_state = request.form.get("initial_state", "IN_STOCK").strip()
            unit_code_prefix = request.form.get("unit_code_prefix", "").strip()
            data["tracking_mode"] = "unit" if inventory_kind == "unit" else "bulk"
            data["qr_mode"] = "individual" if inventory_kind == "unit" else "group"
            data["serial_number"] = ""
            data["project_id"] = None
            initial_location_id = data["location_id"]
            initial_project_id = optional_int(request.form.get("initial_project_id"))
            if not data["asset_tag"]:
                data["asset_tag"] = unique_device_asset_tag(
                    Device,
                    data["model"] or data["product_name"] or data["device_type"],
                )

            if inventory_kind not in {"unit", "bulk"}:
                flash("Válaszd ki, hogy egyedi eszközöket vagy mennyiségi készletet hozol létre.", "danger")
            elif not data["product_name"] or not data["device_type"]:
                flash("A termék neve és a kategória kötelező.", "danger")
            elif data["device_type"] not in DEVICE_CATEGORIES:
                flash("Érvénytelen eszközkategória.", "danger")
            elif data["currency"] and data["currency"] not in DEVICE_CURRENCIES:
                flash("Érvénytelen deviza. Válassz HUF vagy EUR értéket.", "danger")
            elif data["quantity"] is None or data["quantity"] <= 0:
                flash("Pozitív mennyiség megadása kötelező.", "danger")
            elif inventory_kind == "unit" and not float(data["quantity"]).is_integer():
                flash("Egyedi eszközöknél a darabszám csak egész szám lehet.", "danger")
            elif not data["location_id"]:
                flash("A kezdő készlethely megadása kötelező.", "danger")
            elif initial_state not in {"IN_STOCK", "RESERVED", "ISSUED"}:
                flash("Érvénytelen kezdő állapot.", "danger")
            elif initial_state in {"RESERVED", "ISSUED"} and not initial_project_id:
                flash("Projektfoglaláshoz vagy közvetlen kiadáshoz projekt választása kötelező.", "danger")
            elif Device.query.filter_by(asset_tag=data["asset_tag"]).first():
                flash("Ezzel az eszközazonosítóval már létezik eszköz.", "danger")
            else:
                try:
                    device = Device(**data, status="IN_STOCK")
                    db.session.add(device)
                    db.session.flush()
                    if device.tracking_mode == "unit":
                        prefix = unit_code_prefix or data["model"] or data["product_name"]
                        unit_codes = available_unit_codes(
                            DeviceUnit,
                            prefix,
                            1,
                            int(device.quantity),
                        )
                        for unit_code in unit_codes:
                            unit = DeviceUnit(
                                device=device,
                                unit_code=unit_code,
                                status="IN_STOCK",
                                location_id=initial_location_id,
                            )
                            db.session.add(unit)
                            db.session.flush()
                            create_movement(
                                device=device,
                                unit=unit,
                                movement_type="INBOUND",
                                quantity=1,
                                to_location_id=initial_location_id,
                                notes="Új egyedi eszköz kezdő bevételezése.",
                                user_id=session["user_id"],
                            )
                            if initial_state != "IN_STOCK":
                                create_movement(
                                    device=device,
                                    unit=unit,
                                    movement_type=(
                                        "RESERVE"
                                        if initial_state == "RESERVED"
                                        else "ISSUE"
                                    ),
                                    quantity=1,
                                    from_location_id=unit.location_id,
                                    project_id=initial_project_id,
                                    notes="Létrehozáskor választott kezdő projektállapot.",
                                    user_id=session["user_id"],
                                )
                                apply_device_state(
                                    device,
                                    "RESERVE" if initial_state == "RESERVED" else "ISSUE",
                                    None,
                                    initial_project_id,
                                    unit=unit,
                                )
                    else:
                        create_movement(
                            device=device,
                            movement_type="INBOUND",
                            quantity=device.quantity,
                            to_location_id=initial_location_id,
                            notes="Új mennyiségi készlet kezdő bevételezése.",
                            user_id=session["user_id"],
                        )
                        apply_device_state(
                            device,
                            "INBOUND",
                            initial_location_id,
                            None,
                            quantity=device.quantity,
                        )
                        if initial_state != "IN_STOCK":
                            source_balance = infer_bulk_source_balance(
                                device,
                                "RESERVE" if initial_state == "RESERVED" else "ISSUE",
                                device.quantity,
                            )
                            movement_type = (
                                "RESERVE" if initial_state == "RESERVED" else "ISSUE"
                            )
                            create_movement(
                                device=device,
                                movement_type=movement_type,
                                quantity=device.quantity,
                                project_id=initial_project_id,
                                source_balance=source_balance,
                                notes="Létrehozáskor választott kezdő projektállapot.",
                                user_id=session["user_id"],
                            )
                            apply_device_state(
                                device,
                                movement_type,
                                None,
                                initial_project_id,
                                quantity=device.quantity,
                                source_balance=source_balance,
                            )
                    db.session.commit()
                except ValueError as error:
                    db.session.rollback()
                    flash(str(error), "danger")
                else:
                    if device.tracking_mode == "unit":
                        flash(
                            f"Az eszköztétel és {int(device.quantity)} egyedi példány létrejött.",
                            "success",
                        )
                    else:
                        flash("A mennyiségi készlet és a kezdő mozgások létrejöttek.", "success")
                    return redirect(url_for("device_detail", device_id=device.id))
        return render_template(
            "device_new.html",
            projects=projects,
            locations=locations,
            categories=DEVICE_CATEGORIES,
        )

    @app.route("/devices/<int:device_id>")
    @login_required
    def device_detail(device_id):
        device = Device.query.get_or_404(device_id)
        bulk_balances = active_bulk_balances(device)
        projects = Project.query.filter(Project.archived_at.is_(None)).order_by(Project.name.asc()).all()
        locations = active_logistic_locations()
        units = (
            DeviceUnit.query.filter_by(device_id=device.id)
            .filter(DeviceUnit.archived_at.is_(None))
            .order_by(DeviceUnit.unit_code.asc())
            .all()
        )
        unit_summary = {
            "total": len(units),
            **{
                status: sum(1 for unit in units if unit.status == status)
                for status in DEVICE_STATUSES
            },
        }
        movements = (
            StockMovement.query.filter_by(device_id=device.id)
            .order_by(StockMovement.created_at.desc())
            .all()
        )
        return render_template(
            "device_detail.html",
            device=device,
            movements=movements,
            projects=projects,
            locations=locations,
            units=units,
            unit_summary=unit_summary,
            bulk_balances=bulk_balances,
            archive_blockers=device_archive_blockers(device),
            reversed_movement_ids=reversed_movement_ids(movements),
            reversible_movement_ids=reversible_movement_ids(movements),
        )

    @app.route("/devices/<int:device_id>/edit", methods=["GET", "POST"])
    @manager_write_required
    def device_edit(device_id):
        device = Device.query.get_or_404(device_id)
        projects = Project.query.filter(Project.archived_at.is_(None)).order_by(Project.name.asc()).all()
        locations = active_logistic_locations()
        if request.method == "POST":
            data = device_form_data(request.form)
            data.pop("project_id", None)
            data.pop("location_id", None)
            if not data["asset_tag"] or not data["device_type"]:
                flash("Az eszközazonosító és a kategória kötelező.", "danger")
            elif data["device_type"] not in DEVICE_CATEGORIES:
                flash("Érvénytelen eszközkategória.", "danger")
            elif data["currency"] and data["currency"] not in DEVICE_CURRENCIES:
                flash("Érvénytelen deviza. Válassz HUF vagy EUR értéket.", "danger")
            elif data["qr_mode"] not in DEVICE_QR_MODE_LABELS:
                flash("Érvénytelen QR mód.", "danger")
            elif data["tracking_mode"] not in TRACKING_MODES:
                flash("Érvénytelen követési mód.", "danger")
            elif device.tracking_mode == "bulk" and data["tracking_mode"] == "unit":
                flash(
                    "Bulk tételt a Példányok létrehozása művelettel állíts egyedi követésre.",
                    "danger",
                )
            elif (
                data["tracking_mode"] == "bulk"
                and any(unit.archived_at is None for unit in device.units)
            ):
                flash(
                    "Aktív eszközpéldányokkal rendelkező tétel nem állítható vissza mennyiségi követésre.",
                    "danger",
                )
            elif not device_quantity_supports_existing_units(device, data["quantity"]):
                flash(
                    "A mennyiség nem lehet kisebb a már létrehozott aktív példányok számánál.",
                    "danger",
                )
            elif (
                Device.query.filter(Device.id != device.id)
                .filter(Device.asset_tag == data["asset_tag"])
                .first()
            ):
                flash("Ezzel az eszközazonosítóval már létezik másik eszköz.", "danger")
            else:
                for field, value in data.items():
                    setattr(device, field, value)
                db.session.commit()
                flash("Az eszköz módosítva. A státusz nem változott.", "success")
                return redirect(url_for("device_detail", device_id=device.id))
        return render_template(
            "device_edit.html",
            device=device,
            projects=projects,
            locations=locations,
            categories=DEVICE_CATEGORIES,
        )

    @app.route("/devices/<int:device_id>/qr")
    @login_required
    def device_qr(device_id):
        device = Device.query.get_or_404(device_id)
        return qr_png_response(
            url_for("device_detail", device_id=device.id, _external=True),
            f"{device.asset_tag}-qr.png",
        )

    @app.route("/devices/<int:device_id>/label")
    @login_required
    def device_label(device_id):
        device = Device.query.get_or_404(device_id)
        return render_template("device_label.html", device=device)

    @app.route("/devices/<int:device_id>/units")
    @login_required
    def device_units(device_id):
        device = Device.query.get_or_404(device_id)
        units = (
            DeviceUnit.query.filter_by(device_id=device.id)
            .filter(DeviceUnit.archived_at.is_(None))
            .order_by(DeviceUnit.unit_code.asc())
            .all()
        )
        return render_template("device_units.html", device=device, units=units)

    @app.route("/devices/<int:device_id>/units/create", methods=["GET", "POST"])
    @manager_write_required
    def device_units_create(device_id):
        device = Device.query.get_or_404(device_id)
        bulk_balances = active_bulk_balances(device)
        source_balance = bulk_balances[0] if bulk_balances else None
        locations = (
            Location.query.filter(
                Location.archived_at.is_(None),
                Location.location_type.in_(LOGISTIC_LOCATION_TYPES),
            )
            .order_by(Location.name.asc())
            .all()
        )
        if device.tracking_mode == "bulk" and len(bulk_balances) > 1:
            flash(
                "Több készletegyenlegre bontott bulk tétel csak az egyenlegek összevonása után példányosítható.",
                "warning",
            )
            return redirect(url_for("device_detail", device_id=device.id))
        quantity = whole_device_quantity(device)
        active_units = (
            DeviceUnit.query.filter_by(device_id=device.id)
            .filter(DeviceUnit.archived_at.is_(None))
            .count()
        )
        if quantity is None:
            flash("Példányok csak pozitív egész mennyiségű tételből hozhatók létre.", "warning")
            return redirect(url_for("device_detail", device_id=device.id))
        missing_count = max(quantity - active_units, 0)
        if missing_count == 0:
            flash("Ehhez a tételhez már minden példány létrejött.", "info")
            return redirect(url_for("device_units", device_id=device.id))

        prefix = request.form.get("prefix", "").strip() or default_unit_code_prefix(device)
        start_number = optional_int(request.form.get("start_number")) or 1
        generated_codes = available_unit_codes(DeviceUnit, prefix, start_number, missing_count)

        if request.method == "POST":
            if request.form.get("confirm") != "1":
                flash("A példányok létrehozásához erősítsd meg a műveletet.", "warning")
            else:
                initial_location_id = optional_int(request.form.get("initial_location_id"))
                initial_location = (
                    db.session.get(Location, initial_location_id)
                    if initial_location_id
                    else None
                )
                if source_balance is None and initial_location is None:
                    flash(
                        "Új fizikai példányokhoz kezdő készlethely megadása kötelező.",
                        "danger",
                    )
                    return redirect(url_for("device_units_create", device_id=device.id))
                device.tracking_mode = "unit"
                for unit_code in generated_codes:
                    unit = DeviceUnit(
                        device=device,
                        unit_code=unit_code,
                        status=source_balance.status if source_balance else "IN_STOCK",
                        location_id=(
                            source_balance.location_id
                            if source_balance
                            else initial_location.id
                        ),
                        project_id=source_balance.project_id if source_balance else None,
                    )
                    db.session.add(unit)
                    db.session.flush()
                    if source_balance is None:
                        create_movement(
                            device=device,
                            unit=unit,
                            movement_type="INBOUND",
                            quantity=1,
                            to_location_id=initial_location.id,
                            notes="Példány létrehozásakor rögzített kezdő bevételezés.",
                            user_id=session["user_id"],
                        )
                if source_balance is not None:
                    source_balance.quantity = 0
                device.location_id = None
                device.project_id = None
                if device.qr_mode != "individual":
                    device.qr_mode = "individual"
                db.session.commit()
                flash(f"{len(generated_codes)} eszközpéldány létrejött.", "success")
                return redirect(url_for("device_units", device_id=device.id))

        return render_template(
            "device_units_create.html",
            device=device,
            quantity=quantity,
            active_units=active_units,
            missing_count=missing_count,
            prefix=prefix,
            start_number=start_number,
            generated_codes=generated_codes,
            source_balance=source_balance,
            locations=locations,
        )

    @app.route("/devices/<int:device_id>/unit-labels.pdf")
    @login_required
    def device_unit_labels_pdf(device_id):
        device = Device.query.get_or_404(device_id)
        units = (
            DeviceUnit.query.filter_by(device_id=device.id)
            .filter(DeviceUnit.archived_at.is_(None))
            .order_by(DeviceUnit.unit_code.asc())
            .all()
        )
        if not units:
            flash("Nincs nyomtatható eszközpéldány.", "warning")
            return redirect(url_for("device_detail", device_id=device.id))
        unit_urls = {
            unit.id: url_for("device_unit_detail", unit_id=unit.id, _external=True)
            for unit in units
        }
        buffer = build_device_unit_labels_pdf(device, units, unit_urls)
        return send_file(
            buffer,
            mimetype="application/pdf",
            as_attachment=True,
            download_name=secure_filename(f"{device.asset_tag}_QR_cimkek.pdf"),
        )

    @app.route("/device-units/<int:unit_id>")
    @login_required
    def device_unit_detail(unit_id):
        unit = DeviceUnit.query.get_or_404(unit_id)
        projects = Project.query.filter(Project.archived_at.is_(None)).order_by(Project.name.asc()).all()
        locations = active_logistic_locations()
        movements = (
            StockMovement.query.filter_by(unit_id=unit.id)
            .order_by(StockMovement.created_at.desc())
            .all()
        )
        return render_template(
            "device_unit_detail.html",
            unit=unit,
            device=unit.device,
            projects=projects,
            locations=locations,
            movements=movements,
            archive_blockers=device_unit_archive_blockers(unit),
            reversed_movement_ids=reversed_movement_ids(movements),
            reversible_movement_ids=reversible_movement_ids(movements),
        )

    @app.route("/device-units/<int:unit_id>/actions", methods=["POST"])
    @manager_write_required
    def device_unit_action(unit_id):
        unit = DeviceUnit.query.filter_by(id=unit_id).with_for_update().first_or_404()
        device = unit.device
        movement_type = request.form.get("movement_type", "").strip()
        to_location_id = optional_int(request.form.get("to_location_id"))
        project_id = optional_int(request.form.get("project_id"))
        notes = request.form.get("notes", "").strip()
        error = validate_movement(
            device,
            movement_type,
            to_location_id,
            project_id,
            quantity=1,
            unit=unit,
        )
        if error:
            flash(error, "danger")
            return redirect(url_for("device_unit_detail", unit_id=unit.id))
        create_movement(
            device=device,
            unit=unit,
            movement_type=movement_type,
            quantity=1,
            from_location_id=unit.location_id,
            to_location_id=to_location_id,
            project_id=project_id,
            notes=notes,
            user_id=session["user_id"],
        )
        apply_device_state(
            device,
            movement_type,
            to_location_id,
            project_id,
            unit=unit,
        )
        db.session.commit()
        flash(f"Példánymozgás rögzítve: {movement_type_label(movement_type)}.", "success")
        return redirect(url_for("device_unit_detail", unit_id=unit.id))

    @app.route("/device-units/<int:unit_id>/edit", methods=["GET", "POST"])
    @manager_write_required
    def device_unit_edit(unit_id):
        unit = DeviceUnit.query.get_or_404(unit_id)
        if request.method == "POST":
            unit_code = request.form.get("unit_code", "").strip()
            asset_tag = request.form.get("asset_tag", "").strip() or None
            serial_number = request.form.get("serial_number", "").strip() or None
            if not unit_code:
                flash("A példányazonosító kötelező.", "danger")
            elif DeviceUnit.query.filter(DeviceUnit.id != unit.id, DeviceUnit.unit_code == unit_code).first():
                flash("Ezzel a példányazonosítóval már létezik másik példány.", "danger")
            elif asset_tag and Device.query.filter_by(asset_tag=asset_tag).first():
                flash("Ez az eszközazonosító már egy csoportos tételhez tartozik.", "danger")
            elif asset_tag and DeviceUnit.query.filter(
                DeviceUnit.id != unit.id, DeviceUnit.asset_tag == asset_tag
            ).first():
                flash("Ez az eszközazonosító már másik példányhoz tartozik.", "danger")
            else:
                unit.unit_code = unit_code
                unit.asset_tag = asset_tag
                unit.serial_number = serial_number
                unit.notes = request.form.get("notes", "").strip() or None
                db.session.commit()
                flash("Az eszközpéldány módosítva.", "success")
                return redirect(url_for("device_unit_detail", unit_id=unit.id))
        return render_template("device_unit_edit.html", unit=unit, device=unit.device)

    @app.route("/device-units/<int:unit_id>/qr")
    @login_required
    def device_unit_qr(unit_id):
        unit = DeviceUnit.query.get_or_404(unit_id)
        return qr_png_response(
            url_for("device_unit_detail", unit_id=unit.id, _external=True),
            f"{unit.unit_code}-qr.png",
        )

    @app.route("/device-units/<int:unit_id>/label")
    @login_required
    def device_unit_label(unit_id):
        unit = DeviceUnit.query.get_or_404(unit_id)
        return render_template("device_unit_label.html", unit=unit, device=unit.device)

    @app.route("/device-units/<int:unit_id>/archive", methods=["POST"])
    @manager_write_required
    def device_unit_archive(unit_id):
        unit = DeviceUnit.query.get_or_404(unit_id)
        blockers = device_unit_archive_blockers(unit)
        if blockers:
            flash_archive_blockers("Az eszközpéldány nem archiválható.", blockers)
            return redirect(url_for("device_unit_detail", unit_id=unit.id))
        unit.archived_at = now_utc()
        db.session.commit()
        flash("Az eszközpéldány archiválva.", "info")
        return redirect(url_for("device_units", device_id=unit.device_id))

    @app.route("/devices/<int:device_id>/archive", methods=["POST"])
    @manager_write_required
    def device_archive(device_id):
        device = Device.query.get_or_404(device_id)
        blockers = device_archive_blockers(device)
        if blockers:
            flash_archive_blockers("Az eszköztétel nem archiválható.", blockers)
            return redirect(url_for("device_detail", device_id=device.id))
        device.archived_at = now_utc()
        db.session.commit()
        flash("Az eszköz archiválva.", "info")
        return redirect(url_for("devices"))

    @app.route("/devices/<int:device_id>/actions", methods=["POST"])
    @manager_write_required
    def device_action(device_id):
        device = Device.query.get_or_404(device_id)
        source_balance_id = optional_int(request.form.get("source_balance_id"))
        source_balance = (
            db.session.get(BulkStockBalance, source_balance_id)
            if source_balance_id
            else None
        )
        movement_type = request.form.get("movement_type", "").strip()
        to_location_id = optional_int(request.form.get("to_location_id"))
        project_id = optional_int(request.form.get("project_id"))
        quantity = optional_float(request.form.get("quantity"))
        notes = request.form.get("notes", "").strip()
        from_location_id = source_balance.location_id if source_balance else None
        if movement_type not in MOVEMENT_TYPES:
            flash("Érvénytelen készletművelet.", "danger")
            return redirect(url_for("device_detail", device_id=device.id))
        error = validate_movement(
            device,
            movement_type,
            to_location_id,
            project_id,
            quantity=quantity,
            from_location_id=from_location_id,
            source_balance=source_balance,
        )
        if error:
            flash(error, "danger")
            return redirect(url_for("device_detail", device_id=device.id))
        try:
            create_movement(
                device=device,
                movement_type=movement_type,
                quantity=quantity,
                from_location_id=from_location_id,
                to_location_id=to_location_id,
                project_id=project_id,
                notes=notes,
                user_id=session["user_id"],
                source_balance=source_balance,
            )
            apply_device_state(
                device,
                movement_type,
                to_location_id,
                project_id,
                quantity=quantity,
                source_balance=source_balance,
            )
            db.session.commit()
        except ValueError as error:
            db.session.rollback()
            flash(str(error), "danger")
            return redirect(url_for("device_detail", device_id=device.id))
        flash(f"Készletművelet rögzítve: {movement_type_label(movement_type)}.", "success")
        return redirect(url_for("device_detail", device_id=device.id))

    @app.route("/finance")
    @finance_required
    def finance_overview():
        devices = (
            Device.query.filter(Device.archived_at.is_(None))
            .order_by(Device.product_name.asc(), Device.asset_tag.asc())
            .all()
        )
        projects = (
            Project.query.filter(Project.archived_at.is_(None))
            .order_by(Project.code.asc())
            .all()
        )
        locations = (
            Location.query.filter(
                Location.archived_at.is_(None),
                Location.location_type.in_(LOGISTIC_LOCATION_TYPES),
            )
            .order_by(Location.name.asc())
            .all()
        )
        invoice_items = UnassignedInvoiceItem.query.filter(
            UnassignedInvoiceItem.archived_at.is_(None)
        ).all()
        currency_totals = device_currency_totals(devices)
        project_rows = [project_finance_snapshot(project) for project in projects]
        project_rows.sort(
            key=lambda item: item["net_huf"] + item["invoice_value_huf"],
            reverse=True,
        )
        location_rows = []
        for location in locations:
            rows = [
                {
                    "device": unit.device,
                    "unit": unit,
                    "quantity": 1,
                    "status": unit.status,
                }
                for unit in location.device_units
                if unit.archived_at is None and unit.status != "SCRAPPED"
            ] + [
                {
                    "device": balance.device,
                    "unit": None,
                    "quantity": balance.quantity,
                    "status": balance.status,
                }
                for balance in location.bulk_balances
                if balance.quantity > 1e-9 and balance.status != "SCRAPPED"
            ]
            totals = finance_rows_totals(rows)
            location_rows.append(
                {
                    "location": location,
                    "quantity": sum(row["quantity"] for row in rows),
                    **totals,
                }
            )
        location_rows.sort(key=lambda item: item["net_huf"], reverse=True)
        inventory_totals = empty_currency_totals()
        for row in location_rows:
            for key in (
                "net_huf",
                "gross_huf",
                "net_eur",
                "gross_eur",
            ):
                inventory_totals[key] += row[key]
        inventory_totals["missing_currency_count"] = currency_totals[
            "missing_currency_count"
        ]
        supplier_map = {}
        for device in devices:
            supplier = (
                device.supplier_manufacturer
                or device.manufacturer
                or "Nincs megadva"
            )
            entry = supplier_map.setdefault(
                supplier,
                {
                    "name": supplier,
                    "device_count": 0,
                    "net_huf": 0,
                    "net_eur": 0,
                    "invoice_count": 0,
                    "unpaid_count": 0,
                },
            )
            entry["device_count"] += 1
            if device.currency in {"HUF", "EUR"} and device.total_net_price is not None:
                entry[f"net_{device.currency.lower()}"] += device.total_net_price
            if device.supplier_invoice_number:
                entry["invoice_count"] += 1
                if device.supplier_invoice_paid is not True:
                    entry["unpaid_count"] += 1
        supplier_rows = sorted(
            supplier_map.values(),
            key=lambda item: item["net_huf"],
            reverse=True,
        )
        unpaid_supplier_invoices = sum(
            1
            for device in devices
            if device.supplier_invoice_number
            and device.supplier_invoice_paid is not True
        )
        missing_financial_data = sum(
            1
            for device in devices
            if device_finance_issues(device)
        )
        return render_template(
            "finance_overview.html",
            top_projects=project_rows[:8],
            location_rows=location_rows[:8],
            supplier_rows=supplier_rows[:8],
            incomplete_devices=[
                {
                    "device": device,
                    "issues": device_finance_issues(device),
                }
                for device in devices
                if device_finance_issues(device)
            ][:10],
            chart_data={
                "projects": {
                    "labels": [row["project"].code for row in project_rows[:8]],
                    "values": [row["net_huf"] for row in project_rows[:8]],
                },
                "locations": {
                    "labels": [row["location"].name for row in location_rows[:8]],
                    "values": [row["net_huf"] for row in location_rows[:8]],
                },
                "suppliers": {
                    "labels": [row["name"] for row in supplier_rows[:8]],
                    "values": [row["net_huf"] for row in supplier_rows[:8]],
                },
            },
            summary={
                **inventory_totals,
                "unpaid_supplier_invoices": unpaid_supplier_invoices,
                "unassigned_invoice_count": sum(
                    1
                    for item in invoice_items
                    if item.assignment_status == "unassigned"
                ),
                "project_assigned_value_huf": sum(
                    (
                        item.line_gross_amount_huf
                        if item.line_gross_amount_huf is not None
                        else line_net_amount(item) or 0
                    )
                    for item in invoice_items
                    if item.assigned_project_id is not None
                ),
                "device_assigned_value_huf": sum(
                    (
                        item.line_gross_amount_huf
                        if item.line_gross_amount_huf is not None
                        else line_net_amount(item) or 0
                    )
                    for item in invoice_items
                    if item.assigned_device_id is not None
                ),
                "missing_financial_data": missing_financial_data,
            },
        )

    @app.route("/finance/projects")
    @finance_required
    def finance_projects():
        search = request.args.get("q", "").strip()
        status = request.args.get("status", "").strip()
        query = Project.query.filter(Project.archived_at.is_(None))
        if search:
            term = f"%{search}%"
            query = query.filter(
                or_(
                    Project.code.ilike(term),
                    Project.name.ilike(term),
                    Project.customer.ilike(term),
                )
            )
        if status in PROJECT_STATUS_LABELS:
            query = query.filter(Project.status == status)
        rows = [project_finance_snapshot(project) for project in query.order_by(Project.code.asc()).all()]
        rows.sort(key=lambda item: item["net_huf"] + item["invoice_value_huf"], reverse=True)
        return render_template(
            "finance_projects.html",
            rows=rows,
            search=search,
            selected_status=status,
            project_statuses=PROJECT_STATUS_LABELS,
        )

    @app.route("/finance/projects/<int:project_id>")
    @finance_required
    def finance_project_detail(project_id):
        project = Project.query.get_or_404(project_id)
        snapshot = project_finance_snapshot(project)
        bom_rows = project_bom_rows(project)
        invoice_items = (
            UnassignedInvoiceItem.query.filter(
                UnassignedInvoiceItem.archived_at.is_(None),
                UnassignedInvoiceItem.assigned_project_id == project.id,
            )
            .order_by(UnassignedInvoiceItem.invoice_date.desc())
            .all()
        )
        return render_template(
            "finance_project_detail.html",
            project=project,
            snapshot=snapshot,
            bom_rows=bom_rows,
            invoice_items=invoice_items,
        )

    @app.route("/finance/projects/<int:project_id>/bom")
    @finance_required
    def finance_project_bom(project_id):
        project = Project.query.get_or_404(project_id)
        rows = project_bom_rows(project)
        return render_template(
            "finance_bom.html",
            project=project,
            rows=rows,
            totals=finance_rows_totals(
                [
                    {
                        "device": row["device"],
                        "quantity": row["quantity"],
                    }
                    for row in rows
                ]
            ),
        )

    @app.route("/finance/inventory")
    @finance_required
    def finance_inventory():
        location_id = optional_int(request.args.get("location_id"))
        locations = (
            Location.query.filter(
                Location.archived_at.is_(None),
                Location.location_type.in_(LOGISTIC_LOCATION_TYPES),
            )
            .order_by(Location.name.asc())
            .all()
        )
        rows = []
        for location in locations:
            if location_id and location.id != location_id:
                continue
            item_rows = [
                {
                    "device": unit.device,
                    "quantity": 1,
                    "status": unit.status,
                    "unit": unit,
                }
                for unit in location.device_units
                if unit.archived_at is None and unit.status != "SCRAPPED"
            ] + [
                {
                    "device": balance.device,
                    "quantity": balance.quantity,
                    "status": balance.status,
                    "unit": None,
                }
                for balance in location.bulk_balances
                if balance.quantity > 1e-9 and balance.status != "SCRAPPED"
            ]
            rows.append(
                {
                    "location": location,
                    "detail_rows": item_rows,
                    "quantity": sum(item["quantity"] for item in item_rows),
                    **finance_rows_totals(item_rows),
                }
            )
        return render_template(
            "finance_inventory.html",
            rows=rows,
            locations=locations,
            selected_location_id=location_id,
        )

    @app.route("/finance/suppliers")
    @finance_required
    def finance_suppliers():
        search = request.args.get("q", "").strip().lower()
        suppliers = {}
        devices = Device.query.filter(Device.archived_at.is_(None)).all()
        for device in devices:
            supplier = (
                device.supplier_manufacturer
                or device.manufacturer
                or "Nincs megadva"
            )
            if search and search not in supplier.lower():
                continue
            entry = suppliers.setdefault(
                supplier,
                {
                    "name": supplier,
                    "devices": [],
                    "net_huf": 0,
                    "gross_huf": 0,
                    "net_eur": 0,
                    "gross_eur": 0,
                    "invoice_count": 0,
                    "unpaid_count": 0,
                },
            )
            entry["devices"].append(device)
            if device.currency in {"HUF", "EUR"}:
                currency = device.currency.lower()
                entry[f"net_{currency}"] += device.total_net_price or 0
                entry[f"gross_{currency}"] += device.total_gross_price or 0
            if device.supplier_invoice_number:
                entry["invoice_count"] += 1
                if device.supplier_invoice_paid is not True:
                    entry["unpaid_count"] += 1
        rows = sorted(suppliers.values(), key=lambda item: item["net_huf"], reverse=True)
        return render_template("finance_suppliers.html", rows=rows, search=search)

    @app.route("/finance/invoices")
    @finance_required
    def finance_invoices():
        search = request.args.get("q", "").strip()
        payment_status = request.args.get("payment_status", "all").strip()
        device_query = Device.query.filter(
            Device.archived_at.is_(None),
            or_(
                Device.supplier_invoice_number.is_not(None),
                Device.shipping_invoice_number.is_not(None),
            ),
        )
        item_query = UnassignedInvoiceItem.query.filter(
            UnassignedInvoiceItem.archived_at.is_(None)
        )
        if search:
            term = f"%{search}%"
            device_query = device_query.filter(
                or_(
                    Device.supplier_invoice_number.ilike(term),
                    Device.shipping_invoice_number.ilike(term),
                    Device.supplier_manufacturer.ilike(term),
                    Device.product_name.ilike(term),
                )
            )
            item_query = item_query.filter(
                or_(
                    UnassignedInvoiceItem.invoice_number.ilike(term),
                    UnassignedInvoiceItem.partner.ilike(term),
                    UnassignedInvoiceItem.description.ilike(term),
                )
            )
        device_invoices = device_query.order_by(Device.updated_at.desc()).all()
        if payment_status == "unpaid":
            device_invoices = [
                device
                for device in device_invoices
                if is_financially_open(device)
            ]
        elif payment_status == "paid":
            device_invoices = [
                device
                for device in device_invoices
                if not is_financially_open(device)
            ]
        invoice_items = item_query.order_by(UnassignedInvoiceItem.invoice_date.desc()).all()
        return render_template(
            "finance_invoices.html",
            device_invoices=device_invoices,
            invoice_items=invoice_items,
            search=search,
            payment_status=payment_status,
        )

    @app.route("/unassigned-invoices")
    @finance_required
    def unassigned_invoices():
        projects = Project.query.filter(Project.archived_at.is_(None)).order_by(Project.name.asc()).all()
        devices = (
            Device.query.filter(Device.archived_at.is_(None))
            .order_by(Device.device_type.asc(), Device.product_name.asc(), Device.asset_tag.asc())
            .all()
        )
        responsible_users = (
            User.query.filter(
                User.is_active.is_(True),
                User.role.in_(("admin", "manager")),
            )
            .order_by(User.username.asc())
            .all()
        )
        selected_assignment_status = request.args.get(
            "assignment_status", "unassigned"
        ).strip()
        if selected_assignment_status not in {
            "all",
            *ASSIGNMENT_STATUS_LABELS.keys(),
        }:
            selected_assignment_status = "unassigned"
        search = request.args.get("q", "").strip()

        invoice_query = UnassignedInvoiceItem.query.filter(
            UnassignedInvoiceItem.archived_at.is_(None)
        )
        if search:
            term = f"%{search}%"
            invoice_query = invoice_query.filter(
                or_(
                    UnassignedInvoiceItem.invoice_number.ilike(term),
                    UnassignedInvoiceItem.partner.ilike(term),
                    UnassignedInvoiceItem.description.ilike(term),
                )
            )
        if selected_assignment_status == "unassigned":
            invoice_query = invoice_query.filter(
                UnassignedInvoiceItem.assignment_status == "unassigned",
                UnassignedInvoiceItem.assigned_project_id.is_(None),
                UnassignedInvoiceItem.assigned_device_id.is_(None),
            )
        elif selected_assignment_status == "assigned":
            invoice_query = invoice_query.filter(
                or_(
                    UnassignedInvoiceItem.assignment_status == "assigned",
                    UnassignedInvoiceItem.assigned_project_id.is_not(None),
                    UnassignedInvoiceItem.assigned_device_id.is_not(None),
                )
            )
        elif selected_assignment_status in ASSIGNMENT_STATUS_LABELS:
            invoice_query = invoice_query.filter(
                UnassignedInvoiceItem.assignment_status
                == selected_assignment_status
            )
        invoice_items = invoice_query.order_by(
            UnassignedInvoiceItem.created_at.desc()
        ).all()
        return render_template(
            "unassigned_invoices.html",
            invoice_items=invoice_items,
            projects=projects,
            devices=devices,
            responsible_users=responsible_users,
            assignment_statuses=ASSIGNMENT_STATUS_LABELS,
            selected_assignment_status=selected_assignment_status,
            search=search,
        )

    @app.route("/unassigned-invoices/new", methods=["GET", "POST"])
    @finance_required
    def unassigned_invoice_new():
        projects = Project.query.filter(Project.archived_at.is_(None)).order_by(Project.name.asc()).all()
        devices = (
            Device.query.filter(Device.archived_at.is_(None))
            .order_by(Device.device_type.asc(), Device.product_name.asc(), Device.asset_tag.asc())
            .all()
        )
        responsible_users = (
            User.query.filter(
                User.is_active.is_(True),
                User.role.in_(("admin", "manager")),
            )
            .order_by(User.username.asc())
            .all()
        )
        item = UnassignedInvoiceItem()
        if request.method == "POST":
            update_unassigned_invoice_from_form(item, request.form)
            if not item.invoice_number and not item.description:
                flash("A számlaszám vagy a megnevezés megadása kötelező.", "danger")
            else:
                db.session.add(item)
                db.session.commit()
                flash("A manuális számlasor létrejött.", "success")
                return redirect(url_for("unassigned_invoices"))
        return render_template(
            "unassigned_invoice_form.html",
            item=item,
            projects=projects,
            devices=devices,
            responsible_users=responsible_users,
            assignment_statuses=ASSIGNMENT_STATUS_LABELS,
            form_title="Manuális számlasor hozzáadása",
            submit_label="Számlasor mentése",
        )

    @app.route("/unassigned-invoices/<int:item_id>/clarify", methods=["POST"])
    @finance_required
    def unassigned_invoice_clarify(item_id):
        item = UnassignedInvoiceItem.query.get_or_404(item_id)
        assigned_project_id = optional_int(request.form.get("assigned_project_id"))
        assigned_device_id = optional_int(request.form.get("assigned_device_id"))
        responsible_user_id = optional_int(request.form.get("responsible_user_id"))
        assignment_status = request.form.get("assignment_status", "unassigned").strip()
        if assignment_status not in ASSIGNMENT_STATUS_LABELS:
            flash("Érvénytelen hozzárendelési státusz.", "danger")
            return redirect(url_for("unassigned_invoices"))
        if assigned_project_id and db.session.get(Project, assigned_project_id) is None:
            flash("A kiválasztott projekt nem található.", "danger")
            return redirect(url_for("unassigned_invoices"))
        if assigned_device_id and db.session.get(Device, assigned_device_id) is None:
            flash("A kiválasztott eszköz nem található.", "danger")
            return redirect(url_for("unassigned_invoices"))
        responsible = (
            db.session.get(User, responsible_user_id)
            if responsible_user_id
            else None
        )
        if responsible and (
            not responsible.is_active
            or responsible.effective_role not in {"admin", "manager"}
        ):
            flash("A kiválasztott felelős nem jogosult pénzügyi tisztázásra.", "danger")
            return redirect(url_for("unassigned_invoices"))
        if assigned_project_id or assigned_device_id:
            assignment_status = "assigned"
        item.assigned_project_id = assigned_project_id
        item.assigned_device_id = assigned_device_id
        item.responsible_user_id = responsible_user_id
        item.assignment_status = assignment_status
        item.notes = request.form.get("notes", "").strip() or None
        db.session.commit()
        flash("A számlasor tisztázási adatai frissültek.", "success")
        return redirect(url_for("unassigned_invoices"))

    @app.route("/unassigned-invoices/<int:item_id>/edit", methods=["GET", "POST"])
    @finance_required
    def unassigned_invoice_edit(item_id):
        item = UnassignedInvoiceItem.query.get_or_404(item_id)
        projects = Project.query.filter(Project.archived_at.is_(None)).order_by(Project.name.asc()).all()
        devices = (
            Device.query.filter(Device.archived_at.is_(None))
            .order_by(Device.device_type.asc(), Device.product_name.asc(), Device.asset_tag.asc())
            .all()
        )
        responsible_users = (
            User.query.filter(
                User.is_active.is_(True),
                User.role.in_(("admin", "manager")),
            )
            .order_by(User.username.asc())
            .all()
        )
        if request.method == "POST":
            update_unassigned_invoice_from_form(item, request.form)
            db.session.commit()
            flash("A számlasor módosítva.", "success")
            return redirect(url_for("unassigned_invoices"))
        return render_template(
            "unassigned_invoice_edit.html",
            item=item,
            projects=projects,
            devices=devices,
            responsible_users=responsible_users,
            assignment_statuses=ASSIGNMENT_STATUS_LABELS,
        )

    @app.route("/unassigned-invoices/<int:item_id>/archive", methods=["POST"])
    @finance_required
    def unassigned_invoice_archive(item_id):
        item = UnassignedInvoiceItem.query.get_or_404(item_id)
        item.archived_at = now_utc()
        db.session.commit()
        flash("A számlasor archiválva.", "info")
        return redirect(url_for("unassigned_invoices"))

    @app.route("/import", methods=["GET", "POST"])
    @app.route("/legacy/parkl-excel-import", methods=["GET", "POST"])
    @admin_required
    def legacy_parkl_excel_import():
        pending_import = session.get("pending_import")
        preview = None

        if request.method == "POST":
            action = request.form.get("action", "dry_run")
            if action == "confirm":
                if request.form.get("execute_import") != "on":
                    flash("Az importálás végrehajtásához jelöld be a megerősítést.", "danger")
                    return redirect(url_for("legacy_parkl_excel_import"))
                if not pending_import or not os.path.exists(pending_import["path"]):
                    flash("Nincs érvényes előnézeti import. Töltsd fel újra a fájlt.", "danger")
                    return redirect(url_for("legacy_parkl_excel_import"))

                summary = parse_inventory_workbook(pending_import["path"])
                batch = ImportBatch(
                    filename=pending_import["filename"],
                    imported_by_user_id=session.get("user_id"),
                    dry_run_summary_json=json.dumps(
                        summary, ensure_ascii=False, default=str
                    ),
                    warning_count=summary["warning_count"],
                    status="running",
                )
                db.session.add(batch)
                db.session.flush()
                result = import_parsed_workbook(summary, batch.id, session["user_id"])
                batch.created_count = result["created_count"]
                batch.skipped_count = result["skipped_count"]
                batch.updated_count = result["updated_count"]
                batch.status = "completed"
                db.session.commit()
                session.pop("pending_import", None)
                flash(
                    "Importálás kész: "
                    f"{batch.created_count} létrehozva, "
                    f"{batch.skipped_count} kihagyva, "
                    f"{batch.updated_count} frissítve.",
                    "success",
                )
                return redirect(url_for("legacy_parkl_excel_import", batch_id=batch.id))

            upload = request.files.get("excel_file")
            if not upload or upload.filename == "":
                flash("Válassz ki egy .xlsx fájlt.", "danger")
                return redirect(url_for("legacy_parkl_excel_import"))
            if not upload.filename.lower().endswith(".xlsx"):
                flash("Csak .xlsx fájl tölthető fel.", "danger")
                return redirect(url_for("legacy_parkl_excel_import"))

            upload_dir = os.path.join(app.instance_path, UPLOAD_SUBDIR)
            os.makedirs(upload_dir, exist_ok=True)
            safe_name = secure_filename(upload.filename)
            filename = f"{uuid4().hex}_{safe_name}"
            upload_path = os.path.join(upload_dir, filename)
            upload.save(upload_path)
            preview = parse_inventory_workbook(upload_path)
            session["pending_import"] = {
                "path": upload_path,
                "filename": safe_name,
            }
            flash("Előnézet elkészült. Az adatbázis még nem módosult.", "info")

        batch = None
        batch_id = optional_int(request.args.get("batch_id"))
        if batch_id:
            batch = db.session.get(ImportBatch, batch_id)
        if preview is None and pending_import and os.path.exists(pending_import["path"]):
            preview = parse_inventory_workbook(pending_import["path"])

        recent_batches = (
            ImportBatch.query.filter(ImportBatch.archived_at.is_(None))
            .order_by(ImportBatch.created_at.desc())
            .limit(8)
            .all()
        )
        return render_template(
            "import.html",
            preview=preview,
            batch=batch,
            pending_import=session.get("pending_import"),
            recent_batches=recent_batches,
        )

    @app.route("/import-batches/<int:batch_id>")
    @app.route("/legacy/import-batches/<int:batch_id>")
    @admin_required
    def import_batch_detail(batch_id):
        batch = ImportBatch.query.get_or_404(batch_id)
        devices = (
            Device.query.filter_by(import_batch_id=batch.id)
            .order_by(Device.source_row_number.asc())
            .all()
        )
        invoice_items = (
            UnassignedInvoiceItem.query.filter_by(import_batch_id=batch.id)
            .order_by(UnassignedInvoiceItem.source_row_number.asc())
            .all()
        )
        warnings = []
        if batch.dry_run_summary_json:
            try:
                summary = json.loads(batch.dry_run_summary_json)
                warnings = summary.get("warnings", [])
                for sheet in summary.get("sheets", []):
                    warnings.extend(sheet.get("warnings", []))
            except json.JSONDecodeError:
                warnings = []
        return render_template(
            "import_batch_detail.html",
            batch=batch,
            devices=devices,
            invoice_items=invoice_items,
            warnings=warnings,
        )

    @app.route("/import-batches/<int:batch_id>/rollback", methods=["POST"])
    @app.route("/legacy/import-batches/<int:batch_id>/rollback", methods=["POST"])
    @admin_required
    def import_batch_rollback(batch_id):
        batch = ImportBatch.query.get_or_404(batch_id)
        archived_devices = 0
        unsafe_devices = 0
        archived_invoice_items = 0
        for device in batch.devices:
            if device.archived_at:
                continue
            if len(device.movements) <= 1:
                device.archived_at = now_utc()
                archived_devices += 1
            else:
                unsafe_devices += 1
        for item in batch.unassigned_invoice_items:
            if not item.archived_at:
                item.archived_at = now_utc()
                archived_invoice_items += 1
        batch.status = "rolled_back" if unsafe_devices == 0 else "partial_rollback"
        batch.archived_at = now_utc()
        db.session.commit()
        if unsafe_devices:
            flash(
                "Import részben visszavonva. "
                f"{archived_devices} eszköz és {archived_invoice_items} számlasor archiválva, "
                f"{unsafe_devices} eszköz kézi mozgások miatt megmaradt.",
                "warning",
            )
        else:
            flash(
                f"Import visszavonva: {archived_devices} eszköz és "
                f"{archived_invoice_items} számlasor archiválva.",
                "success",
            )
        return redirect(url_for("import_batch_detail", batch_id=batch.id))

    @app.route("/import-batches/<int:batch_id>/archive", methods=["POST"])
    @app.route("/legacy/import-batches/<int:batch_id>/archive", methods=["POST"])
    @admin_required
    def import_batch_archive(batch_id):
        batch = ImportBatch.query.get_or_404(batch_id)
        batch.archived_at = now_utc()
        db.session.commit()
        flash("Az importcsomag archiválva.", "info")
        return redirect(url_for("legacy_parkl_excel_import"))

    @app.route("/locations")
    @write_required
    def locations():
        selected_type = request.args.get("location_type", "").strip()
        location_query = Location.query.filter(
            Location.archived_at.is_(None),
            Location.location_type.in_(LOGISTIC_LOCATION_TYPES),
        )
        if selected_type in LOCATION_TYPE_LABELS:
            location_query = location_query.filter(Location.location_type == selected_type)
        location_list = location_query.order_by(Location.name.asc()).all()
        return render_template(
            "locations.html",
            locations=location_list,
            location_inventory={
                location.id: location_inventory_summary(location)
                for location in location_list
            },
            location_types=LOCATION_TYPE_LABELS,
            selected_type=selected_type,
        )

    @app.route("/locations/new", methods=["GET", "POST"])
    @manager_write_required
    def location_new():
        if request.method == "POST":
            name = request.form.get("name", "").strip()
            location_type = request.form.get("location_type", "warehouse").strip()
            address = request.form.get("address", "").strip()
            notes = request.form.get("notes", "").strip()
            if not name:
                flash("A készlethely neve kötelező.", "danger")
            elif location_type not in LOGISTIC_LOCATION_TYPES:
                flash("Csak logisztikai készlethelytípus választható.", "danger")
            else:
                location = Location(
                    name=name,
                    location_type=location_type,
                    address=address,
                    notes=notes,
                )
                db.session.add(location)
                db.session.commit()
                flash("A készlethely létrejött.", "success")
                return redirect(url_for("location_detail", location_id=location.id))
        return render_template(
            "location_form.html",
            location=None,
            location_types=LOCATION_TYPE_LABELS,
            form_title="Új készlethely",
            submit_label="Készlethely létrehozása",
            cancel_url=url_for("locations"),
        )

    @app.route("/locations/<int:location_id>")
    @login_required
    def location_detail(location_id):
        location = Location.query.get_or_404(location_id)
        devices = (
            Device.query.join(BulkStockBalance)
            .filter(
                BulkStockBalance.location_id == location.id,
                BulkStockBalance.quantity > 0,
                BulkStockBalance.status.in_(PHYSICAL_LOCATION_STATUSES),
                Device.archived_at.is_(None),
                Device.tracking_mode == "bulk",
            )
            .distinct()
            .order_by(Device.asset_tag.asc())
            .all()
        )
        bulk_balances = (
            BulkStockBalance.query.filter(
                BulkStockBalance.location_id == location.id,
                BulkStockBalance.quantity > 0,
                BulkStockBalance.status.in_(PHYSICAL_LOCATION_STATUSES),
            )
            .join(Device)
            .filter(Device.archived_at.is_(None), Device.tracking_mode == "bulk")
            .order_by(BulkStockBalance.id.asc())
            .all()
        )
        units = (
            DeviceUnit.query.filter_by(location_id=location.id)
            .filter(
                DeviceUnit.archived_at.is_(None),
                DeviceUnit.status.in_(PHYSICAL_LOCATION_STATUSES),
            )
            .join(Device)
            .filter(Device.archived_at.is_(None), Device.tracking_mode == "unit")
            .order_by(DeviceUnit.unit_code.asc())
            .all()
        )
        movements = (
            StockMovement.query.filter(
                or_(
                    StockMovement.from_location_id == location.id,
                    StockMovement.to_location_id == location.id,
                )
            )
            .order_by(StockMovement.created_at.desc())
            .limit(50)
            .all()
        )
        free_bulk_balances = [
            balance for balance in bulk_balances if balance.status in FREE_STOCK_STATUSES
        ]
        reserved_bulk_balances = [
            balance for balance in bulk_balances if balance.status == "RESERVED"
        ]
        service_bulk_balances = [
            balance for balance in bulk_balances if balance.status == "IN_SERVICE"
        ]
        free_units = [unit for unit in units if unit.status in FREE_STOCK_STATUSES]
        reserved_units = [unit for unit in units if unit.status == "RESERVED"]
        service_units = [unit for unit in units if unit.status == "IN_SERVICE"]
        location_inventory_rows = [
            {
                "device": balance.device,
                "unit": None,
                "quantity": balance.quantity,
                "status": balance.status,
                "location": location,
            }
            for balance in bulk_balances
        ] + [
            {
                "device": unit.device,
                "unit": unit,
                "quantity": 1,
                "status": unit.status,
                "location": location,
            }
            for unit in units
        ]
        return render_template(
            "location_detail.html",
            location=location,
            devices=devices,
            units=units,
            bulk_balances=bulk_balances,
            free_bulk_balances=free_bulk_balances,
            reserved_bulk_balances=reserved_bulk_balances,
            service_bulk_balances=service_bulk_balances,
            free_units=free_units,
            reserved_units=reserved_units,
            service_units=service_units,
            movements=movements,
            location_summary={
                "quantity": sum(balance.quantity for balance in bulk_balances) + len(units),
                "free_quantity": sum(
                    balance.quantity for balance in free_bulk_balances
                )
                + len(free_units),
                "reserved_quantity": sum(
                    balance.quantity for balance in reserved_bulk_balances
                )
                + len(reserved_units),
                "service_quantity": sum(
                    balance.quantity for balance in service_bulk_balances
                )
                + len(service_units),
                **inventory_rows_currency_totals(location_inventory_rows),
            },
            archive_blockers=location_archive_blockers(location),
        )

    @app.route("/locations/<int:location_id>/edit", methods=["GET", "POST"])
    @manager_write_required
    def location_edit(location_id):
        location = Location.query.get_or_404(location_id)
        if request.method == "POST":
            name = request.form.get("name", "").strip()
            location_type = request.form.get("location_type", "warehouse").strip()
            if not name:
                flash("A készlethely neve kötelező.", "danger")
            elif location_type not in LOGISTIC_LOCATION_TYPES:
                flash("Csak logisztikai készlethelytípus választható.", "danger")
            else:
                location.name = name
                location.location_type = location_type
                location.address = request.form.get("address", "").strip()
                location.notes = request.form.get("notes", "").strip()
                db.session.commit()
                flash("A készlethely módosítva.", "success")
                return redirect(url_for("location_detail", location_id=location.id))
        return render_template(
            "location_form.html",
            location=location,
            location_types=LOCATION_TYPE_LABELS,
            form_title="Készlethely szerkesztése",
            submit_label="Mentés",
            cancel_url=url_for("location_detail", location_id=location.id),
        )

    @app.route("/locations/<int:location_id>/qr")
    @login_required
    def location_qr(location_id):
        location = Location.query.get_or_404(location_id)
        return qr_png_response(
            url_for("location_detail", location_id=location.id, _external=True),
            f"keszlethely-{location.id}-qr.png",
        )

    @app.route("/locations/<int:location_id>/archive", methods=["POST"])
    @manager_write_required
    def location_archive(location_id):
        location = Location.query.get_or_404(location_id)
        blockers = location_archive_blockers(location)
        if blockers:
            flash_archive_blockers("A készlethely nem archiválható.", blockers)
            return redirect(url_for("location_detail", location_id=location.id))
        location.archived_at = now_utc()
        db.session.commit()
        flash("A készlethely archiválva.", "info")
        return redirect(url_for("locations"))

    @app.route("/movements", methods=["GET", "POST"])
    @write_required
    def movements():
        search = request.args.get("q", "").strip()
        date_from = optional_date(request.args.get("date_from"))
        date_to = optional_date(request.args.get("date_to"))
        selected_device_id = optional_int(request.args.get("device_id"))
        selected_unit_id = optional_int(request.args.get("unit_id"))
        selected_project_id = optional_int(request.args.get("project_id"))
        selected_location_id = optional_int(request.args.get("location_id"))
        selected_movement_type = request.args.get("movement_type", "").strip()
        selected_user_id = optional_int(request.args.get("user_id"))
        group_by = request.args.get("group_by", "").strip()
        if group_by not in {"", "date", "device", "project"}:
            group_by = ""
        devices = (
            Device.query.filter(Device.archived_at.is_(None))
            .order_by(Device.device_type.asc(), Device.product_name.asc(), Device.asset_tag.asc())
            .all()
        )
        locations = active_logistic_locations()
        projects = Project.query.filter(Project.archived_at.is_(None)).order_by(Project.name.asc()).all()
        units = (
            DeviceUnit.query.filter(DeviceUnit.archived_at.is_(None))
            .join(Device)
            .filter(Device.archived_at.is_(None))
            .order_by(DeviceUnit.unit_code.asc())
            .all()
        )
        bulk_balances = (
            BulkStockBalance.query.filter(BulkStockBalance.quantity > 0)
            .join(Device)
            .filter(Device.archived_at.is_(None), Device.tracking_mode == "bulk")
            .order_by(BulkStockBalance.device_id.asc(), BulkStockBalance.id.asc())
            .all()
        )
        movement_users = User.query.order_by(User.username.asc()).all()

        if request.method == "POST":
            device_id = optional_int(request.form.get("device_id"))
            unit_id = optional_int(request.form.get("unit_id"))
            source_balance_id = optional_int(request.form.get("source_balance_id"))
            quantity = optional_float(request.form.get("quantity"))
            movement_type = request.form.get("movement_type", "").strip()
            from_location_id = optional_int(request.form.get("from_location_id"))
            to_location_id = optional_int(request.form.get("to_location_id"))
            project_id = optional_int(request.form.get("project_id"))
            notes = request.form.get("notes", "").strip()

            device = db.session.get(Device, device_id) if device_id else None
            unit = (
                DeviceUnit.query.filter_by(id=unit_id).with_for_update().one_or_none()
                if unit_id
                else None
            )
            source_balance = (
                db.session.get(BulkStockBalance, source_balance_id)
                if source_balance_id
                else None
            )
            if device is None or not movement_type:
                flash("Az eszköz és a mozgástípus kötelező.", "danger")
            else:
                error = validate_movement(
                    device,
                    movement_type,
                    to_location_id,
                    project_id,
                    quantity=quantity,
                    unit=unit,
                    from_location_id=from_location_id,
                    source_balance=source_balance,
                )
                if error:
                    flash(error, "danger")
                else:
                    try:
                        create_movement(
                            device=device,
                            movement_type=movement_type,
                            quantity=quantity,
                            unit=unit,
                            from_location_id=from_location_id,
                            to_location_id=to_location_id,
                            project_id=project_id,
                            notes=notes,
                            user_id=session["user_id"],
                            source_balance=source_balance,
                        )
                        apply_device_state(
                            device,
                            movement_type,
                            to_location_id,
                            project_id,
                            unit=unit,
                            quantity=quantity,
                            source_balance=source_balance,
                        )
                        db.session.commit()
                    except ValueError as error:
                        db.session.rollback()
                        flash(str(error), "danger")
                        return redirect(url_for("movements"))
                    flash("A készletmozgás rögzítve.", "success")
                    return redirect(url_for("movements"))

        movement_query = StockMovement.query
        if search:
            term = f"%{search}%"
            movement_query = movement_query.filter(
                or_(
                    StockMovement.notes.ilike(term),
                    StockMovement.device.has(
                        or_(
                            Device.asset_tag.ilike(term),
                            Device.product_name.ilike(term),
                            Device.model.ilike(term),
                        )
                    ),
                    StockMovement.unit.has(
                        or_(
                            DeviceUnit.unit_code.ilike(term),
                            DeviceUnit.asset_tag.ilike(term),
                            DeviceUnit.serial_number.ilike(term),
                        )
                    ),
                )
            )
        if date_from:
            movement_query = movement_query.filter(
                StockMovement.created_at >= datetime.combine(
                    date_from, datetime.min.time(), tzinfo=timezone.utc
                )
            )
        if date_to:
            movement_query = movement_query.filter(
                StockMovement.created_at < datetime.combine(
                    date_to + timedelta(days=1),
                    datetime.min.time(),
                    tzinfo=timezone.utc,
                )
            )
        if selected_device_id:
            movement_query = movement_query.filter(
                StockMovement.device_id == selected_device_id
            )
        if selected_unit_id:
            movement_query = movement_query.filter(
                StockMovement.unit_id == selected_unit_id
            )
        if selected_project_id:
            movement_query = movement_query.filter(
                or_(
                    StockMovement.project_id == selected_project_id,
                    StockMovement.from_project_id == selected_project_id,
                    StockMovement.to_project_id == selected_project_id,
                )
            )
        if selected_location_id:
            movement_query = movement_query.filter(
                or_(
                    StockMovement.from_location_id == selected_location_id,
                    StockMovement.to_location_id == selected_location_id,
                )
            )
        if selected_movement_type in MOVEMENT_TYPE_LABELS:
            movement_query = movement_query.filter(
                StockMovement.movement_type == selected_movement_type
            )
        if selected_user_id:
            movement_query = movement_query.filter(
                StockMovement.created_by_id == selected_user_id
            )

        movement_list = movement_query.order_by(StockMovement.created_at.desc()).all()
        movement_groups = []
        if group_by:
            grouped = {}
            for movement in movement_list:
                if group_by == "date":
                    key = movement.created_at.date().isoformat()
                    label = movement.created_at.strftime("%Y. %m. %d.")
                elif group_by == "device":
                    key = movement.device_id
                    label = device_primary_label(movement.device)
                else:
                    project = movement.to_project or movement.from_project or movement.project
                    key = project.id if project else "none"
                    label = (
                        f"{project.code} - {project.name}"
                        if project
                        else "Projekt nélküli mozgások"
                    )
                grouped.setdefault(key, {"label": label, "movements": []})[
                    "movements"
                ].append(movement)
            movement_groups = list(grouped.values())
        else:
            movement_groups = [{"label": None, "movements": movement_list}]
        return render_template(
            "movements.html",
            movements=movement_list,
            movement_groups=movement_groups,
            devices=devices,
            locations=locations,
            projects=projects,
            units=units,
            bulk_balances=bulk_balances,
            movement_users=movement_users,
            movement_types=MOVEMENT_TYPES,
            movement_filter_types=MOVEMENT_TYPE_LABELS,
            search=search,
            date_from=date_from,
            date_to=date_to,
            selected_device_id=selected_device_id,
            selected_unit_id=selected_unit_id,
            selected_project_id=selected_project_id,
            selected_location_id=selected_location_id,
            selected_movement_type=selected_movement_type,
            selected_user_id=selected_user_id,
            group_by=group_by,
            reversed_movement_ids=reversed_movement_ids(movement_list),
            reversible_movement_ids=reversible_movement_ids(movement_list),
        )

    @app.route("/movements/<int:movement_id>/reverse", methods=["POST"])
    @manager_write_required
    def movement_reverse(movement_id):
        movement = StockMovement.query.get_or_404(movement_id)
        try:
            reversal = reverse_stock_movement(movement, session["user_id"])
            db.session.commit()
        except ValueError as error:
            db.session.rollback()
            flash(str(error), "danger")
        else:
            flash(
                f"A(z) #{movement.id} készletmozgás ellenmozgással visszavonva "
                f"(új mozgás: #{reversal.id}).",
                "success",
            )

        project_id = movement.to_project_id or movement.from_project_id or movement.project_id
        if request.form.get("return_to") == "project" and project_id:
            return redirect(url_for("project_detail", project_id=project_id))
        if movement.unit_id:
            return redirect(url_for("device_unit_detail", unit_id=movement.unit_id))
        return redirect(url_for("device_detail", device_id=movement.device_id))

    return app


def reset_demo_dataset(
    app,
    User,
    Project,
    Location,
    Device,
    DeviceUnit,
    BulkStockBalance,
    StockMovement,
    UnassignedInvoiceItem,
    ImportBatch,
    ProjectDrawing,
    WorkOrder,
    WorkOrderMaterial,
    WorkOrderMeasurement,
    WorkOrderPhoto,
    WorkOrderTemplate,
):
    username = app.config["ADMIN_USERNAME"]
    password = app.config["ADMIN_PASSWORD"]
    user = User.query.filter_by(username=username).first()
    if user is None:
        if not password:
            raise click.ClickException(
                "Nincs admin felhasználó és ADMIN_PASSWORD sincs beállítva. "
                "Futtasd előbb: flask --app app seed-admin --password"
            )
        user = User(
            username=username,
            password_hash=generate_password_hash(password),
            is_admin=True,
            role="admin",
            is_active=True,
            force_password_change=True,
        )
        db.session.add(user)
        db.session.flush()

    ProjectDrawing.query.delete()
    WorkOrderPhoto.query.delete()
    WorkOrderMeasurement.query.delete()
    WorkOrderMaterial.query.delete()
    WorkOrder.query.delete()
    WorkOrderTemplate.query.delete()
    UnassignedInvoiceItem.query.delete()
    StockMovement.query.delete()
    DeviceUnit.query.delete()
    BulkStockBalance.query.delete()
    Device.query.delete()
    ImportBatch.query.delete()
    Project.query.delete()
    Location.query.delete()
    db.session.flush()
    seed_work_order_templates(WorkOrderTemplate)

    projects = {
        "PRK-001": Project(
            code="PRK-001",
            name="Arena EV Upgrade",
            customer="Arena",
            site_name="Arena helyszín",
            address="Stefánia út 2.",
            city="Budapest",
            country="Magyarország",
            site_notes="Demó telepítési helyszín.",
            status="active",
            notes="Demó EV-töltő bővítési projekt.",
        ),
        "PRK-002": Project(
            code="PRK-002",
            name="Office Park Sorompó projekt",
            customer="Office Park",
            site_name="Office Park helyszín",
            address="Váci út 99.",
            city="Budapest",
            country="Magyarország",
            site_notes="Demó sorompótelepítési helyszín.",
            status="active",
            notes="Demó sorompó és beléptetési projekt.",
        ),
    }
    db.session.add_all(projects.values())

    locations = {
        "warehouse": Location(name="Fő raktár", location_type="warehouse"),
        "service_car": Location(name="Szervizautó 1", location_type="service_vehicle"),
        "service": Location(name="Szerviz / javítás", location_type="service"),
    }
    db.session.add_all(locations.values())
    db.session.flush()

    devices = {
        "EV-BATCH": Device(
            asset_tag="EV-BATCH",
            device_type="EV charger",
            product_name="Schneider EVlink Pro AC",
            manufacturer="Schneider",
            model="EVlink Pro AC",
            quantity=3,
            currency="HUF",
            unit_net_price=350000,
            tracking_mode="unit",
            qr_mode="individual",
        ),
        "MAT-001": Device(
            asset_tag="MAT-001",
            device_type="Sticker",
            product_name="Matrica csomag",
            manufacturer="Parkl",
            model="Matrica csomag",
            quantity=50,
            currency="HUF",
            unit_net_price=500,
            tracking_mode="bulk",
        ),
    }
    db.session.add_all(devices.values())
    db.session.flush()

    warehouse_id = locations["warehouse"].id
    sticker = devices["MAT-001"]
    create_movement(
        device=sticker,
        movement_type="INBOUND",
        quantity=50,
        to_location_id=warehouse_id,
        notes="Demo kezdő bevételezés: 50 db matrica.",
        user_id=user.id,
    )
    apply_device_state(sticker, "INBOUND", warehouse_id, None, quantity=50)

    charger = devices["EV-BATCH"]
    charger.location_id = None
    charger.project_id = None
    units = []
    for number in range(1, 4):
        unit = DeviceUnit(
            device=charger,
            unit_code=f"SCH-EV-{number:03d}",
            asset_tag=f"SCH-EV-{number:03d}",
            status="IN_STOCK",
            location_id=warehouse_id,
        )
        db.session.add(unit)
        db.session.flush()
        create_movement(
            device=charger,
            unit=unit,
            movement_type="INBOUND",
            quantity=1,
            to_location_id=warehouse_id,
            notes="Demo egyedi példány kezdő bevételezése.",
            user_id=user.id,
        )
        units.append(unit)

    reserved_unit = units[0]
    create_movement(
        device=charger,
        unit=reserved_unit,
        movement_type="RESERVE",
        quantity=1,
        project_id=projects["PRK-001"].id,
        notes="Demo: egy töltő előfoglalása PRK-001 projektre.",
        user_id=user.id,
    )
    apply_device_state(
        charger,
        "RESERVE",
        None,
        projects["PRK-001"].id,
        unit=reserved_unit,
    )

    sticker_source = infer_bulk_source_balance(sticker, "RESERVE", 20)
    create_movement(
        device=sticker,
        movement_type="RESERVE",
        quantity=20,
        project_id=projects["PRK-001"].id,
        source_balance=sticker_source,
        notes="Demo: 20 db matrica előfoglalása PRK-001 projektre.",
        user_id=user.id,
    )
    apply_device_state(
        sticker,
        "RESERVE",
        None,
        projects["PRK-001"].id,
        quantity=20,
        source_balance=sticker_source,
    )

    invoice_items = [
        UnassignedInvoiceItem(
            invoice_number="TEL-2026-001",
            partner="Teltonika",
            invoice_date=date.today(),
            payment_deadline=date.today(),
            gross_amount_huf=113030,
            currency="HUF",
            description="Teltonika RUTX11 router számla",
            quantity=1,
            unit_price_huf=89000,
            net_amount_huf=89000,
            vat_amount_huf=24030,
            line_gross_amount_huf=113030,
            assignment_status="unassigned",
            notes="Demo nyitott, még nem hozzárendelt számlasor.",
        ),
        UnassignedInvoiceItem(
            invoice_number="SERV-2026-014",
            partner="Kamera Szerviz Kft.",
            invoice_date=date.today(),
            payment_deadline=date.today(),
            gross_amount_huf=63500,
            currency="HUF",
            description="Hikvision ANPR kamera szervizdíj",
            quantity=1,
            unit_price_huf=50000,
            net_amount_huf=50000,
            vat_amount_huf=13500,
            line_gross_amount_huf=63500,
            assignment_status="unassigned",
            notes="Demo nyitott szerviz számlasor.",
        ),
    ]
    db.session.add_all(invoice_items)
    db.session.commit()

    return {
        "projects": len(projects),
        "locations": len(locations),
        "devices": len(devices),
        "device_units": len(units),
        "bulk_physical": 50,
        "bulk_reserved": 20,
        "unit_reserved": 1,
        "movements": StockMovement.query.count(),
        "invoice_items": len(invoice_items),
    }


def allowed_drawing_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_DRAWING_EXTENSIONS


def save_drawing_background(app, upload, project_id):
    upload_dir = os.path.join(app.instance_path, DRAWING_UPLOAD_SUBDIR)
    os.makedirs(upload_dir, exist_ok=True)
    extension = upload.filename.rsplit(".", 1)[1].lower()
    base_name = secure_filename(upload.filename.rsplit(".", 1)[0]) or "alaprajz"
    filename = f"project-{project_id}-{uuid4().hex}-{base_name}.{extension}"
    upload.save(os.path.join(upload_dir, filename))
    return filename


def optional_int(value):
    if not value:
        return None
    try:
        return int(value)
    except ValueError:
        return None


def optional_float(value):
    if value in (None, ""):
        return None
    normalized = str(value).strip().replace(" ", "").replace(",", ".")
    try:
        return float(normalized)
    except ValueError:
        return None


def optional_decimal(value):
    if value in (None, ""):
        return None
    normalized = str(value).strip().replace(" ", "").replace(",", ".")
    try:
        return Decimal(normalized)
    except (InvalidOperation, ValueError):
        return None


def optional_date(value):
    if not value:
        return None
    try:
        return date.fromisoformat(value)
    except ValueError:
        return None


def m2m_teltonika_devices(Device):
    return (
        Device.query.filter(Device.archived_at.is_(None))
        .order_by(Device.product_name.asc(), Device.asset_tag.asc())
        .all()
    )


def update_m2m_subscription_from_form(subscription, form, Device):
    status = form.get("status", "active").strip()
    if status not in M2M_STATUS_LABELS:
        return "Érvénytelen előfizetés-státusz."
    phone_number = form.get("phone_number", "").strip() or None
    sim_number = form.get("sim_number", "").strip() or None
    device_number = form.get("device_number", "").strip() or None
    if not any((phone_number, sim_number, device_number)):
        return "Legalább a hívószám, a SIM-szám vagy az eszközszám megadása kötelező."
    teltonika_device_id = optional_int(form.get("teltonika_device_id"))
    if teltonika_device_id and db.session.get(Device, teltonika_device_id) is None:
        return "A kiválasztott ERP-eszköz nem található."
    monthly_fee = optional_decimal(form.get("current_monthly_fee"))
    if monthly_fee is not None and monthly_fee < 0:
        return "A havidíj nem lehet negatív."

    subscription.subscriber_name = form.get("subscriber_name", "").strip() or None
    subscription.account_number = form.get("account_number", "").strip() or None
    subscription.contract_number = form.get("contract_number", "").strip() or None
    subscription.registration_date = optional_date(form.get("registration_date"))
    subscription.phone_number = phone_number
    subscription.device_number = device_number
    subscription.location_name = form.get("location_name", "").strip() or None
    subscription.device_identifier = (
        form.get("device_identifier", "").strip() or None
    )
    subscription.sim_number = sim_number
    subscription.tariff_name = form.get("tariff_name", "").strip() or None
    subscription.current_package = form.get("current_package", "").strip() or None
    subscription.current_monthly_fee = monthly_fee
    subscription.status = status
    subscription.notes = form.get("notes", "").strip() or None
    subscription.teltonika_device_id = teltonika_device_id
    return None


M2M_IMPORT_ALIASES = {
    "subscriber_name": ("elofizeto", "subscriber name"),
    "account_number": ("folyoszamlaszam", "account number"),
    "contract_number": ("szerzodesszam", "contract number"),
    "registration_date": ("rogzites datuma", "registration date"),
    "phone_number": ("hivoszam", "telefonszam", "phone number"),
    "device_number": ("eszkoz szam", "device number"),
    "location_name": ("helyszin", "location"),
    "device_identifier": ("eszkoz azonosito", "device identifier"),
    "sim_number": ("sim", "sim szam", "sim number", "iccid"),
    "tariff_name": ("dijcsomag", "tariff name"),
    "current_package": ("csomag", "aktualis csomag", "current package"),
    "current_monthly_fee": ("havidij", "aktualis havidij", "monthly fee"),
    "status": ("statusz", "status"),
    "notes": ("megjegyzes", "notes"),
}


def parse_m2m_import_file(path):
    if path.lower().endswith(".xlsx"):
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        values = list(sheet.iter_rows(values_only=True))
        workbook.close()
        if not values:
            return []
        headers = [clean_string(value) or "" for value in values[0]]
        return [
            {
                "row_number": row_number,
                "values": {
                    headers[index]: value
                    for index, value in enumerate(row)
                    if index < len(headers) and headers[index]
                },
            }
            for row_number, row in enumerate(values[1:], start=2)
            if any(value not in (None, "") for value in row)
        ]

    with open(path, "rb") as source:
        raw = source.read()
    try:
        text = raw.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = raw.decode("cp1250")
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=";,\\t")
        reader = csv.DictReader(StringIO(text), dialect=dialect)
    except csv.Error:
        reader = csv.DictReader(StringIO(text), delimiter=";")
    if not reader.fieldnames:
        return []
    return [
        {"row_number": row_number, "values": dict(row)}
        for row_number, row in enumerate(reader, start=2)
        if any(value not in (None, "") for value in row.values())
    ]


def m2m_import_value(values, field):
    normalized_values = {
        normalize_key(header): value for header, value in values.items()
    }
    for alias in M2M_IMPORT_ALIASES[field]:
        value = normalized_values.get(normalize_key(alias))
        if value not in (None, ""):
            return value
    return None


def parse_m2m_period_header(header):
    normalized = normalize_key(header)
    match = re.search(r"\b(20\d{2})\s+(0?[1-9]|1[0-2])\b", normalized)
    if not match:
        return None
    if not any(token in normalized for token in ("forgalom", "usage", "mb")):
        return None
    return int(match.group(1)), int(match.group(2))


def parse_m2m_status(value):
    normalized = normalize_key(value)
    mapping = {
        "aktiv": "active",
        "active": "active",
        "felfuggesztve": "suspended",
        "suspended": "suspended",
        "inaktiv": "inactive",
        "inactive": "inactive",
        "megszuntetve": "cancelled",
        "cancelled": "cancelled",
        "torolt": "cancelled",
    }
    return mapping.get(normalized, "active" if not normalized else None)


def import_m2m_rows(
    rows,
    M2MSubscription,
    M2MMonthlyUsage,
    M2MPackageHistory,
    Device,
):
    result = {"created": 0, "updated": 0, "usages": 0, "errors": []}
    for raw_row in rows:
        values = raw_row["values"]
        row_number = raw_row["row_number"]
        sim_number = clean_string(m2m_import_value(values, "sim_number"))
        phone_number = clean_string(m2m_import_value(values, "phone_number"))
        device_number = clean_string(m2m_import_value(values, "device_number"))
        if not any((sim_number, phone_number, device_number)):
            result["errors"].append(
                f"{row_number}. sor: hiányzik a SIM, hívószám és eszközszám."
            )
            continue

        subscription = None
        if sim_number:
            subscription = M2MSubscription.query.filter_by(
                sim_number=sim_number
            ).first()
        if subscription is None and phone_number:
            subscription = M2MSubscription.query.filter_by(
                phone_number=phone_number
            ).first()
        if subscription is None and device_number:
            subscription = M2MSubscription.query.filter_by(
                device_number=device_number
            ).first()
        created = subscription is None
        if created:
            subscription = M2MSubscription(status="active")
            db.session.add(subscription)

        status = parse_m2m_status(m2m_import_value(values, "status"))
        if status is None:
            result["errors"].append(
                f"{row_number}. sor: ismeretlen státusz."
            )
            if created:
                db.session.expunge(subscription)
            continue

        field_values = {
            "subscriber_name": clean_string(m2m_import_value(values, "subscriber_name")),
            "account_number": clean_string(m2m_import_value(values, "account_number")),
            "contract_number": clean_string(m2m_import_value(values, "contract_number")),
            "registration_date": date_value(m2m_import_value(values, "registration_date")),
            "phone_number": phone_number,
            "device_number": device_number,
            "location_name": clean_string(m2m_import_value(values, "location_name")),
            "device_identifier": clean_string(m2m_import_value(values, "device_identifier")),
            "sim_number": sim_number,
            "tariff_name": clean_string(m2m_import_value(values, "tariff_name")),
            "current_package": clean_string(m2m_import_value(values, "current_package")),
            "current_monthly_fee": optional_decimal(
                m2m_import_value(values, "current_monthly_fee")
            ),
            "status": status,
            "notes": clean_string(m2m_import_value(values, "notes")),
        }
        for field, value in field_values.items():
            if value is not None or field == "status":
                setattr(subscription, field, value)
        db.session.flush()

        if created:
            result["created"] += 1
            if subscription.current_package:
                db.session.add(
                    M2MPackageHistory(
                        subscription_id=subscription.id,
                        package_name=subscription.current_package,
                        monthly_fee=subscription.current_monthly_fee,
                        valid_from=subscription.registration_date or date.today(),
                        notes="Importált kezdő csomag.",
                    )
                )
        else:
            result["updated"] += 1

        for header, raw_usage in values.items():
            period = parse_m2m_period_header(header)
            if period is None or raw_usage in (None, ""):
                continue
            usage_mb = optional_decimal(raw_usage)
            if usage_mb is None or usage_mb < 0:
                result["errors"].append(
                    f"{row_number}. sor: hibás havi forgalom ({header})."
                )
                continue
            year, month = period
            usage = M2MMonthlyUsage.query.filter_by(
                subscription_id=subscription.id,
                year=year,
                month=month,
                source="import",
            ).first()
            if usage is None:
                usage = M2MMonthlyUsage(
                    subscription_id=subscription.id,
                    year=year,
                    month=month,
                    source="import",
                )
                db.session.add(usage)
            usage.usage_mb = usage_mb
            result["usages"] += 1
    return result


def build_m2m_import_template():
    from openpyxl.styles import Font, PatternFill

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "M2M SIM-ek"
    headers = [
        "Előfizető",
        "Folyószámlaszám",
        "Szerződésszám",
        "Rögzítés dátuma",
        "Hívószám",
        "Eszköz szám",
        "Helyszín",
        "Eszköz azonosító",
        "SIM",
        "Díjcsomag",
        "Csomag",
        "Havidíj",
        "Státusz",
        "Megjegyzés",
        f"{date.today().year}-{date.today().month:02d} forgalom (MB)",
    ]
    sheet.append(headers)
    sheet.append(
        [
            "Parkl Digital Technologies Kft.",
            "12345678",
            "M2M-2026-001",
            date.today(),
            "+36301234567",
            "DEV-001",
            "Arena helyszín",
            "RUT241-001",
            "8944100000000000001",
            "M2M adat",
            "1 GB",
            1990,
            "Aktív",
            "Minta sor, import előtt törölhető.",
            420,
        ]
    )
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{sheet.cell(1, len(headers)).column_letter}2"
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(fill_type="solid", fgColor="6D45C4")
    widths = [28, 18, 18, 18, 18, 16, 24, 22, 24, 18, 16, 14, 16, 34, 24]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[sheet.cell(1, index).column_letter].width = width
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def project_form_data(form):
    latitude_raw = form.get("latitude", "").strip()
    longitude_raw = form.get("longitude", "").strip()
    latitude = optional_decimal(latitude_raw)
    longitude = optional_decimal(longitude_raw)
    error = None
    if latitude_raw and latitude is None:
        error = "A latitude értéke nem érvényes szám."
    elif longitude_raw and longitude is None:
        error = "A longitude értéke nem érvényes szám."
    elif latitude is not None and not Decimal("-90") <= latitude <= Decimal("90"):
        error = "A latitude értékének -90 és 90 között kell lennie."
    elif longitude is not None and not Decimal("-180") <= longitude <= Decimal("180"):
        error = "A longitude értékének -180 és 180 között kell lennie."
    return {
        "name": form.get("name", "").strip(),
        "code": form.get("code", "").strip(),
        "customer": form.get("customer", "").strip(),
        "site_name": form.get("site_name", "").strip() or None,
        "address": form.get("address", "").strip() or None,
        "city": form.get("city", "").strip() or None,
        "country": form.get("country", "").strip() or None,
        "latitude": latitude,
        "longitude": longitude,
        "google_maps_url": form.get("google_maps_url", "").strip() or None,
        "site_notes": form.get("site_notes", "").strip() or None,
        "status": form.get("status", "planned").strip() or "planned",
        "notes": form.get("notes", "").strip(),
    }, error


def checkbox_value(value):
    return True if value == "on" else False


def optional_time(value):
    if not value:
        return None
    try:
        return datetime.strptime(value, "%H:%M").time()
    except ValueError:
        return None


def work_order_type_label(value):
    return WORK_ORDER_TYPE_LABELS.get(value, value)


def work_order_status_label(value):
    return WORK_ORDER_STATUS_LABELS.get(value, value)


def work_order_photo_category_label(value):
    return WORK_ORDER_PHOTO_CATEGORY_LABELS.get(value, value)


def format_duration(minutes):
    if minutes is None:
        return "-"
    hours, remaining = divmod(minutes, 60)
    if hours and remaining:
        return f"{hours} óra {remaining} perc"
    if hours:
        return f"{hours} óra"
    return f"{remaining} perc"


def next_work_order_number(work_order_model):
    prefix = date.today().strftime("%Y%m%d")
    count = work_order_model.query.filter(work_order_model.number.like(f"{prefix}%")).count()
    return f"{prefix}_{count + 1:03d}"


def seed_work_order_templates(template_model):
    defaults = [
        ("AC töltő karbantartás", "maintenance"),
        ("Schneider hibaelhárítás", "troubleshooting"),
        ("Circontrol hibaelhárítás", "troubleshooting"),
        ("Teltonika csere", "troubleshooting"),
        ("Kábelcsere", "cable_replacement"),
        ("Helyszíni felmérés", "inspection"),
    ]
    existing = {item.name for item in template_model.query.all()}
    for name, work_type in defaults:
        if name not in existing:
            db.session.add(template_model(name=name, work_type=work_type))


def update_work_order_from_form(work_order, form):
    number = form.get("number", "").strip()
    work_type = form.get("work_type", "").strip()
    status = form.get("status", "draft").strip()
    if not number:
        return "A munkalap száma kötelező."
    if work_type not in WORK_ORDER_TYPE_LABELS:
        return "Válassz érvényes munkalap típust."
    if status not in WORK_ORDER_STATUS_LABELS:
        return "Válassz érvényes munkalap státuszt."

    work_order.number = number
    work_order.work_type = work_type
    work_order.created_date = optional_date(form.get("created_date")) or date.today()
    work_order.work_date = optional_date(form.get("work_date"))
    work_order.status = status
    work_order.customer_name = form.get("customer_name", "").strip() or None
    work_order.customer_address = form.get("customer_address", "").strip() or None
    work_order.contact_name = form.get("contact_name", "").strip() or None
    work_order.phone = form.get("phone", "").strip() or None
    work_order.email = form.get("email", "").strip() or None
    work_order.site_name = form.get("site_name", "").strip() or None
    work_order.site_address = form.get("site_address", "").strip() or None
    work_order.site_city = form.get("site_city", "").strip() or None
    work_order.site_notes = form.get("site_notes", "").strip() or None
    work_order.device_manufacturer = form.get("device_manufacturer", "").strip() or None
    work_order.device_type = form.get("device_type", "").strip() or None
    work_order.device_serial_number = form.get("device_serial_number", "").strip() or None
    work_order.device_purchase_date = optional_date(form.get("device_purchase_date"))
    work_order.arrival_time = optional_time(form.get("arrival_time"))
    work_order.departure_time = optional_time(form.get("departure_time"))
    work_order.fault_description = form.get("fault_description", "").strip() or None
    work_order.work_performed = form.get("work_performed", "").strip() or None
    work_order.labor_settlement = form.get("labor_settlement", "").strip() or None
    work_order.material_settlement = form.get("material_settlement", "").strip() or None
    work_order.notes = form.get("notes", "").strip() or None
    work_order.technician_name = form.get("technician_name", "").strip() or None
    work_order.second_technician = form.get("second_technician", "").strip() or None
    work_order.subcontractor = form.get("subcontractor", "").strip() or None
    return None


def replace_work_order_rows(work_order, form, material_model, measurement_model):
    work_order.materials.clear()
    material_names = form.getlist("material_name[]")
    for index, name in enumerate(material_names):
        name = name.strip()
        if not name:
            continue
        work_order.materials.append(
            material_model(
                name=name,
                item_number=list_value(form, "material_item_number[]", index),
                quantity=optional_float(list_value(form, "material_quantity[]", index)),
                unit=list_value(form, "material_unit[]", index),
                notes=list_value(form, "material_notes[]", index),
            )
        )

    work_order.measurements.clear()
    measurement_names = form.getlist("measurement_name[]")
    for index, name in enumerate(measurement_names):
        name = name.strip()
        if not name:
            continue
        work_order.measurements.append(
            measurement_model(
                name=name,
                value=list_value(form, "measurement_value[]", index),
                unit=list_value(form, "measurement_unit[]", index),
                notes=list_value(form, "measurement_notes[]", index),
            )
        )


def list_value(form, key, index):
    values = form.getlist(key)
    if index >= len(values):
        return None
    return values[index].strip() or None


def save_work_order_uploads(app, work_order, files, form, photo_model):
    upload_dir = os.path.join(app.instance_path, WORK_ORDER_UPLOAD_SUBDIR)
    os.makedirs(upload_dir, exist_ok=True)

    delete_ids = {optional_int(value) for value in form.getlist("delete_photo_ids[]")}
    for photo in list(work_order.photos):
        if photo.id in delete_ids:
            work_order.photos.remove(photo)

    for category in WORK_ORDER_PHOTO_CATEGORY_LABELS:
        for upload in files.getlist(f"photos_{category}"):
            if not upload or not upload.filename or not allowed_photo_file(upload.filename):
                continue
            extension = upload.filename.rsplit(".", 1)[1].lower()
            filename = secure_filename(f"{work_order.number}_{category}_{uuid4().hex}.{extension}")
            path = os.path.join(upload_dir, filename)
            upload.save(path)
            if valid_image_file(path):
                work_order.photos.append(photo_model(category=category, filename=filename))
            else:
                os.remove(path)

    for field, attribute in (
        ("technician_signature", "technician_signature_filename"),
        ("customer_signature", "customer_signature_filename"),
    ):
        data_url = form.get(field, "")
        if data_url.startswith("data:image/png;base64,"):
            filename = secure_filename(f"{work_order.number}_{field}_{uuid4().hex}.png")
            path = os.path.join(upload_dir, filename)
            with open(path, "wb") as output:
                output.write(base64.b64decode(data_url.split(",", 1)[1]))
            if valid_image_file(path):
                setattr(work_order, attribute, filename)
            else:
                os.remove(path)


def allowed_photo_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_PHOTO_EXTENSIONS


def valid_image_file(path):
    try:
        with PILImage.open(path) as image:
            image.verify()
        return True
    except (UnidentifiedImageError, OSError, ValueError):
        return False


def template_json_rows(template, attribute):
    if template is None:
        return []
    try:
        return json.loads(getattr(template, attribute) or "[]")
    except json.JSONDecodeError:
        return []


def update_work_order_template_from_form(template, form):
    name = form.get("name", "").strip()
    work_type = form.get("work_type", "").strip() or None
    if not name:
        return "A sablon neve kötelező."
    if work_type and work_type not in WORK_ORDER_TYPE_LABELS:
        return "Válassz érvényes munkalap típust."
    template.name = name
    template.work_type = work_type
    template.fault_description = form.get("fault_description", "").strip() or None
    template.work_performed = form.get("work_performed", "").strip() or None
    template.notes = form.get("notes", "").strip() or None
    template.materials_json = json.dumps(form_rows_to_dicts(form, "material"), ensure_ascii=False)
    template.measurements_json = json.dumps(form_rows_to_dicts(form, "measurement"), ensure_ascii=False)
    return None


def form_rows_to_dicts(form, prefix):
    names = form.getlist(f"{prefix}_name[]")
    keys = (
        ("item_number", f"{prefix}_item_number[]"),
        ("quantity", f"{prefix}_quantity[]"),
        ("unit", f"{prefix}_unit[]"),
        ("value", f"{prefix}_value[]"),
        ("notes", f"{prefix}_notes[]"),
    )
    rows = []
    for index, name in enumerate(names):
        if not name.strip():
            continue
        row = {"name": name.strip()}
        for output_key, form_key in keys:
            value = list_value(form, form_key, index)
            if value is not None:
                row[output_key] = value
        rows.append(row)
    return rows


def copy_work_order(source, user_id, work_order_model, material_model, measurement_model):
    copied = work_order_model(
        number=f"{source.number}-M-{uuid4().hex[:4].upper()}",
        work_type=source.work_type,
        created_date=date.today(),
        work_date=source.work_date,
        status="draft",
        customer_name=source.customer_name,
        customer_address=source.customer_address,
        contact_name=source.contact_name,
        phone=source.phone,
        email=source.email,
        site_name=source.site_name,
        site_address=source.site_address,
        site_city=source.site_city,
        site_notes=source.site_notes,
        device_manufacturer=source.device_manufacturer,
        device_type=source.device_type,
        device_serial_number=source.device_serial_number,
        device_purchase_date=source.device_purchase_date,
        fault_description=source.fault_description,
        work_performed=source.work_performed,
        labor_settlement=source.labor_settlement,
        material_settlement=source.material_settlement,
        notes=source.notes,
        technician_name=source.technician_name,
        second_technician=source.second_technician,
        subcontractor=source.subcontractor,
        created_by_id=user_id,
    )
    copied.materials = [
        material_model(
            name=item.name,
            item_number=item.item_number,
            quantity=item.quantity,
            unit=item.unit,
            notes=item.notes,
        )
        for item in source.materials
    ]
    copied.measurements = [
        measurement_model(name=item.name, value=item.value, unit=item.unit, notes=item.notes)
        for item in source.measurements
    ]
    return copied


def calculate_huf_value(quantity, unit_net_price, currency):
    if currency == "HUF" and quantity is not None and unit_net_price is not None:
        return quantity * unit_net_price
    return None


def calculate_imported_huf_value(quantity, unit_net_price, currency, excel_huf_value):
    if excel_huf_value is not None:
        if quantity is not None and unit_net_price is not None:
            return quantity * excel_huf_value
        return excel_huf_value
    return calculate_huf_value(quantity, unit_net_price, currency)


def calculate_line_net_amount(quantity, unit_price):
    if quantity is not None and unit_price is not None:
        return quantity * unit_price
    return None


def device_currency_totals(devices):
    totals = {
        "net_huf": 0,
        "net_eur": 0,
        "gross_huf": 0,
        "gross_eur": 0,
        "missing_currency_count": 0,
    }
    for device in devices:
        if device.currency not in {"HUF", "EUR"}:
            totals["missing_currency_count"] += 1
            continue
        currency_key = device.currency.lower()
        if device.total_net_price is not None:
            totals[f"net_{currency_key}"] += device.total_net_price
        if device.total_gross_price is not None:
            totals[f"gross_{currency_key}"] += device.total_gross_price
    return totals


def empty_currency_totals():
    return {
        "net_huf": 0,
        "net_eur": 0,
        "gross_huf": 0,
        "gross_eur": 0,
        "missing_currency_count": 0,
    }


def add_finance_value(totals, device, quantity):
    if device.currency not in {"HUF", "EUR"}:
        totals["missing_currency_count"] += 1
        return
    currency = device.currency.lower()
    if device.unit_net_price is not None:
        totals[f"net_{currency}"] += device.unit_net_price * quantity
    if device.unit_gross_price is not None:
        totals[f"gross_{currency}"] += device.unit_gross_price * quantity


def finance_rows_totals(rows):
    totals = empty_currency_totals()
    for row in rows:
        add_finance_value(totals, row["device"], row["quantity"])
    return totals


def device_finance_issues(device):
    issues = []
    if device.quantity in (None, 0):
        issues.append("Hiányzó vagy nulla mennyiség")
    if device.currency not in {"HUF", "EUR"}:
        issues.append("Hiányzó vagy ismeretlen deviza")
    if device.unit_net_price is None and device.huf_value is None:
        issues.append("Hiányzó nettó egységár vagy importált összérték")
    if device.vat_rate is None:
        issues.append("Hiányzó ÁFA")
    if not (device.supplier_manufacturer or device.manufacturer):
        issues.append("Hiányzó beszállító / gyártó")
    return issues


def invoice_item_value(item):
    if item.line_gross_amount_huf is not None:
        return item.line_gross_amount_huf
    if item.gross_amount_huf is not None:
        return item.gross_amount_huf
    return line_net_amount(item) or 0


def project_finance_snapshot(project):
    rows = project_inventory_rows(project)
    totals = finance_rows_totals(rows)
    invoice_items = [
        item
        for item in project.unassigned_invoice_items
        if item.archived_at is None
    ]
    unique_devices = {row["device"].id: row["device"] for row in rows}.values()
    totals.update(
        {
            "project": project,
            "rows": rows,
            "item_quantity": sum(row["quantity"] for row in rows),
            "invoice_count": len(invoice_items),
            "invoice_value_huf": sum(invoice_item_value(item) for item in invoice_items),
            "unpaid_count": sum(
                1
                for device in unique_devices
                if device.supplier_invoice_number
                and device.supplier_invoice_paid is not True
            ),
            "missing_count": sum(
                1 for device in unique_devices if device_finance_issues(device)
            ),
        }
    )
    return totals


def device_active_project_codes(device):
    if device.tracking_mode == "unit":
        projects = {
            unit.project.code
            for unit in active_device_units(device)
            if unit.project is not None and unit.status in PROJECT_ACTIVE_STATUSES
        }
    else:
        projects = {
            balance.project.code
            for balance in active_bulk_balances(device)
            if balance.project is not None and balance.status in PROJECT_ACTIVE_STATUSES
        }
    return ", ".join(sorted(projects)) or "-"


def project_bom_rows(project):
    grouped = {}
    for row in project_inventory_rows(project):
        device = row["device"]
        key = (device.id, row["status"])
        if key not in grouped:
            grouped[key] = {
                "device": device,
                "status": row["status"],
                "quantity": 0,
                "unit_count": 0,
                "net_total": None,
                "gross_total": None,
            }
        grouped[key]["quantity"] += row["quantity"]
        grouped[key]["unit_count"] += 1 if row["unit"] else 0
    for item in grouped.values():
        device = item["device"]
        if device.unit_net_price is not None and device.currency in {"HUF", "EUR"}:
            item["net_total"] = device.unit_net_price * item["quantity"]
        if device.unit_gross_price is not None and device.currency in {"HUF", "EUR"}:
            item["gross_total"] = device.unit_gross_price * item["quantity"]
    return sorted(
        grouped.values(),
        key=lambda item: (
            item["device"].device_type or "",
            item["device"].product_name or item["device"].asset_tag,
            item["status"],
        ),
    )


def line_net_amount(item):
    if item.net_amount_huf is not None:
        return item.net_amount_huf
    return calculate_line_net_amount(item.quantity, item.unit_price_huf)


def status_label(value):
    return STATUS_LABELS.get(value, value)


def movement_type_label(value):
    return MOVEMENT_TYPE_LABELS.get(value, value)


def category_label(value):
    return CATEGORY_LABELS.get(value, value)


def device_display_label(device):
    return device.human_label


def device_primary_label(device):
    return device.primary_label


def device_qr_mode_label(value):
    return DEVICE_QR_MODE_LABELS.get(value, value)


def tracking_mode_label(value):
    return {
        "bulk": "Mennyiségi követés",
        "unit": "Egyedi példánykövetés",
    }.get(value, value)


def bulk_balance_summary(device, field):
    balances = [
        balance
        for balance in device.bulk_balances
        if balance.quantity is not None and balance.quantity > 1e-9
    ]
    if not balances:
        return "-"
    if field == "status":
        values = {status_label(balance.status) for balance in balances}
    elif field == "project":
        values = {
            balance.project.code if balance.project else "Nincs projekt"
            for balance in balances
        }
    elif field == "location":
        values = {
            balance.location.name if balance.location else "Nincs készlethely"
            for balance in balances
        }
    else:
        return "-"
    return next(iter(values)) if len(values) == 1 else f"{len(values)} féle"


def device_inventory_subjects(device):
    if device.tracking_mode == "unit":
        return active_device_units(device)
    return [
        balance
        for balance in device.bulk_balances
        if balance.quantity is not None and balance.quantity > 1e-9
    ]


def device_inventory_values(device, field):
    subjects = device_inventory_subjects(device)

    if field == "project":
        values = {
            subject.project.code
            for subject in subjects
            if subject.project is not None
            and subject.status in PROJECT_ACTIVE_STATUSES
        }
    elif field == "location":
        values = {
            subject.location.name
            for subject in subjects
            if subject.location is not None
            and subject.status in PHYSICAL_LOCATION_STATUSES
        }
    elif field == "status":
        values = {subject.status for subject in subjects}
    else:
        values = set()
    return sorted(values)


def device_inventory_export_value(device, field):
    subjects = device_inventory_subjects(device)
    values = device_inventory_values(device, field)
    if field == "status":
        if len(values) == 1:
            return values[0]
        return f"MIXED: {'; '.join(values)}" if values else ""

    attribute = "project_id" if field == "project" else "location_id"
    raw_values = {getattr(subject, attribute) for subject in subjects}
    if len(raw_values) == 1:
        return values[0] if values else ""
    return "MIXED"


def available_device_movements(subject):
    transitions = {
        "IN_STOCK": ["RESERVE", "ISSUE", "TRANSFER", "SERVICE", "SCRAP"],
        "RESERVED": ["ISSUE", "RELEASE", "SCRAP"],
        "ISSUED": ["INSTALL", "RETURN", "SERVICE"],
        "INSTALLED": ["RETURN", "SERVICE", "SCRAP"],
        "RETURNED": ["INBOUND", "TRANSFER", "ISSUE", "INSTALL"],
        "IN_SERVICE": ["RETURN", "SCRAP"],
        "SCRAPPED": [],
    }
    return transitions.get(subject.status, [])


def movement_allowed_statuses():
    return {
        "INBOUND": {"RETURNED", "IN_SERVICE"},
        "RESERVE": {"IN_STOCK"},
        "ISSUE": {"IN_STOCK", "RESERVED", "RETURNED"},
        "INSTALL": {"ISSUED", "RETURNED"},
        "RETURN": {"ISSUED", "INSTALLED", "IN_SERVICE"},
        "SERVICE": {"IN_STOCK", "RETURNED", "ISSUED", "INSTALLED"},
        "SCRAP": None,
        "TRANSFER": {"IN_STOCK", "RETURNED"},
        "RELEASE": {"RESERVED"},
    }


def reversed_movement_ids(movements):
    from models import StockMovement

    movement_ids = [movement.id for movement in movements if movement.id is not None]
    if not movement_ids:
        return set()
    return {
        movement_id
        for (movement_id,) in db.session.query(StockMovement.reversal_of_movement_id)
        .filter(StockMovement.reversal_of_movement_id.in_(movement_ids))
        .all()
        if movement_id is not None
    }


def reversible_movement_ids(movements):
    return {
        movement.id
        for movement in movements
        if movement.id is not None and not movement_reversal_blockers(movement)
    }


def movement_reversal_blockers(movement):
    from models import BulkStockBalance, StockMovement

    blockers = []
    if movement.reversal_of_movement_id is not None or movement.movement_type == "REVERSAL":
        blockers.append("Ellenmozgás nem vonható vissza újabb ellenmozgással.")
        return blockers
    if StockMovement.query.filter_by(reversal_of_movement_id=movement.id).first():
        blockers.append("Ezt a mozgást már visszavonták.")
        return blockers
    if movement.quantity is None or movement.quantity <= 0:
        blockers.append("A régi mozgásnak nincs egyértelmű mennyisége.")
        return blockers
    if movement.device.archived_at is not None:
        blockers.append("Az eszköztétel archiválva van.")

    if movement.unit_id is not None:
        if StockMovement.query.filter(
            StockMovement.unit_id == movement.unit_id,
            StockMovement.id > movement.id,
        ).first():
            blockers.append(
                "Csak a példány legutolsó mozgása vonható vissza; későbbi "
                "mozgás már épül erre az állapotra."
            )
            return blockers
        unit = movement.unit
        if unit is None or unit.archived_at is not None:
            blockers.append("Az érintett eszközpéldány nem aktív.")
            return blockers
        if movement.from_status is None:
            blockers.append("A példány eredeti kiinduló állapota nem ismert.")
            return blockers
        if (
            unit.status != movement.to_status
            or unit.location_id != movement.to_location_id
            or unit.project_id != movement.to_project_id
        ):
            blockers.append(
                "A példány aktuális állapota már nem egyezik a mozgás eredményével."
            )
        return blockers

    if movement.device.tracking_mode != "bulk":
        blockers.append("A mozgás nem köthető egyértelműen bulk készlethez vagy példányhoz.")
        return blockers
    if StockMovement.query.filter(
        StockMovement.device_id == movement.device_id,
        StockMovement.unit_id.is_(None),
        StockMovement.id > movement.id,
    ).first():
        blockers.append(
            "Csak a bulk tétel legutolsó mozgása vonható vissza; későbbi "
            "mennyiségi mozgás már épül az egyenlegekre."
        )
        return blockers
    target_balance = BulkStockBalance.query.filter_by(
        device_id=movement.device_id,
        status=movement.to_status,
        location_id=movement.to_location_id,
        project_id=movement.to_project_id,
    ).first()
    if target_balance is None or target_balance.quantity + 1e-9 < movement.quantity:
        blockers.append(
            "A mozgás eredményoldalán már nincs elegendő, azonos állapotú mennyiség."
        )
    return blockers


def reverse_stock_movement(original, user_id):
    from models import BulkStockBalance, DeviceUnit, StockMovement

    blockers = movement_reversal_blockers(original)
    if blockers:
        raise ValueError("A mozgás nem vonható vissza: " + " ".join(blockers))

    quantity = original.quantity
    if original.unit_id is not None:
        unit = (
            DeviceUnit.query.filter_by(id=original.unit_id)
            .with_for_update()
            .one()
        )
        reversal = StockMovement(
            device_id=original.device_id,
            unit_id=unit.id,
            movement_type="REVERSAL",
            quantity=1,
            from_location_id=original.to_location_id,
            to_location_id=original.from_location_id,
            project_id=original.from_project_id,
            from_project_id=original.to_project_id,
            to_project_id=original.from_project_id,
            reversal_of_movement_id=original.id,
            from_status=original.to_status,
            to_status=original.from_status,
            notes=f"#{original.id} mozgás visszavonása ellenmozgással.",
            created_by_id=user_id,
        )
        db.session.add(reversal)
        unit.status = original.from_status
        unit.location_id = original.from_location_id
        unit.project_id = original.from_project_id
        unit.updated_at = now_utc()
        db.session.flush()
        return reversal

    target_balance = (
        BulkStockBalance.query.filter_by(
            device_id=original.device_id,
            status=original.to_status,
            location_id=original.to_location_id,
            project_id=original.to_project_id,
        )
        .with_for_update()
        .one()
    )
    if target_balance.quantity + 1e-9 < quantity:
        raise ValueError("A visszavonáshoz nincs elegendő mennyiség a célállapotban.")
    target_balance.quantity -= quantity
    if abs(target_balance.quantity) <= 1e-9:
        target_balance.quantity = 0

    if original.from_status is not None:
        restored_balance = find_or_create_bulk_balance(
            original.device,
            original.from_status,
            original.from_location_id,
            original.from_project_id,
        )
        restored_balance.quantity += quantity
    else:
        original.device.quantity = max((original.device.quantity or 0) - quantity, 0)

    reversal = StockMovement(
        device_id=original.device_id,
        movement_type="REVERSAL",
        quantity=quantity,
        from_location_id=original.to_location_id,
        to_location_id=original.from_location_id,
        project_id=original.from_project_id,
        from_project_id=original.to_project_id,
        to_project_id=original.from_project_id,
        reversal_of_movement_id=original.id,
        from_status=original.to_status,
        to_status=original.from_status,
        notes=f"#{original.id} mozgás visszavonása ellenmozgással.",
        created_by_id=user_id,
    )
    db.session.add(reversal)
    sync_bulk_device_legacy_state(original.device)
    db.session.flush()
    return reversal


def flash_archive_blockers(title, blockers):
    flash(title, "danger")
    for blocker in blockers:
        flash(blocker, "warning")


def device_unit_archive_blockers(unit):
    if unit.status != "SCRAPPED":
        return [
            f"A példány jelenlegi státusza {status_label(unit.status)}. "
            "Archiválás előtt selejtezési mozgás szükséges."
        ]
    return []


def device_archive_blockers(device):
    blockers = []
    active_balances = [
        balance
        for balance in device.bulk_balances
        if balance.quantity is not None
        and balance.quantity > 1e-9
        and balance.status != "SCRAPPED"
    ]
    if active_balances:
        quantity = sum(balance.quantity for balance in active_balances)
        blockers.append(
            f"{format_number(quantity)} aktív készletmennyiség maradt "
            f"{len(active_balances)} egyenlegen. Mozgasd vagy selejtezd a készletet."
        )
    active_units = [
        unit
        for unit in device.units
        if unit.archived_at is None and unit.status != "SCRAPPED"
    ]
    if active_units:
        blockers.append(
            f"{len(active_units)} aktív eszközpéldány kapcsolódik a tételhez. "
            "Előbb selejtezd vagy rendezd a példányokat."
        )
    scrapped_units = [
        unit
        for unit in device.units
        if unit.archived_at is None and unit.status == "SCRAPPED"
    ]
    if scrapped_units:
        blockers.append(
            f"{len(scrapped_units)} selejtezett, de még nem archivált példány kapcsolódik "
            "a tételhez. Archiváld előbb a példányokat."
        )
    return blockers


def project_archive_blockers(project):
    from models import DeviceUnit

    blockers = []
    if project.status != "completed":
        blockers.append(
            f"A projekt státusza {project_status_label(project.status)}. "
            "Csak lezárt projekt archiválható."
        )
    active_balances = [
        balance
        for balance in project.bulk_balances
        if balance.quantity is not None
        and balance.quantity > 1e-9
        and balance.status != "SCRAPPED"
    ]
    if active_balances:
        blockers.append(
            f"{format_number(sum(balance.quantity for balance in active_balances))} "
            "aktív bulk készletmennyiség van még a projekten."
        )
    active_units = [
        unit
        for unit in DeviceUnit.query.filter_by(project_id=project.id)
        .filter(DeviceUnit.archived_at.is_(None))
        .all()
        if unit.status != "SCRAPPED"
    ]
    if active_units:
        blockers.append(
            f"{len(active_units)} aktív egyedi eszközpéldány van még a projekthez rendelve."
        )
    return blockers


def location_archive_blockers(location):
    blockers = []
    active_balances = [
        balance
        for balance in location.bulk_balances
        if balance.quantity is not None
        and balance.quantity > 1e-9
        and balance.status != "SCRAPPED"
    ]
    if active_balances:
        blockers.append(
            f"{format_number(sum(balance.quantity for balance in active_balances))} "
            "aktív bulk készletmennyiség található ezen a készlethelyen."
        )
    active_units = [
        unit
        for unit in location.device_units
        if unit.archived_at is None and unit.status != "SCRAPPED"
    ]
    if active_units:
        blockers.append(
            f"{len(active_units)} aktív egyedi eszközpéldány található ezen a készlethelyen."
        )
    return blockers


def whole_device_quantity(device):
    if device.quantity is None or device.quantity <= 0 or not float(device.quantity).is_integer():
        return None
    return int(device.quantity)


def device_quantity_supports_existing_units(device, quantity):
    active_unit_count = sum(1 for unit in device.units if unit.archived_at is None)
    if active_unit_count == 0:
        return True
    return (
        quantity is not None
        and float(quantity).is_integer()
        and int(quantity) >= active_unit_count
    )


def default_unit_code_prefix(device):
    value = device.asset_tag or f"DEVICE-{device.id}"
    return re.sub(r"[^A-Za-z0-9]+", "-", value).strip("-").upper()


def unique_device_asset_tag(device_model, value):
    normalized = unicodedata.normalize("NFKD", value or "ESZKOZ")
    ascii_value = normalized.encode("ascii", "ignore").decode("ascii")
    prefix = re.sub(r"[^A-Za-z0-9]+", "-", ascii_value).strip("-").upper()
    prefix = (prefix or "ESZKOZ")[:60]
    candidate = prefix
    number = 1
    while device_model.query.filter_by(asset_tag=candidate).first():
        number += 1
        candidate = f"{prefix[:54]}-{number:03d}"
    return candidate


def available_unit_codes(device_unit_model, prefix, start_number, count):
    clean_prefix = re.sub(r"[^A-Za-z0-9]+", "-", prefix).strip("-").upper() or "UNIT"
    width = max(3, len(str(start_number + count - 1)))
    codes = []
    number = max(start_number, 1)
    while len(codes) < count:
        code = f"{clean_prefix}-{number:0{width}d}"
        if not device_unit_model.query.filter_by(unit_code=code).first():
            codes.append(code)
        number += 1
    return codes


def status_badge_class(value):
    return {
        "IN_STOCK": "status-in-stock",
        "RESERVED": "status-reserved",
        "ISSUED": "status-issued",
        "INSTALLED": "status-installed",
        "RETURNED": "status-returned",
        "IN_SERVICE": "status-service",
        "SCRAPPED": "status-scrapped",
    }.get(value, "status-neutral")


def movement_badge_class(value):
    return {
        "INBOUND": "movement-inbound",
        "RESERVE": "movement-reserve",
        "ISSUE": "movement-issue",
        "INSTALL": "movement-install",
        "RETURN": "movement-return",
        "SERVICE": "movement-service",
        "SCRAP": "movement-scrap",
        "TRANSFER": "movement-transfer",
        "RELEASE": "movement-release",
        "REVERSAL": "movement-reversal",
    }.get(value, "movement-neutral")


def location_type_label(value):
    return LOCATION_TYPE_LABELS.get(
        value,
        LEGACY_LOCATION_TYPE_LABELS.get(value, value),
    )


def active_logistic_locations():
    from models import Location

    return (
        Location.query.filter(
            Location.archived_at.is_(None),
            Location.location_type.in_(LOGISTIC_LOCATION_TYPES),
        )
        .order_by(Location.name.asc())
        .all()
    )


def project_status_label(value):
    return PROJECT_STATUS_LABELS.get(value, value)


def assignment_status_label(value):
    return ASSIGNMENT_STATUS_LABELS.get(value, value)


def m2m_status_label(value):
    return M2M_STATUS_LABELS.get(value, value or "–")


def m2m_usage_source_label(value):
    return M2M_USAGE_SOURCE_LABELS.get(value, value or "–")


def m2m_package_limit_mb(package_name):
    if not package_name:
        return None
    text = str(package_name).lower().replace(",", ".")
    if "korlátlan" in text or "unlimited" in text:
        return None
    matches = re.findall(r"(\d+(?:\.\d+)?)\s*(tb|gb|mb)\b", text)
    if not matches:
        return None
    value, unit = matches[-1]
    multiplier = {"mb": Decimal("1"), "gb": Decimal("1024"), "tb": Decimal("1048576")}
    return Decimal(value) * multiplier[unit]


def m2m_usage_state(package_name, usage_mb):
    limit_mb = m2m_package_limit_mb(package_name)
    if limit_mb is None:
        return {
            "key": "unknown",
            "label": "Nincs limitadat",
            "percent": None,
            "class": "status-neutral",
            "limit_mb": None,
        }
    usage = Decimal(str(usage_mb or 0))
    percent = float(usage / limit_mb * 100) if limit_mb else 0
    if usage > limit_mb:
        key, label, badge_class = "exceeded", "Túllépve", "status-scrapped"
    elif percent >= 80:
        key, label, badge_class = "warning", "Limit közelében", "status-reserved"
    else:
        key, label, badge_class = "normal", "Rendben", "status-in-stock"
    return {
        "key": key,
        "label": label,
        "percent": percent,
        "class": badge_class,
        "limit_mb": limit_mb,
    }


def m2m_subscription_usage_state(subscription, usage_mb):
    if subscription.connection_type == "wired":
        return {
            "key": "wired",
            "label": "Vezetékes kapcsolat",
            "percent": None,
            "class": "status-neutral",
            "limit_mb": None,
        }
    return m2m_usage_state(subscription.current_package, usage_mb)


def m2m_effective_usage_history(subscription):
    source_priority = {"import": 1, "manual": 2, "teltonika_api": 3}
    def created_rank(value):
        if value is None:
            return 0
        if value.tzinfo is None:
            value = value.replace(tzinfo=timezone.utc)
        return value.timestamp()

    monthly = {}
    for usage in subscription.monthly_usages:
        key = (usage.year, usage.month)
        current = monthly.get(key)
        current_rank = (
            source_priority.get(current.source, 0),
            created_rank(current.created_at),
            current.id or 0,
        ) if current else None
        usage_rank = (
            source_priority.get(usage.source, 0),
            created_rank(usage.created_at),
            usage.id or 0,
        )
        if current is None or usage_rank > current_rank:
            monthly[key] = usage
    return [
        {
            "year": year,
            "month": month,
            "usage_mb": usage.usage_mb,
            "source": usage.source,
            "created_at": usage.created_at,
        }
        for (year, month), usage in sorted(monthly.items())
    ]


def m2m_current_usage(subscription):
    today = date.today()
    for item in reversed(m2m_effective_usage_history(subscription)):
        if item["year"] == today.year and item["month"] == today.month:
            return item["usage_mb"]
    return None


def import_status_label(value):
    return IMPORT_STATUS_LABELS.get(value, value)


def yes_no_label(value):
    if value is True:
        return "Igen"
    if value is False:
        return "Nem"
    return "-"


def format_number(value):
    if value is None:
        return "-"
    if float(value).is_integer():
        return f"{int(value):,}".replace(",", " ")
    return f"{value:,.2f}".replace(",", " ").replace(".", ",")


def format_vat_rate(value):
    if value is None:
        return "Nincs megadva"
    return f"{format_number(value)}%"


def device_money_text(device, field):
    value_map = {
        "unit_net": device.unit_net_price,
        "total_net": device.total_net_price,
        "unit_gross": device.unit_gross_price,
        "total_gross": device.total_gross_price,
    }
    value = value_map.get(field)
    if value is not None:
        currency = f" {device.currency}" if device.currency in DEVICE_CURRENCIES else ""
        suffix = "" if currency else " (deviza hiányzik)"
        return f"{format_number(value)}{currency}{suffix}"

    if device.currency not in DEVICE_CURRENCIES:
        return "Nem számolható: hiányzik a deviza"
    if field == "total_net":
        if device.quantity is None:
            return "Nem számolható: hiányzik a mennyiség"
        if device.unit_net_price is None and device.huf_value is None:
            return "Nem számolható: hiányzik az egységár"
    if field in {"unit_gross", "total_gross"}:
        if device.vat_rate is None:
            return "Nem számolható: hiányzik az ÁFA"
        if device.unit_net_price is None:
            return "Nem számolható: hiányzik az egységár"
        if field == "total_gross" and device.quantity is None and device.huf_value is None:
            return "Nem számolható: hiányzik a mennyiség"
    if field == "unit_net":
        return "Nincs megadva"
    return "Nem számolható"


def is_awaiting_arrival(device):
    return device.is_ordered is True and device.has_arrived is not True


def active_device_units(device):
    return [unit for unit in device.units if unit.archived_at is None]


def device_has_status(device, status):
    if device.tracking_mode == "unit":
        return any(unit.status == status for unit in active_device_units(device))
    return any(
        balance.status == status and balance.quantity > 1e-9
        for balance in device.bulk_balances
    )


def device_has_project(device):
    if device.tracking_mode == "unit":
        return any(
            unit.project_id is not None and unit.status in PROJECT_ACTIVE_STATUSES
            for unit in active_device_units(device)
        )
    return any(
        balance.project_id is not None
        and balance.quantity > 1e-9
        and balance.status in PROJECT_ACTIVE_STATUSES
        for balance in device.bulk_balances
    )


def project_inventory_rows(project):
    from models import BulkStockBalance, Device, DeviceUnit

    rows = []
    for balance in BulkStockBalance.query.filter(
        BulkStockBalance.project_id == project.id,
        BulkStockBalance.quantity > 1e-9,
        BulkStockBalance.status.in_(PROJECT_ACTIVE_STATUSES),
    ).join(Device).filter(Device.archived_at.is_(None)).all():
        rows.append(
            {
                "device": balance.device,
                "unit": None,
                "quantity": balance.quantity,
                "status": balance.status,
                "location": balance.location,
            }
        )
    for unit in (
        DeviceUnit.query.filter(
            DeviceUnit.project_id == project.id,
            DeviceUnit.archived_at.is_(None),
            DeviceUnit.status.in_(PROJECT_ACTIVE_STATUSES),
        )
        .join(Device)
        .filter(Device.archived_at.is_(None))
        .all()
    ):
        rows.append(
            {
                "device": unit.device,
                "unit": unit,
                "quantity": 1,
                "status": unit.status,
                "location": unit.location,
            }
        )
    return rows


def project_inventory_summary(project):
    rows = project_inventory_rows(project)
    return {
        "quantity": sum(row["quantity"] for row in rows),
        "unit_count": sum(1 for row in rows if row["unit"] is not None),
        "bulk_quantity": sum(
            row["quantity"] for row in rows if row["unit"] is None
        ),
        "reserved": sum(
            row["quantity"] for row in rows if row["status"] == "RESERVED"
        ),
        "issued": sum(
            row["quantity"] for row in rows if row["status"] == "ISSUED"
        ),
        "installed": sum(
            row["quantity"] for row in rows if row["status"] == "INSTALLED"
        ),
    }


def location_inventory_summary(location):
    bulk_balances = [
        balance
        for balance in location.bulk_balances
        if balance.quantity > 1e-9
        and balance.status in PHYSICAL_LOCATION_STATUSES
        and balance.device.archived_at is None
    ]
    units = [
        unit
        for unit in location.device_units
        if unit.archived_at is None
        and unit.status in PHYSICAL_LOCATION_STATUSES
        and unit.device.archived_at is None
    ]
    reserved = sum(
        balance.quantity
        for balance in bulk_balances
        if balance.status == "RESERVED"
    ) + sum(1 for unit in units if unit.status == "RESERVED")
    total = sum(balance.quantity for balance in bulk_balances) + len(units)
    free = sum(
        balance.quantity
        for balance in bulk_balances
        if balance.status in FREE_STOCK_STATUSES
    ) + sum(1 for unit in units if unit.status in FREE_STOCK_STATUSES)
    service = sum(
        balance.quantity
        for balance in bulk_balances
        if balance.status == "IN_SERVICE"
    ) + sum(1 for unit in units if unit.status == "IN_SERVICE")
    return {
        "physical": total,
        "reserved": reserved,
        "free": free,
        "service": service,
    }


def inventory_rows_currency_totals(rows):
    totals = {
        "net_huf": 0,
        "net_eur": 0,
        "gross_huf": 0,
        "gross_eur": 0,
        "missing_currency_count": 0,
    }
    for row in rows:
        device = row["device"]
        quantity = row["quantity"]
        if device.currency not in {"HUF", "EUR"}:
            totals["missing_currency_count"] += 1
            continue
        currency_key = device.currency.lower()
        if device.unit_net_price is not None:
            totals[f"net_{currency_key}"] += device.unit_net_price * quantity
        if device.unit_gross_price is not None:
            totals[f"gross_{currency_key}"] += device.unit_gross_price * quantity
    return totals


def inventory_status_quantity(devices, units, status):
    bulk_quantity = sum(
        balance.quantity
        for device in devices
        if device.tracking_mode == "bulk"
        for balance in device.bulk_balances
        if balance.status == status and balance.quantity > 1e-9
    )
    return bulk_quantity + sum(1 for unit in units if unit.status == status)


def is_arrived_unassigned(device):
    if device.tracking_mode == "unit":
        units = active_device_units(device)
        return device.has_arrived is True and (
            not units
            or all(not unit.project_id and not unit.location_id for unit in units)
        )
    balances = [
        balance for balance in device.bulk_balances if balance.quantity > 1e-9
    ]
    if balances:
        return device.has_arrived is True and all(
            not balance.project_id and not balance.location_id for balance in balances
        )
    return device.has_arrived is True


def is_financially_open(device):
    return (
        bool(device.supplier_invoice_number) and device.supplier_invoice_paid is not True
    ) or (
        bool(device.shipping_invoice_number) and device.shipping_invoice_paid is not True
    )


def device_attention_reasons(device, include_finance=True):
    reasons = []
    today = date.today()
    if (
        device.is_ordered is True
        and device.has_arrived is not True
        and device.planned_arrival_date
        and device.planned_arrival_date < today
    ):
        reasons.append("Megrendelve, de a tervezett érkezési dátum lejárt.")
    if is_arrived_unassigned(device):
        reasons.append("Megérkezett, de nincs projekthez vagy készlethelyhez rendelve.")
    if include_finance and device.supplier_invoice_number and device.supplier_invoice_paid is not True:
        reasons.append("Beszállítói számla van, de nincs fizetettként jelölve.")
    if include_finance and device.shipping_invoice_number and device.shipping_invoice_paid is not True:
        reasons.append("Szállítmányozói számla van, de nincs fizetettként jelölve.")
    if device.tracking_mode == "bulk":
        projects = {
            balance.project.id: balance.project
            for balance in device.bulk_balances
            if balance.quantity > 1e-9 and balance.project
        }
        if any(
            not project.code
            or not project.name
            or project.name == project.code
            or not project.customer
            for project in projects.values()
        ):
            reasons.append("A kapcsolódó projekt adatai hiányosak.")
    if device.source_sheet and not device.product_name:
        reasons.append("Importált sorból hiányzik a terméknév.")
    if not device.device_type or device.device_type not in CATEGORY_LABELS:
        reasons.append("Hiányzó vagy ismeretlen kategória.")
    if device.source_sheet and (device.quantity is None or device.quantity <= 0):
        reasons.append("Hiányzó vagy nulla mennyiség.")
    if device.tracking_mode == "unit":
        for unit in active_device_units(device):
            if unit.status not in STATUS_LABELS:
                reasons.append(f"{unit.unit_code}: ismeretlen példánystátusz.")
                continue
            state_error = inventory_state_error(
                unit.status,
                unit.location_id,
                unit.project_id,
            )
            if state_error:
                reasons.append(f"{unit.unit_code}: {state_error}.")
    else:
        for balance in device.bulk_balances:
            if balance.quantity <= 1e-9:
                continue
            if balance.status not in STATUS_LABELS:
                reasons.append("Ismeretlen bulk készletstátusz.")
                continue
            state_error = inventory_state_error(
                balance.status,
                balance.location_id,
                balance.project_id,
            )
            if state_error:
                reasons.append(f"Bulk egyenleg: {state_error}.")
    return reasons


def invoice_attention_reasons(item):
    reasons = []
    if item.assignment_status == "unassigned":
        reasons.append("A számlasor nincs hozzárendelve.")
    if not item.invoice_number:
        reasons.append("Hiányzik a számlaszám.")
    if not item.partner:
        reasons.append("Hiányzik a partner.")
    if not item.description:
        reasons.append("Hiányzik a megnevezés.")
    if item.line_gross_amount_huf is None and item.gross_amount_huf is None:
        reasons.append("Hiányzik a bruttó összeg.")
    return reasons


def build_attention_items(devices, invoice_items, include_finance=True):
    items = []
    for device in devices:
        reasons = device_attention_reasons(device, include_finance=include_finance)
        if reasons:
            project_codes = device_inventory_values(device, "project")
            items.append(
                {
                    "type": "Eszköz",
                    "name": device_primary_label(device),
                    "reasons": reasons,
                    "project": ", ".join(project_codes) if project_codes else "-",
                    "supplier": device.supplier_manufacturer or device.manufacturer or "-",
                    "invoice": (
                        device.supplier_invoice_number
                        or device.shipping_invoice_number
                        or "-"
                        if include_finance
                        else "-"
                    ),
                    "source": f"{device.source_sheet or '-'} / {device.source_row_number or '-'}",
                    "detail_url": url_for("device_detail", device_id=device.id),
                    "edit_url": url_for("device_edit", device_id=device.id),
                }
            )
    for item in invoice_items:
        reasons = invoice_attention_reasons(item)
        if reasons:
            items.append(
                {
                    "type": "Számlasor",
                    "name": item.description or item.invoice_number or "Névtelen számlasor",
                    "reasons": reasons,
                    "project": item.assigned_project.code if item.assigned_project else "-",
                    "supplier": item.partner or "-",
                    "invoice": item.invoice_number or "-",
                    "source": f"{item.source_sheet or '-'} / {item.source_row_number or '-'}",
                    "detail_url": None,
                    "edit_url": url_for("unassigned_invoice_edit", item_id=item.id),
                }
            )
    return items


def build_project_pdf(project, inventory_rows, pdf_type):
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=1.2 * cm,
        leftMargin=1.2 * cm,
        topMargin=1.2 * cm,
        bottomMargin=1.2 * cm,
    )
    styles = unicode_pdf_styles()
    story = []
    titles = {
        "equipment": "Projekt eszközlista",
        "issue": "Kiadási lista",
        "installation": "Telepítési lista",
        "finance": "Pénzügyi összesítő",
    }
    story.append(Paragraph(pdf_escape(titles[pdf_type]), styles["Title"]))
    story.append(Paragraph(pdf_escape(f"{project.code} - {project.name}"), styles["Heading2"]))
    story.append(Paragraph(pdf_escape(f"Ügyfél: {project.customer or '-'}"), styles["Normal"]))
    site_parts = [
        project.site_name,
        project.address,
        project.city,
        project.country,
    ]
    site_text = ", ".join(str(part) for part in site_parts if part) or "-"
    story.append(Paragraph(pdf_escape(f"Helyszín: {site_text}"), styles["Normal"]))
    if project.latitude is not None and project.longitude is not None:
        story.append(
            Paragraph(
                pdf_escape(f"GPS: {project.latitude}, {project.longitude}"),
                styles["Normal"],
            )
        )
    if project.google_maps_url:
        story.append(
            Paragraph(
                pdf_escape(f"Google Maps: {project.google_maps_url}"),
                styles["Normal"],
            )
        )
    if project.site_notes:
        story.append(
            Paragraph(
                pdf_escape(f"Helyszín megjegyzés: {project.site_notes}"),
                styles["Normal"],
            )
        )
    story.append(Paragraph(pdf_escape(f"Dátum: {date.today().isoformat()}"), styles["Normal"]))
    story.append(Spacer(1, 0.4 * cm))

    if pdf_type == "equipment":
        rows = [
            [
                "Tétel",
                "Mennyiség",
                "Deviza",
                "Egység nettó",
                "Összes nettó",
                "ÁFA %",
                "Egység bruttó",
                "Összes bruttó",
                "Státusz",
                "Lokáció",
                "Megjegyzés",
            ]
        ]
        pdf_rows = inventory_rows
        for item in pdf_rows:
            device = item["device"]
            unit = item["unit"]
            quantity = item["quantity"]
            rows.append(
                [
                    unit.unit_code if unit else device_primary_label(device),
                    format_number(quantity),
                    device.currency or "hiányzik",
                    device_money_text(device, "unit_net"),
                    (
                        f"{format_number(device.unit_net_price * quantity)} {device.currency}"
                        if device.unit_net_price is not None and device.currency
                        else "Nem számolható"
                    ),
                    format_vat_rate(device.vat_rate),
                    device_money_text(device, "unit_gross"),
                    (
                        f"{format_number(device.unit_gross_price * quantity)} {device.currency}"
                        if device.unit_gross_price is not None and device.currency
                        else "Nem számolható"
                    ),
                    status_label(item["status"]),
                    item["location"].name if item["location"] else "-",
                    device.assignment_notes or device.subtype_note or "-",
                ]
            )
    elif pdf_type in {"issue", "installation"}:
        wanted_status = "ISSUED" if pdf_type == "issue" else "INSTALLED"
        rows = [["Azonosító", "Termék", "Mennyiség", "Lokáció", "Megjegyzés"]]
        pdf_rows = [item for item in inventory_rows if item["status"] == wanted_status]
        for item in pdf_rows:
            device = item["device"]
            unit = item["unit"]
            rows.append(
                [
                    unit.unit_code if unit else device.asset_tag or "-",
                    device.product_name or device.model or "-",
                    format_number(item["quantity"]),
                    item["location"].name if item["location"] else "-",
                    device.assignment_notes or device.subtype_note or "-",
                ]
            )
    else:
        rows = [["Beszállító", "Számlaszám", "Fizetve", "Deviza", "Egység nettó", "Összes nettó", "ÁFA %", "Egység bruttó", "Összes bruttó", "Tétel"]]
        pdf_rows = inventory_rows
        for item in pdf_rows:
            device = item["device"]
            quantity = item["quantity"]
            invoice_number = device.supplier_invoice_number or device.shipping_invoice_number or "-"
            paid = "Igen" if (device.supplier_invoice_paid or device.shipping_invoice_paid) else "Nem"
            rows.append(
                [
                    device.supplier_manufacturer or device.manufacturer or "-",
                    invoice_number,
                    paid,
                    device.currency or "hiányzik",
                    device_money_text(device, "unit_net"),
                    (
                        f"{format_number(device.unit_net_price * quantity)} {device.currency}"
                        if device.unit_net_price is not None and device.currency
                        else "Nem számolható"
                    ),
                    format_vat_rate(device.vat_rate),
                    device_money_text(device, "unit_gross"),
                    (
                        f"{format_number(device.unit_gross_price * quantity)} {device.currency}"
                        if device.unit_gross_price is not None and device.currency
                        else "Nem számolható"
                    ),
                    device.product_name or device.model or device.asset_tag or "-",
                ]
            )
        totals = inventory_rows_currency_totals(inventory_rows)
        unpaid_count = len(
            {
                item["device"].id
                for item in inventory_rows
                if is_financially_open(item["device"])
            }
        )
        story.append(
            Paragraph(
                pdf_escape(
                    "Összes nettó projektérték: "
                    f"{format_number(totals['net_huf'])} HUF, "
                    f"{format_number(totals['net_eur'])} EUR; "
                    f"nyitott számlás tételek: {unpaid_count}"
                ),
                styles["Normal"],
            )
        )
        story.append(Spacer(1, 0.3 * cm))

    if len(rows) == 1:
        rows.append(["Nincs megjeleníthető tétel."] + [""] * (len(rows[0]) - 1))

    table = Table([[pdf_cell(cell, styles) for cell in row] for row in rows], repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e9eef5")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#1d2733")),
                ("FONTNAME", (0, 0), (-1, 0), PDF_FONT_BOLD),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#cbd3df")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f7f9fb")]),
            ]
        )
    )
    story.append(table)

    if pdf_type in {"issue", "installation"}:
        story.append(Spacer(1, 1.2 * cm))
        signature_rows = [
            ["Előkészítette", "Átvette"],
            ["\n\n____________________________", "\n\n____________________________"],
        ]
        signature_table = Table(signature_rows, colWidths=[8 * cm, 8 * cm])
        signature_table.setStyle(TableStyle([("ALIGN", (0, 0), (-1, -1), "CENTER")]))
        story.append(signature_table)

    doc.build(story)
    buffer.seek(0)
    return buffer


def build_device_unit_labels_pdf(device, units, unit_urls):
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=1.2 * cm,
        leftMargin=1.2 * cm,
        topMargin=1.2 * cm,
        bottomMargin=1.2 * cm,
    )
    styles = unicode_pdf_styles()
    rows = []
    current_row = []
    for unit in units:
        qr_buffer = BytesIO()
        qr = qrcode.QRCode(
            version=None,
            error_correction=qrcode.constants.ERROR_CORRECT_M,
            box_size=8,
            border=3,
        )
        qr.add_data(unit_urls[unit.id])
        qr.make(fit=True)
        qr.make_image(fill_color="#21182f", back_color="white").save(qr_buffer, format="PNG")
        qr_buffer.seek(0)
        qr_image = Image(qr_buffer, width=3.3 * cm, height=3.3 * cm)
        title = unit.asset_tag or unit.unit_code
        details = [
            Paragraph(pdf_escape(title), styles["Heading3"]),
            Paragraph(pdf_escape(device.product_name or device.model or device.asset_tag), styles["BodyText"]),
            Paragraph(pdf_escape(f"Példány: {unit.unit_code}"), styles["BodyText"]),
            Paragraph(pdf_escape(f"Sorozatszám: {unit.serial_number or '-'}"), styles["BodyText"]),
            qr_image,
        ]
        current_row.append(details)
        if len(current_row) == 2:
            rows.append(current_row)
            current_row = []
    if current_row:
        current_row.append("")
        rows.append(current_row)

    table = Table(rows, colWidths=[9.1 * cm, 9.1 * cm])
    table.setStyle(
        TableStyle(
            [
                ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#6f42c1")),
                ("INNERGRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#d8d0e5")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("PADDING", (0, 0), (-1, -1), 8),
            ]
        )
    )
    doc.build([table])
    buffer.seek(0)
    return buffer


def build_work_order_pdf(app, work_order):
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=1.4 * cm,
        leftMargin=1.4 * cm,
        topMargin=1.6 * cm,
        bottomMargin=1.8 * cm,
        title=f"Munkalap {work_order.number}",
    )
    styles = unicode_pdf_styles()
    styles["Title"].textColor = colors.HexColor("#5b3f92")
    styles["Heading2"].textColor = colors.HexColor("#5b3f92")
    story = []

    logo_path = os.path.join(app.static_folder, "parkl-logo.png")
    if os.path.exists(logo_path):
        story.append(cropped_pdf_image(logo_path, max_width=3 * cm, max_height=1.8 * cm))
        story.append(Spacer(1, 0.2 * cm))
    story.append(Paragraph("Szerviz megrendelő / Munkalap / Jegyzőkönyv", styles["Title"]))
    story.append(Paragraph("Parkl Digital Technologies Kft. · 1051 Budapest, Arany János utca 15.", styles["Normal"]))
    story.append(Spacer(1, 0.35 * cm))

    story.append(
        work_order_pdf_key_value_table(
            [
                ("Munkalap száma", work_order.number),
                ("Munkalap típusa", work_order_type_label(work_order.work_type)),
                ("Létrehozás dátuma", work_order.created_date),
                ("Munkavégzés dátuma", work_order.work_date),
                ("Státusz", work_order_status_label(work_order.status)),
            ],
            styles,
        )
    )
    story.append(Spacer(1, 0.3 * cm))

    add_work_order_pdf_section(
        story,
        "Ügyfél adatok",
        [
            ("Ügyfél neve", work_order.customer_name),
            ("Cím", work_order.customer_address),
            ("Kapcsolattartó", work_order.contact_name),
            ("Telefonszám", work_order.phone),
            ("E-mail", work_order.email),
        ],
        styles,
    )
    add_work_order_pdf_section(
        story,
        "Helyszín",
        [
            ("Helyszín neve", work_order.site_name),
            ("Cím", work_order.site_address),
            ("Város", work_order.site_city),
            ("Megjegyzés", work_order.site_notes),
        ],
        styles,
    )
    add_work_order_pdf_section(
        story,
        "Készülék adatok",
        [
            ("Gyártó", work_order.device_manufacturer),
            ("Típus", work_order.device_type),
            ("Gyári szám", work_order.device_serial_number),
            ("Vásárlás dátuma", work_order.device_purchase_date),
        ],
        styles,
    )
    add_work_order_pdf_section(
        story,
        "Munkavégzés",
        [
            ("Érkezés időpontja", format_pdf_time(work_order.arrival_time)),
            ("Távozás időpontja", format_pdf_time(work_order.departure_time)),
            ("Helyszínen töltött idő", format_duration(work_order.duration_minutes)),
            ("Munkát végezte", work_order.technician_name),
            ("Második technikus", work_order.second_technician),
            ("Alvállalkozó", work_order.subcontractor),
        ],
        styles,
    )
    add_work_order_pdf_text(story, "Hiba leírása", work_order.fault_description, styles)
    add_work_order_pdf_text(story, "Elvégzett munka", work_order.work_performed, styles)
    add_work_order_pdf_section(
        story,
        "Elszámolás és megjegyzés",
        [
            ("Munka elszámolása", work_order.labor_settlement),
            ("Anyag elszámolása", work_order.material_settlement),
            ("Megjegyzés", work_order.notes),
        ],
        styles,
    )

    story.append(Paragraph("Felhasznált anyagok", styles["Heading2"]))
    material_rows = [["Anyag megnevezése", "Cikkszám", "Mennyiség", "Mértékegység", "Megjegyzés"]]
    for item in work_order.materials:
        material_rows.append(
            [item.name, item.item_number or "-", format_number(item.quantity), item.unit or "-", item.notes or "-"]
        )
    if len(material_rows) == 1:
        material_rows.append(["Nincs rögzített anyag.", "", "", "", ""])
    story.append(work_order_pdf_table(material_rows, styles))
    story.append(Spacer(1, 0.3 * cm))

    story.append(Paragraph("Mérések", styles["Heading2"]))
    measurement_rows = [["Mérés", "Érték", "Mértékegység", "Megjegyzés"]]
    for item in work_order.measurements:
        measurement_rows.append([item.name, item.value or "-", item.unit or "-", item.notes or "-"])
    if len(measurement_rows) == 1:
        measurement_rows.append(["Nincs rögzített mérés.", "", "", ""])
    story.append(work_order_pdf_table(measurement_rows, styles))
    story.append(Spacer(1, 0.4 * cm))

    upload_dir = os.path.join(app.instance_path, WORK_ORDER_UPLOAD_SUBDIR)
    signature_cells = []
    for label, filename in (
        ("Technikus aláírása", work_order.technician_signature_filename),
        ("Ügyfél aláírása", work_order.customer_signature_filename),
    ):
        content = [Paragraph(pdf_escape(label), styles["Normal"])]
        path = os.path.join(upload_dir, filename) if filename else None
        if path and os.path.exists(path) and valid_image_file(path):
            content.append(Image(path, width=6.5 * cm, height=2.4 * cm))
        else:
            content.append(Spacer(1, 2.4 * cm))
        signature_cells.append(content)
    signature_table = Table([signature_cells], colWidths=[8.5 * cm, 8.5 * cm])
    signature_table.setStyle(
        TableStyle(
            [
                ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#cfc6df")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("PADDING", (0, 0), (-1, -1), 8),
            ]
        )
    )
    story.append(signature_table)

    if work_order.photos:
        story.append(PageBreak())
        story.append(Paragraph("Fotódokumentáció", styles["Title"]))
        for photo in work_order.photos:
            path = os.path.join(upload_dir, photo.filename)
            if not os.path.exists(path) or not valid_image_file(path):
                continue
            story.append(Paragraph(pdf_escape(work_order_photo_category_label(photo.category)), styles["Heading2"]))
            story.append(Image(path, width=16 * cm, height=10 * cm, kind="proportional"))
            story.append(Spacer(1, 0.35 * cm))

    def footer(canvas, document):
        canvas.saveState()
        canvas.setFont(PDF_FONT_REGULAR, 8)
        canvas.setFillColor(colors.HexColor("#6e647d"))
        canvas.drawString(1.4 * cm, 0.8 * cm, f"Munkalap: {work_order.number}")
        canvas.drawRightString(A4[0] - 1.4 * cm, 0.8 * cm, f"{document.page}. oldal")
        canvas.restoreState()

    doc.build(story, onFirstPage=footer, onLaterPages=footer)
    buffer.seek(0)
    return buffer


def add_work_order_pdf_section(story, title, rows, styles):
    story.append(Paragraph(pdf_escape(title), styles["Heading2"]))
    story.append(work_order_pdf_key_value_table(rows, styles))
    story.append(Spacer(1, 0.3 * cm))


def add_work_order_pdf_text(story, title, text, styles):
    story.append(Paragraph(pdf_escape(title), styles["Heading2"]))
    story.append(Paragraph(pdf_escape(text or "-").replace("\n", "<br/>"), styles["BodyText"]))
    story.append(Spacer(1, 0.3 * cm))


def work_order_pdf_key_value_table(rows, styles):
    data = [
        [pdf_cell(label, styles), pdf_cell(value if value not in (None, "") else "-", styles)]
        for label, value in rows
    ]
    table = Table(data, colWidths=[5 * cm, 12 * cm])
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#f2eef8")),
                ("FONTNAME", (0, 0), (0, -1), PDF_FONT_BOLD),
                ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#d8d0e5")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("PADDING", (0, 0), (-1, -1), 5),
            ]
        )
    )
    return table


def work_order_pdf_table(rows, styles):
    table = Table([[pdf_cell(cell, styles) for cell in row] for row in rows], repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#5b3f92")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), PDF_FONT_BOLD),
                ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#d8d0e5")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#faf8fd")]),
                ("PADDING", (0, 0), (-1, -1), 5),
            ]
        )
    )
    return table


def format_pdf_time(value):
    return value.strftime("%H:%M") if value else "-"


def unicode_pdf_styles():
    register_pdf_fonts()
    styles = getSampleStyleSheet()
    for style in styles.byName.values():
        style.fontName = PDF_FONT_REGULAR
        if style.name in {"Title", "Heading1", "Heading2", "Heading3", "Heading4"}:
            style.fontName = PDF_FONT_BOLD
    return styles


def register_pdf_fonts():
    if PDF_FONT_REGULAR in pdfmetrics.getRegisteredFontNames():
        return
    font_dir = os.path.join(os.path.dirname(reportlab.__file__), "fonts")
    regular_path = first_existing_path(
        [
            os.environ.get("PDF_FONT_REGULAR_PATH"),
            "/System/Library/Fonts/Supplemental/Arial.ttf",
            "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
            os.path.join(font_dir, "Vera.ttf"),
        ]
    )
    bold_path = first_existing_path(
        [
            os.environ.get("PDF_FONT_BOLD_PATH"),
            "/System/Library/Fonts/Supplemental/Arial Bold.ttf",
            "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
            os.path.join(font_dir, "VeraBd.ttf"),
        ]
    )
    pdfmetrics.registerFont(TTFont(PDF_FONT_REGULAR, regular_path))
    pdfmetrics.registerFont(TTFont(PDF_FONT_BOLD, bold_path))
    pdfmetrics.registerFontFamily(
        PDF_FONT_REGULAR,
        normal=PDF_FONT_REGULAR,
        bold=PDF_FONT_BOLD,
        italic=PDF_FONT_REGULAR,
        boldItalic=PDF_FONT_BOLD,
    )


def first_existing_path(paths):
    for path in paths:
        if path and os.path.exists(path):
            return path
    raise RuntimeError("Nem található használható PDF betűkészlet.")


def cropped_pdf_image(path, max_width, max_height):
    with PILImage.open(path) as source:
        source.load()
        if source.mode in {"RGBA", "LA"}:
            alpha = source.getchannel("A")
            bounding_box = alpha.getbbox()
        else:
            bounding_box = source.getbbox()
        cropped = source.crop(bounding_box) if bounding_box else source.copy()
        image_buffer = BytesIO()
        cropped.save(image_buffer, format="PNG")
        image_buffer.seek(0)
        width, height = cropped.size
    scale = min(max_width / width, max_height / height)
    return Image(image_buffer, width=width * scale, height=height * scale)


def pdf_escape(value):
    text = "" if value is None else str(value)
    return text.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def pdf_cell(value, styles):
    return Paragraph(pdf_escape(value), styles["BodyText"])


def now_utc():
    return datetime.now(timezone.utc)


def qr_png_response(target_url, filename):
    qr = qrcode.QRCode(
        version=None,
        error_correction=qrcode.constants.ERROR_CORRECT_M,
        box_size=10,
        border=4,
    )
    qr.add_data(target_url)
    qr.make(fit=True)
    image = qr.make_image(fill_color="#21182f", back_color="white")
    buffer = BytesIO()
    image.save(buffer, format="PNG")
    buffer.seek(0)
    return send_file(
        buffer,
        mimetype="image/png",
        as_attachment=request.args.get("download") == "1",
        download_name=secure_filename(filename),
    )


def device_form_data(form):
    quantity = optional_float(form.get("quantity"))
    unit_net_price = optional_float(form.get("unit_net_price"))
    currency = form.get("currency", "").strip().upper() or None
    huf_value = optional_float(form.get("huf_value"))
    return {
        "asset_tag": form.get("asset_tag", "").strip(),
        "serial_number": form.get("serial_number", "").strip(),
        "device_type": form.get("device_type", "").strip(),
        "manufacturer": form.get("manufacturer", "").strip(),
        "model": form.get("model", "").strip(),
        "product_name": form.get("product_name", "").strip() or None,
        "subtype_note": form.get("subtype_note", "").strip() or None,
        "supplier_manufacturer": form.get("supplier_manufacturer", "").strip() or None,
        "version": form.get("version", "").strip() or None,
        "quantity": quantity,
        "unit_net_price": unit_net_price,
        "currency": currency,
        "vat_rate": optional_float(form.get("vat_rate")),
        "qr_mode": form.get("qr_mode", "group").strip() or "group",
        "tracking_mode": form.get("tracking_mode", "bulk").strip() or "bulk",
        "huf_value": huf_value
        if huf_value is not None
        else calculate_huf_value(quantity, unit_net_price, currency),
        "assignment_quantity": optional_float(form.get("assignment_quantity")),
        "assignment_notes": form.get("assignment_notes", "").strip() or None,
        "order_date": optional_date(form.get("order_date")),
        "is_ordered": checkbox_value(form.get("is_ordered")),
        "planned_arrival_date": optional_date(form.get("planned_arrival_date")),
        "actual_arrival_date": optional_date(form.get("actual_arrival_date")),
        "has_arrived": checkbox_value(form.get("has_arrived")),
        "shipping_cost": optional_float(form.get("shipping_cost")),
        "shipping_date": optional_date(form.get("shipping_date")),
        "supplier_invoice_number": form.get("supplier_invoice_number", "").strip() or None,
        "supplier_invoice_paid": checkbox_value(form.get("supplier_invoice_paid")),
        "invoice_value": optional_float(form.get("invoice_value")),
        "shipping_invoice_number": form.get("shipping_invoice_number", "").strip() or None,
        "shipping_invoice_paid": checkbox_value(form.get("shipping_invoice_paid")),
        "project_id": optional_int(form.get("project_id")),
        "location_id": optional_int(form.get("location_id")),
    }


def update_unassigned_invoice_from_form(item, form):
    assigned_project_id = optional_int(form.get("assigned_project_id"))
    assigned_device_id = optional_int(form.get("assigned_device_id"))
    responsible_user_id = optional_int(form.get("responsible_user_id"))
    assignment_status = form.get("assignment_status", "unassigned").strip()
    if assignment_status not in ASSIGNMENT_STATUS_LABELS:
        assignment_status = "unassigned"
    if assigned_project_id or assigned_device_id:
        assignment_status = "assigned"
    item.invoice_number = form.get("invoice_number", "").strip() or None
    item.partner = form.get("partner", "").strip() or None
    item.invoice_date = optional_date(form.get("invoice_date"))
    item.accounting_fulfillment_date = optional_date(
        form.get("accounting_fulfillment_date")
    )
    item.payment_deadline = optional_date(form.get("payment_deadline"))
    item.gross_amount_huf = optional_float(form.get("gross_amount_huf"))
    item.currency = form.get("currency", "").strip().upper() or None
    item.description = form.get("description", "").strip() or None
    item.quantity = optional_float(form.get("quantity"))
    item.unit_price_huf = optional_float(form.get("unit_price_huf"))
    net_amount_huf = optional_float(form.get("net_amount_huf"))
    item.net_amount_huf = (
        net_amount_huf
        if net_amount_huf is not None
        else calculate_line_net_amount(item.quantity, item.unit_price_huf)
    )
    item.vat_amount_huf = optional_float(form.get("vat_amount_huf"))
    item.line_gross_amount_huf = optional_float(form.get("line_gross_amount_huf"))
    item.assignment_status = assignment_status
    item.notes = form.get("notes", "").strip() or None
    item.assigned_project_id = assigned_project_id
    item.assigned_device_id = assigned_device_id
    item.responsible_user_id = responsible_user_id


def build_import_template_workbook():
    workbook = Workbook()
    projects = workbook.active
    projects.title = "Projects"
    projects.append(TEMPLATE_PROJECT_HEADERS)
    projects.append([
        "PRK-100",
        "Minta EV projekt",
        "Minta Ügyfél Kft.",
        "Minta helyszín",
        "Minta utca 1.",
        "Budapest",
        "Magyarország",
        47.4979,
        19.0402,
        "https://maps.google.com/?q=47.4979,19.0402",
        "Bejárat a főkapu felől.",
        "active",
        "Példasor, import előtt törölhető.",
    ])

    devices = workbook.create_sheet("Devices")
    devices.append(TEMPLATE_DEVICE_HEADERS)
    devices.append([
        "PRK-100", "EV charger", "Schneider EVlink Pro AC", "Schneider",
        "EVB3S22N4", "", "EV-MINTA-001", 2, "HUF", 250000, 500000, 27,
        317500, 635000, "Fő raktár", "IN_STOCK", "unit", "yes",
        "EV-MINTA", "Példasor, import előtt törölhető.",
    ])

    locations = workbook.create_sheet("Locations")
    locations.append(TEMPLATE_LOCATION_HEADERS)
    locations.append(["Fő raktár", "warehouse", "Budapest", "Példa készlethely."])

    instructions = workbook.create_sheet("Instructions")
    instructions.append(["Parkl Infra Manager import sablon"])
    instructions.append(["A Projects és Devices munkalap használható. A Locations munkalap opcionális."])
    instructions.append(["Kötelező Project mezők: project_code, project_name új projekt esetén."])
    instructions.append(["A Projects site_name/address/city/country és térképes mezői közvetlenül a projekthez tartoznak; nem hoznak létre készlethelyet."])
    instructions.append(["Kötelező Device mezők: project_code, product_name, quantity, currency."])
    instructions.append(["tracking_mode: bulk vagy unit. Ha üres, az alapértelmezés bulk."])
    instructions.append(["unit_generation: unit követésnél yes esetén quantity darab DeviceUnit jön létre."])
    instructions.append(["unit_code_prefix: opcionális példányazonosító prefix, például EV-MINTA."])
    instructions.append(["Elfogadott deviza: HUF, EUR. Elfogadott Device státuszok: " + ", ".join(sorted(STATUS_LABELS))])
    instructions.append(["Elfogadott projekt státuszok: " + ", ".join(sorted(PROJECT_STATUS_LABELS))])
    instructions.append(["A példa sorokat import előtt töröld vagy írd át."])

    style_workbook_headers(workbook)
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def style_workbook_headers(workbook):
    from openpyxl.styles import Font, PatternFill

    for sheet in workbook.worksheets:
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="5B3F92")
        sheet.freeze_panes = "A2"
        for column in sheet.columns:
            width = max(len(str(cell.value or "")) for cell in column) + 2
            sheet.column_dimensions[column[0].column_letter].width = min(max(width, 14), 34)


def build_data_export_workbook(export_type, Project, Device, Location):
    workbook = Workbook()
    sheet = workbook.active
    if export_type == "projects":
        sheet.title = "Projects"
        sheet.append(TEMPLATE_PROJECT_HEADERS)
        for project in Project.query.filter(Project.archived_at.is_(None)).order_by(Project.code).all():
            sheet.append([
                project.code,
                project.name,
                project.customer,
                project.site_name,
                project.address,
                project.city,
                project.country,
                project.latitude,
                project.longitude,
                project.google_maps_url,
                project.site_notes,
                project.status,
                project.notes,
            ])
    elif export_type == "locations":
        sheet.title = "Locations"
        sheet.append(TEMPLATE_LOCATION_HEADERS)
        for location in (
            Location.query.filter(
                Location.archived_at.is_(None),
                Location.location_type.in_(LOGISTIC_LOCATION_TYPES),
            )
            .order_by(Location.name)
            .all()
        ):
            sheet.append([location.name, location.location_type, location.address, location.notes])
    else:
        sheet.title = "Devices"
        sheet.append(TEMPLATE_DEVICE_HEADERS)
        for device in Device.query.filter(Device.archived_at.is_(None)).order_by(Device.asset_tag).all():
            sheet.append([
                device_inventory_export_value(device, "project"),
                device.device_type,
                device.product_name or "",
                device.manufacturer,
                device.model,
                device.serial_number,
                device.asset_tag,
                device.quantity,
                device.currency,
                device.unit_net_price,
                device.total_net_price,
                device.vat_rate,
                device.unit_gross_price,
                device.total_gross_price,
                device_inventory_export_value(device, "location"),
                device_inventory_export_value(device, "status"),
                device.tracking_mode,
                "yes" if device.tracking_mode == "unit" else "no",
                default_unit_code_prefix(device) if device.tracking_mode == "unit" else "",
                device.subtype_note or "",
            ])
    style_workbook_headers(workbook)
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def parse_template_workbook(path, Project, Device, Location):
    workbook = load_workbook(path, read_only=True, data_only=True)
    summary = {
        "projects": [],
        "devices": [],
        "locations": [],
        "errors": [],
        "new_project_count": 0,
        "existing_project_count": 0,
        "new_device_count": 0,
        "new_unit_count": 0,
        "new_location_count": 0,
        "critical_error_count": 0,
    }
    sheets = {sheet.title.lower(): sheet for sheet in workbook.worksheets}
    project_sheet = sheets.get("projects")
    device_sheet = sheets.get("devices")
    location_sheet = sheets.get("locations")
    if project_sheet is None:
        summary["errors"].append({"sheet": "Projects", "row": "-", "message": "Hiányzik a Projects munkalap."})
    if device_sheet is None:
        summary["errors"].append({"sheet": "Devices", "row": "-", "message": "Hiányzik a Devices munkalap."})

    existing_projects = {project.code: project for project in Project.query.all()}
    existing_locations = {
        location.name.lower(): location
        for location in Location.query.filter(
            Location.archived_at.is_(None),
            Location.location_type.in_(LOGISTIC_LOCATION_TYPES),
        ).all()
    }
    existing_asset_tags = {value for (value,) in db.session.query(Device.asset_tag).all() if value}
    existing_serials = {value for (value,) in db.session.query(Device.serial_number).all() if value}
    seen_project_codes, seen_asset_tags, seen_serials, seen_locations = set(), set(), set(), set()

    if project_sheet:
        rows, errors = template_sheet_rows(
            project_sheet,
            TEMPLATE_PROJECT_HEADERS,
            optional_headers=OPTIONAL_TEMPLATE_PROJECT_HEADERS,
        )
        summary["errors"].extend(errors)
        for row_number, row in rows:
            code = clean_string(row.get("project_code"))
            name = clean_string(row.get("project_name"))
            if not code:
                add_template_error(summary, "Projects", row_number, "Hiányzó project_code.")
                continue
            if code in seen_project_codes:
                add_template_error(summary, "Projects", row_number, f"Duplikált project_code a fájlban: {code}.")
                continue
            seen_project_codes.add(code)
            if code not in existing_projects and not name:
                add_template_error(summary, "Projects", row_number, "Hiányzó project_name új projektnél.")
                continue
            status = clean_string(row.get("status")) or "planned"
            if status not in PROJECT_STATUS_LABELS:
                add_template_error(summary, "Projects", row_number, f"Ismeretlen projekt státusz: {status}.")
                continue
            parsed = {key: clean_string(value) for key, value in row.items()}
            latitude = optional_decimal(row.get("latitude"))
            longitude = optional_decimal(row.get("longitude"))
            if meaningful_value(row.get("latitude")) and latitude is None:
                add_template_error(summary, "Projects", row_number, "Hibás latitude érték.")
            elif latitude is not None and not Decimal("-90") <= latitude <= Decimal("90"):
                add_template_error(summary, "Projects", row_number, "A latitude -90 és 90 közötti lehet.")
            if meaningful_value(row.get("longitude")) and longitude is None:
                add_template_error(summary, "Projects", row_number, "Hibás longitude érték.")
            elif longitude is not None and not Decimal("-180") <= longitude <= Decimal("180"):
                add_template_error(summary, "Projects", row_number, "A longitude -180 és 180 közötti lehet.")
            if any(
                error["sheet"] == "Projects" and error["row"] == row_number
                for error in summary["errors"]
            ):
                continue
            parsed.update({"project_code": code, "project_name": name, "status": status})
            parsed["latitude"] = latitude
            parsed["longitude"] = longitude
            summary["projects"].append(parsed)
            if code in existing_projects:
                summary["existing_project_count"] += 1
            else:
                summary["new_project_count"] += 1

    project_rows_by_code = {row["project_code"]: row for row in summary["projects"]}

    if location_sheet:
        rows, errors = template_sheet_rows(location_sheet, TEMPLATE_LOCATION_HEADERS)
        summary["errors"].extend(errors)
        for row_number, row in rows:
            name = clean_string(row.get("location_name"))
            if not name:
                add_template_error(summary, "Locations", row_number, "Hiányzó location_name.")
                continue
            key = name.lower()
            if key in seen_locations:
                add_template_error(summary, "Locations", row_number, f"Duplikált location_name a fájlban: {name}.")
                continue
            seen_locations.add(key)
            parsed = {field: clean_string(value) for field, value in row.items()}
            parsed["location_name"] = name
            parsed["location_type"] = parsed.get("location_type") or "warehouse"
            if parsed["location_type"] not in LOGISTIC_LOCATION_TYPES:
                add_template_error(
                    summary,
                    "Locations",
                    row_number,
                    "Ismeretlen vagy nem logisztikai location_type: "
                    f"{parsed['location_type']}.",
                )
                continue
            summary["locations"].append(parsed)
            if key not in existing_locations:
                summary["new_location_count"] += 1

    if device_sheet:
        rows, errors = template_sheet_rows(
            device_sheet,
            TEMPLATE_DEVICE_HEADERS,
            optional_headers=OPTIONAL_TEMPLATE_DEVICE_HEADERS,
        )
        summary["errors"].extend(errors)
        for row_number, row in rows:
            parsed = parse_template_device_row(
                row_number, row, existing_projects, project_rows_by_code,
                existing_asset_tags, existing_serials, seen_asset_tags, seen_serials, summary
            )
            if parsed:
                summary["devices"].append(parsed)
                summary["new_device_count"] += 1
                if parsed["tracking_mode"] == "unit":
                    summary["new_unit_count"] += int(parsed["quantity"])
                location_name = parsed.get("location_name")
                if location_name and location_name.lower() not in existing_locations and location_name.lower() not in seen_locations:
                    seen_locations.add(location_name.lower())
                    summary["new_location_count"] += 1

    used_project_codes = {
        row["project_code"] for row in summary["projects"]
    } | {
        row["project_code"] for row in summary["devices"]
    }
    summary["existing_project_count"] = sum(
        1 for code in used_project_codes if code in existing_projects
    )
    summary["new_project_count"] = sum(
        1 for code in used_project_codes if code not in existing_projects
    )
    summary["critical_error_count"] = len(summary["errors"])
    return summary


def template_sheet_rows(sheet, expected_headers, optional_headers=None):
    optional_headers = optional_headers or set()
    values = list(sheet.iter_rows(values_only=True))
    if not values:
        return [], [{"sheet": sheet.title, "row": "-", "message": "Üres munkalap."}]
    headers = [clean_string(value) or "" for value in values[0]]
    missing = [
        header
        for header in expected_headers
        if header not in headers and header not in optional_headers
    ]
    errors = [{"sheet": sheet.title, "row": 1, "message": f"Hiányzó oszlop: {header}."} for header in missing]
    rows = []
    for row_number, values_row in enumerate(values[1:], start=2):
        row = {header: values_row[index] if index < len(values_row) else None for index, header in enumerate(headers) if header}
        if any(meaningful_value(value) for value in row.values()):
            rows.append((row_number, row))
    return rows, errors


def add_template_error(summary, sheet, row, message):
    summary["errors"].append({"sheet": sheet, "row": row, "message": message})


def parse_template_boolean(value):
    if not meaningful_value(value):
        return False
    normalized = clean_string(value)
    if normalized is None:
        return False
    normalized = normalized.lower()
    if normalized in {"yes", "true", "1", "igen", "i"}:
        return True
    if normalized in {"no", "false", "0", "nem", "n"}:
        return False
    return None


def parse_template_device_row(row_number, row, existing_projects, project_rows_by_code, existing_asset_tags, existing_serials, seen_asset_tags, seen_serials, summary):
    from models import TRACKING_MODES

    project_code = clean_string(row.get("project_code"))
    product_name = clean_string(row.get("product_name"))
    currency = (clean_string(row.get("currency")) or "").upper()
    quantity = number_value(row.get("quantity"))
    tracking_mode = (clean_string(row.get("tracking_mode")) or "bulk").lower()
    unit_generation = parse_template_boolean(row.get("unit_generation"))
    if not project_code:
        add_template_error(summary, "Devices", row_number, "Hiányzó project_code.")
    elif project_code not in existing_projects and project_code not in project_rows_by_code:
        add_template_error(summary, "Devices", row_number, f"A projekt nem létezik és nincs a Projects lapon: {project_code}.")
    if not product_name:
        add_template_error(summary, "Devices", row_number, "Hiányzó product_name.")
    if not meaningful_value(row.get("quantity")):
        add_template_error(summary, "Devices", row_number, "Hiányzó quantity.")
    elif quantity is None or quantity <= 0:
        add_template_error(summary, "Devices", row_number, "Hibás quantity.")
    if not currency:
        add_template_error(summary, "Devices", row_number, "Hiányzó currency.")
    elif currency not in DEVICE_CURRENCIES:
        add_template_error(summary, "Devices", row_number, f"Ismeretlen currency érték: {currency}.")
    if tracking_mode not in TRACKING_MODES:
        add_template_error(
            summary,
            "Devices",
            row_number,
            f"Ismeretlen tracking_mode érték: {tracking_mode}.",
        )
    if meaningful_value(row.get("unit_generation")) and unit_generation is None:
        add_template_error(
            summary,
            "Devices",
            row_number,
            "A unit_generation értéke yes/no, igen/nem vagy true/false lehet.",
        )
    if tracking_mode == "unit":
        if quantity is not None and not float(quantity).is_integer():
            add_template_error(
                summary,
                "Devices",
                row_number,
                "Unit követésnél a quantity csak pozitív egész szám lehet.",
            )
        if unit_generation is not True:
            add_template_error(
                summary,
                "Devices",
                row_number,
                "Unit követésű importnál unit_generation=yes szükséges.",
            )
    elif unit_generation is True:
        add_template_error(
            summary,
            "Devices",
            row_number,
            "Bulk követésnél nem kérhető unit példánygenerálás.",
        )

    numeric_fields = {}
    for field in ["unit_net_price", "total_net_price", "vat_rate", "unit_gross_price", "total_gross_price"]:
        raw = row.get(field)
        value = number_value(raw)
        if meaningful_value(raw) and value is None:
            add_template_error(summary, "Devices", row_number, f"Nem numerikus ár mező: {field}.")
        numeric_fields[field] = value

    asset_tag = clean_string(row.get("asset_tag"))
    serial_number = clean_string(row.get("serial_number"))
    if asset_tag and (asset_tag in existing_asset_tags or asset_tag in seen_asset_tags):
        add_template_error(summary, "Devices", row_number, f"Duplikált asset_tag: {asset_tag}.")
    if serial_number and (serial_number in existing_serials or serial_number in seen_serials):
        add_template_error(summary, "Devices", row_number, f"Duplikált serial_number: {serial_number}.")
    if asset_tag:
        seen_asset_tags.add(asset_tag)
    if serial_number:
        seen_serials.add(serial_number)

    status = clean_string(row.get("status")) or "IN_STOCK"
    if status not in STATUS_LABELS:
        add_template_error(summary, "Devices", row_number, f"Ismeretlen status érték: {status}.")
    category = clean_string(row.get("category")) or "Other"
    if category not in CATEGORY_LABELS:
        add_template_error(summary, "Devices", row_number, f"Ismeretlen category érték: {category}.")

    unit_net = numeric_fields["unit_net_price"]
    total_net = numeric_fields["total_net_price"]
    if quantity and unit_net is not None and total_net is None:
        total_net = quantity * unit_net
    elif quantity and total_net is not None and unit_net is None:
        unit_net = total_net / quantity
    elif quantity and unit_net is not None and total_net is not None and abs(quantity * unit_net - total_net) > 0.01:
        add_template_error(summary, "Devices", row_number, "A quantity × unit_net_price nem egyezik a total_net_price értékkel.")

    vat_rate = numeric_fields["vat_rate"]
    unit_gross = unit_net * (1 + vat_rate / 100) if unit_net is not None and vat_rate is not None else None
    total_gross = total_net * (1 + vat_rate / 100) if total_net is not None and vat_rate is not None else None
    if numeric_fields["unit_gross_price"] is not None and unit_gross is not None and abs(numeric_fields["unit_gross_price"] - unit_gross) > 0.01:
        add_template_error(summary, "Devices", row_number, "A unit_gross_price nem egyezik a nettó érték és ÁFA alapján számolt értékkel.")
    if numeric_fields["total_gross_price"] is not None and total_gross is not None and abs(numeric_fields["total_gross_price"] - total_gross) > 0.01:
        add_template_error(summary, "Devices", row_number, "A total_gross_price nem egyezik a nettó érték és ÁFA alapján számolt értékkel.")

    if any(error["sheet"] == "Devices" and error["row"] == row_number for error in summary["errors"]):
        return None
    return {
        "project_code": project_code,
        "category": category,
        "product_name": product_name,
        "manufacturer": clean_string(row.get("manufacturer")) or "",
        "model": clean_string(row.get("model")) or "",
        "serial_number": serial_number or "",
        "asset_tag": asset_tag,
        "quantity": quantity,
        "currency": currency,
        "unit_net_price": unit_net,
        "total_net_price": total_net,
        "vat_rate": vat_rate,
        "location_name": clean_string(row.get("location_name")),
        "status": status,
        "tracking_mode": tracking_mode,
        "unit_generation": unit_generation is True,
        "unit_code_prefix": clean_string(row.get("unit_code_prefix")),
        "notes": clean_string(row.get("notes")),
    }


def import_template_workbook(summary, Project, Device, Location, user_id):
    from models import DeviceUnit

    projects = {project.code: project for project in Project.query.all()}
    locations = {
        location.name.lower(): location
        for location in Location.query.filter(
            Location.archived_at.is_(None),
            Location.location_type.in_(LOGISTIC_LOCATION_TYPES),
        ).all()
    }
    result = {
        "projects_created": 0,
        "locations_created": 0,
        "devices_created": 0,
        "units_created": 0,
    }
    for row in summary["projects"]:
        if row["project_code"] in projects:
            continue
        project = Project(
            code=row["project_code"],
            name=row["project_name"],
            customer=row.get("customer_name") or "",
            site_name=row.get("site_name") or None,
            address=row.get("address") or None,
            city=row.get("city") or None,
            country=row.get("country") or None,
            latitude=row.get("latitude"),
            longitude=row.get("longitude"),
            google_maps_url=row.get("google_maps_url") or None,
            site_notes=row.get("site_notes") or None,
            status=row.get("status") or "planned",
            notes=row.get("notes") or "",
        )
        db.session.add(project)
        db.session.flush()
        projects[project.code] = project
        result["projects_created"] += 1
    for row in summary["locations"]:
        key = row["location_name"].lower()
        if key in locations:
            continue
        location = Location(name=row["location_name"], location_type=row.get("location_type") or "warehouse", address=row.get("address") or "", notes=row.get("notes") or "")
        db.session.add(location)
        db.session.flush()
        locations[key] = location
        result["locations_created"] += 1
    for row in summary["devices"]:
        location = None
        if row.get("location_name"):
            key = row["location_name"].lower()
            location = locations.get(key)
            if location is None:
                location = Location(name=row["location_name"], location_type="warehouse")
                db.session.add(location)
                db.session.flush()
                locations[key] = location
                result["locations_created"] += 1
        if location is None:
            location = locations.get("fő raktár")
            if location is None:
                location = Location(name="Fő raktár", location_type="warehouse")
                db.session.add(location)
                db.session.flush()
                locations["fő raktár"] = location
                result["locations_created"] += 1
        project = projects[row["project_code"]]
        asset_tag = row.get("asset_tag") or unique_import_asset_tag(Device)
        device = Device(
            asset_tag=asset_tag,
            serial_number=row.get("serial_number") or "",
            device_type=row.get("category") or "Other",
            manufacturer=row.get("manufacturer") or "",
            model=row.get("model") or "",
            product_name=row.get("product_name"),
            subtype_note=row.get("notes"),
            quantity=row.get("quantity"),
            currency=row.get("currency"),
            unit_net_price=row.get("unit_net_price"),
            huf_value=row.get("total_net_price") if row.get("currency") == "HUF" else None,
            vat_rate=row.get("vat_rate"),
            tracking_mode=row.get("tracking_mode") or "bulk",
            qr_mode="individual" if row.get("tracking_mode") == "unit" else "group",
            status="IN_STOCK",
            location=location if row.get("tracking_mode") != "unit" else None,
        )
        db.session.add(device)
        db.session.flush()
        if device.tracking_mode == "unit":
            unit_count = int(device.quantity)
            prefix = row.get("unit_code_prefix") or default_unit_code_prefix(device)
            unit_codes = available_unit_codes(DeviceUnit, prefix, 1, unit_count)
            for unit_code in unit_codes:
                unit = DeviceUnit(
                    device=device,
                    unit_code=unit_code,
                    status="IN_STOCK",
                    location_id=location.id,
                )
                db.session.add(unit)
                db.session.flush()
                create_movement(
                    device=device,
                    unit=unit,
                    movement_type="INBOUND",
                    quantity=1,
                    to_location_id=location.id,
                    notes="Sablon alapú unit import.",
                    user_id=user_id,
                )
                apply_imported_device_status(
                    device,
                    row.get("status") or "IN_STOCK",
                    user_id,
                    project_id=project.id,
                    stock_location_id=location.id,
                    unit=unit,
                )
                result["units_created"] += 1
        else:
            create_movement(
                device=device,
                movement_type="INBOUND",
                quantity=device.quantity,
                to_location_id=location.id,
                notes="Sablon alapú bulk import.",
                user_id=user_id,
            )
            apply_device_state(
                device,
                "INBOUND",
                location.id,
                None,
                quantity=device.quantity,
            )
            apply_imported_device_status(
                device,
                row.get("status") or "IN_STOCK",
                user_id,
                project_id=project.id,
                stock_location_id=location.id,
            )
        result["devices_created"] += 1
    return result


def apply_imported_device_status(
    device,
    target_status,
    user_id,
    project_id=None,
    stock_location_id=None,
    unit=None,
):
    movement_path = {
        "IN_STOCK": [],
        "RESERVED": ["RESERVE"],
        "ISSUED": ["ISSUE"],
        "INSTALLED": ["ISSUE", "INSTALL"],
        "RETURNED": ["ISSUE", "RETURN"],
        "IN_SERVICE": ["SERVICE"],
        "SCRAPPED": ["SCRAP"],
    }
    for movement_type in movement_path.get(target_status, []):
        source_balance = None
        if unit is None:
            source_balance = infer_bulk_source_balance(
                device,
                movement_type,
                device.quantity,
            )
        target_location_id = (
            stock_location_id
            if movement_type in {"RETURN", "SERVICE", "INBOUND", "TRANSFER"}
            else None
        )
        target_project_id = (
            project_id
            if movement_type in {"RESERVE", "ISSUE", "INSTALL"}
            else None
        )
        create_movement(
            device=device,
            unit=unit,
            movement_type=movement_type,
            quantity=1 if unit is not None else device.quantity,
            from_location_id=unit.location_id if unit is not None else (
                source_balance.location_id if source_balance else None
            ),
            to_location_id=target_location_id,
            project_id=target_project_id,
            source_balance=source_balance,
            notes="Sablon alapú import státuszbeállítás.",
            user_id=user_id,
        )
        apply_device_state(
            device,
            movement_type,
            target_location_id,
            target_project_id,
            unit=unit,
            quantity=1 if unit is not None else device.quantity,
            source_balance=source_balance,
        )


def unique_import_asset_tag(Device):
    while True:
        asset_tag = f"IMP-{uuid4().hex[:10].upper()}"
        if not Device.query.filter_by(asset_tag=asset_tag).first():
            return asset_tag


def parse_inventory_workbook(path):
    workbook = load_workbook(path, read_only=True, data_only=True)
    summary = {
        "sheets": [],
        "preview_rows": [],
        "warnings": [],
        "warning_count": 0,
        "total_rows": 0,
        "total_skipped": 0,
        "inventory_rows": [],
        "invoice_rows": [],
    }

    for sheet in workbook.worksheets:
        normalized_sheet = normalize_key(sheet.title)
        if normalized_sheet in IGNORED_IMPORT_SHEETS:
            continue
        if normalized_sheet == ORPHAN_INVOICE_SHEET:
            sheet_summary = parse_unassigned_invoice_sheet(sheet, summary)
        elif normalized_sheet in INVENTORY_SHEETS:
            sheet_summary = parse_inventory_sheet(sheet, summary)
        else:
            sheet_summary = {
                "sheet_name": sheet.title,
                "type": "kihagyva",
                "parsed_count": 0,
                "skipped_count": 0,
                "warnings": [f"A(z) {sheet.title} munkalap importálása nincs beállítva."],
            }
        summary["sheets"].append(sheet_summary)

    summary["warning_count"] = len(summary["warnings"]) + sum(
        len(sheet["warnings"]) for sheet in summary["sheets"]
    )
    return summary


def parse_inventory_sheet(sheet, summary):
    header_row, header_map = find_header_map(sheet, max_scan_rows=12)
    warnings = []
    parsed_count = 0
    skipped_count = 0
    if not header_row:
        warning = f"A(z) {sheet.title} munkalapon nem található fejlécsor."
        summary["warnings"].append(warning)
        return {
            "sheet_name": sheet.title,
            "type": "eszköz",
            "parsed_count": 0,
            "skipped_count": 0,
            "warnings": [warning],
        }

    required_candidates = [
        ("termék vagy típus", ["termek", "tipus", "tolto tipusa", "kamera tipusa"]),
        ("mennyiség", ["mennyiseg"]),
        ("projekt kód", ["projekt kod", "projekt kod"]),
    ]
    for label, candidates in required_candidates:
        if not any(candidate in header_map for candidate in candidates):
            warnings.append(f"Hiányzó vagy eltérő fejléc: {label}.")

    for row_number, row in iter_sheet_rows(sheet, header_row + 1):
        parsed = parse_inventory_row(sheet.title, row_number, row, header_map)
        if parsed is None:
            skipped_count += 1
            continue
        parsed_count += 1
        summary["inventory_rows"].append(parsed)
        if len(summary["preview_rows"]) < 10:
            summary["preview_rows"].append(preview_row("Eszköz", parsed))

    summary["total_rows"] += parsed_count
    summary["total_skipped"] += skipped_count
    return {
        "sheet_name": sheet.title,
        "type": "eszköz",
        "parsed_count": parsed_count,
        "skipped_count": skipped_count,
        "warnings": warnings,
    }


def parse_unassigned_invoice_sheet(sheet, summary):
    header_row, header_map = find_header_map(sheet, max_scan_rows=5)
    warnings = []
    parsed_count = 0
    skipped_count = 0
    if not header_row:
        warning = "A Gazdátlanul munkalapon nem található fejlécsor."
        summary["warnings"].append(warning)
        return {
            "sheet_name": sheet.title,
            "type": "gazdátlan számlasor",
            "parsed_count": 0,
            "skipped_count": 0,
            "warnings": [warning],
        }

    for label in ["szamlaszam", "partner", "megnevezes"]:
        if label not in header_map:
            warnings.append(f"Hiányzó vagy eltérő fejléc: {label}.")

    for row_number, row in iter_sheet_rows(sheet, header_row + 1):
        parsed = parse_unassigned_invoice_row(sheet.title, row_number, row, header_map)
        if parsed is None:
            skipped_count += 1
            continue
        parsed_count += 1
        summary["invoice_rows"].append(parsed)
        if len(summary["preview_rows"]) < 10:
            summary["preview_rows"].append(preview_row("Gazdátlan számlasor", parsed))

    summary["total_rows"] += parsed_count
    summary["total_skipped"] += skipped_count
    return {
        "sheet_name": sheet.title,
        "type": "gazdátlan számlasor",
        "parsed_count": parsed_count,
        "skipped_count": skipped_count,
        "warnings": warnings,
    }


def parse_inventory_row(sheet_name, row_number, row, header_map):
    product_name = first_value(row, header_map, ["termek", "tolto marka"])
    type_value = first_value(row, header_map, ["tipus", "tolto tipusa", "kamera tipusa"])
    subtype_note = first_value(row, header_map, ["altipus", "altipus megjegyzes"])
    row_note = first_value(row, header_map, ["megjegyzes"])
    supplier = first_value(
        row,
        header_map,
        ["beszallito", "gyartasert felelos", "gyarto", "maganszemely neve"],
    )
    quantity = number_value(first_value(row, header_map, ["mennyiseg"]))
    unit_net_price = number_value(first_value(row, header_map, ["netto egysegar"]))
    currency = clean_string(first_value(row, header_map, ["deviza", "penznem"]))
    huf_value = number_value(first_value(row, header_map, ["ertek huf", "ertek"]))
    project_code = clean_project_code(
        first_value(row, header_map, ["projekt kod", "projekt kód"])
    )
    internal_id = clean_string(first_value(row, header_map, ["id", "po szam", "po"]))
    version = clean_string(first_value(row, header_map, ["verzio"]))
    supplier_invoice_number = clean_string(
        first_value(
            row,
            header_map,
            [
                "kapcsolodo szamla sorszama beszallito",
                "kapcsolodo eloleg szamla sorszama beszallito",
                "kapcsolodo elolegszamla szamla sorszama beszallito",
                "kapcsolodo vegszamla szamla sorszama beszallito",
                "kapcsolodo elolegszamla sorszama beszallito",
                "kapcsolodo vegszamla sorszama beszallito",
            ],
        )
    )
    shipping_invoice_number = clean_string(
        first_value(
            row,
            header_map,
            [
                "kapcsolodo szamla sorszama szallitmanyozo",
                "kapcsolodo szallitmanyozoi szamla sorszama",
            ],
        )
    )
    invoice_value = number_value(first_value(row, header_map, ["szamla erteke"]))

    if not any(
        [
            product_name,
            type_value,
            subtype_note,
            supplier,
            quantity,
            project_code,
            supplier_invoice_number,
            shipping_invoice_number,
            internal_id,
        ]
    ):
        return None

    asset_tag = internal_id or f"{normalize_key(sheet_name).replace(' ', '-').upper()}-{row_number}"
    category = infer_device_type(sheet_name, type_value, product_name)
    return {
        "source_sheet": sheet_name,
        "source_row_number": row_number,
        "asset_tag": asset_tag,
        "serial_number": internal_id or "",
        "device_type": category,
        "manufacturer": supplier or "",
        "model": clean_string(product_name or type_value or subtype_note) or "",
        "product_name": clean_string(product_name or type_value),
        "subtype_note": clean_string(subtype_note),
        "supplier_manufacturer": supplier,
        "version": version,
        "quantity": quantity,
        "unit_net_price": unit_net_price,
        "currency": currency.upper() if currency else None,
        "huf_value": calculate_imported_huf_value(
            quantity,
            unit_net_price,
            currency.upper() if currency else None,
            huf_value,
        ),
        "project_code": project_code,
        "notes": clean_string(row_note),
        "order_date": date_value(first_value(row, header_map, ["rendeles napja"])),
        "is_ordered": bool_value(first_value(row, header_map, ["megrendelve"])),
        "planned_arrival_date": date_value(
            first_value(row, header_map, ["tervezett erkezes napja"])
        ),
        "actual_arrival_date": date_value(first_value(row, header_map, ["erkezes napja"])),
        "has_arrived": bool_value(first_value(row, header_map, ["megerkezett"])),
        "shipping_cost": number_value(
            first_value(
                row,
                header_map,
                ["szallitasi ktg", "netto szallitasi ktg eur huf", "szallitasi koltseg"],
            )
        ),
        "shipping_date": date_value(first_value(row, header_map, ["elszallitas napja"])),
        "supplier_invoice_number": supplier_invoice_number,
        "supplier_invoice_paid": bool_value(
            first_value(
                row,
                header_map,
                [
                    "beszallito szamla fizetve",
                    "beszallito elolegszamla fizetve",
                    "beszallito vegszamla fizetve",
                ],
            )
        ),
        "invoice_value": invoice_value,
        "shipping_invoice_number": shipping_invoice_number,
        "shipping_invoice_paid": bool_value(
            first_value(row, header_map, ["szallitmanyozo szamla fizetve"])
        ),
    }


def parse_unassigned_invoice_row(sheet_name, row_number, row, header_map):
    invoice_number = clean_string(first_value(row, header_map, ["szamlaszam"]))
    partner = clean_string(first_value(row, header_map, ["partner"]))
    description = clean_string(first_value(row, header_map, ["megnevezes"]))
    line_gross_amount_huf = number_value(
        first_value(row, header_map, ["szamla sor brutto osszeg huf"])
    )
    quantity = number_value(first_value(row, header_map, ["mennyiseg"]))
    unit_price_huf = number_value(first_value(row, header_map, ["egysegar huf"]))
    net_amount_huf = number_value(
        first_value(row, header_map, ["szamla sor netto osszeg huf"])
    )
    if not any([invoice_number, partner, description, line_gross_amount_huf]):
        return None

    return {
        "source_sheet": sheet_name,
        "source_row_number": row_number,
        "invoice_number": invoice_number,
        "partner": partner,
        "invoice_date": date_value(first_value(row, header_map, ["szamla kelte"])),
        "accounting_fulfillment_date": date_value(
            first_value(row, header_map, ["szamviteli teljesites datuma"])
        ),
        "payment_deadline": date_value(first_value(row, header_map, ["fizetesi hatarido"])),
        "gross_amount_huf": number_value(
            first_value(row, header_map, ["brutto osszeg huf"])
        ),
        "currency": clean_string(first_value(row, header_map, ["penznem", "deviza"])),
        "description": description,
        "quantity": quantity,
        "unit_price_huf": unit_price_huf,
        "net_amount_huf": net_amount_huf
        if net_amount_huf is not None
        else calculate_line_net_amount(quantity, unit_price_huf),
        "vat_amount_huf": number_value(
            first_value(row, header_map, ["szamla sor afa osszeg huf"])
        ),
        "line_gross_amount_huf": line_gross_amount_huf,
        "assignment_status": "unassigned",
    }


def import_parsed_workbook(summary, import_batch_id, user_id):
    from models import Device, Location, Project, UnassignedInvoiceItem

    created = 0
    skipped = 0
    updated = 0
    imported_at = datetime.now(timezone.utc)
    warehouse = Location.query.filter_by(
        name="Fő raktár",
        archived_at=None,
    ).first()
    if warehouse is None:
        warehouse = Location(name="Fő raktár", location_type="warehouse")
        db.session.add(warehouse)
        db.session.flush()

    for row in summary["inventory_rows"]:
        existing = Device.query.filter_by(
            source_sheet=row["source_sheet"],
            asset_tag=row["asset_tag"],
            product_name=row["product_name"],
            supplier_invoice_number=row["supplier_invoice_number"],
        ).first()
        asset_tag_conflict = Device.query.filter_by(asset_tag=row["asset_tag"]).first()
        if existing or asset_tag_conflict:
            skipped += 1
            continue

        project = get_or_create_project(row.get("project_code"))
        device = Device(
            asset_tag=row["asset_tag"],
            serial_number=row["serial_number"],
            device_type=row["device_type"],
            manufacturer=row["manufacturer"],
            model=row["model"],
            product_name=row["product_name"],
            subtype_note=row["subtype_note"],
            supplier_manufacturer=row["supplier_manufacturer"],
            version=row["version"],
            quantity=row["quantity"],
            unit_net_price=row["unit_net_price"],
            currency=row["currency"],
            huf_value=row["huf_value"]
            if row["huf_value"] is not None
            else calculate_huf_value(row["quantity"], row["unit_net_price"], row["currency"]),
            assignment_quantity=row["quantity"],
            assignment_notes=row["notes"],
            order_date=row["order_date"],
            is_ordered=row["is_ordered"],
            planned_arrival_date=row["planned_arrival_date"],
            actual_arrival_date=row["actual_arrival_date"],
            has_arrived=row["has_arrived"],
            shipping_cost=row["shipping_cost"],
            shipping_date=row["shipping_date"],
            supplier_invoice_number=row["supplier_invoice_number"],
            supplier_invoice_paid=row["supplier_invoice_paid"],
            invoice_value=row["invoice_value"],
            shipping_invoice_number=row["shipping_invoice_number"],
            shipping_invoice_paid=row["shipping_invoice_paid"],
            location_id=warehouse.id,
            status="IN_STOCK",
            source_sheet=row["source_sheet"],
            source_row_number=row["source_row_number"],
            import_batch_id=import_batch_id,
            imported_at=imported_at,
        )
        db.session.add(device)
        db.session.flush()
        create_movement(
            device=device,
            movement_type="INBOUND",
            quantity=device.quantity,
            to_location_id=warehouse.id,
            notes=f"Excel import: {row['source_sheet']} #{row['source_row_number']}",
            user_id=user_id,
        )
        apply_device_state(
            device,
            "INBOUND",
            warehouse.id,
            None,
            quantity=device.quantity,
        )
        if project is not None:
            source_balance = infer_bulk_source_balance(
                device,
                "RESERVE",
                device.quantity,
            )
            create_movement(
                device=device,
                movement_type="RESERVE",
                quantity=device.quantity,
                project_id=project.id,
                source_balance=source_balance,
                notes=f"Excel import projektfoglalás: {project.code}",
                user_id=user_id,
            )
            apply_device_state(
                device,
                "RESERVE",
                None,
                project.id,
                quantity=device.quantity,
                source_balance=source_balance,
            )
        created += 1

    for row in summary["invoice_rows"]:
        existing = UnassignedInvoiceItem.query.filter_by(
            invoice_number=row["invoice_number"],
            partner=row["partner"],
            description=row["description"],
            line_gross_amount_huf=row["line_gross_amount_huf"],
        ).first()
        if existing:
            skipped += 1
            continue
        item = UnassignedInvoiceItem(
            invoice_number=row["invoice_number"],
            partner=row["partner"],
            invoice_date=row["invoice_date"],
            accounting_fulfillment_date=row["accounting_fulfillment_date"],
            payment_deadline=row["payment_deadline"],
            gross_amount_huf=row["gross_amount_huf"],
            currency=row["currency"],
            description=row["description"],
            quantity=row["quantity"],
            unit_price_huf=row["unit_price_huf"],
            net_amount_huf=row["net_amount_huf"],
            vat_amount_huf=row["vat_amount_huf"],
            line_gross_amount_huf=row["line_gross_amount_huf"],
            assignment_status=row["assignment_status"],
            source_sheet=row["source_sheet"],
            source_row_number=row["source_row_number"],
            import_batch_id=import_batch_id,
            imported_at=imported_at,
        )
        db.session.add(item)
        created += 1

    return {
        "created_count": created,
        "skipped_count": skipped,
        "updated_count": updated,
    }


def get_or_create_project(project_code):
    from models import Project

    if not project_code:
        return None
    project = Project.query.filter_by(code=project_code).first()
    if project:
        return project
    project = Project(name=project_code, code=project_code, status="active")
    db.session.add(project)
    db.session.flush()
    return project


def normalize_key(value):
    text = "" if value is None else str(value)
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.lower()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def find_header_map(sheet, max_scan_rows=12):
    for row_number, cells in enumerate(
        sheet.iter_rows(min_row=1, max_row=max_scan_rows, values_only=True), start=1
    ):
        normalized = [normalize_key(value) for value in cells]
        if any(
            key in normalized
            for key in ["termek", "tipus", "szamlaszam", "projekt kod", "mennyiseg"]
        ):
            header_map = {}
            for index, key in enumerate(normalized):
                if key:
                    header_map.setdefault(key, []).append(index)
            return row_number, header_map
    return None, {}


def iter_sheet_rows(sheet, start_row):
    for row_number, row in enumerate(
        sheet.iter_rows(min_row=start_row, values_only=True), start=start_row
    ):
        yield row_number, list(row)


def first_value(row, header_map, candidates):
    for candidate in candidates:
        key = normalize_key(candidate)
        for index in header_map.get(key, []):
            if index < len(row):
                value = row[index]
                if meaningful_value(value):
                    return value
    return None


def meaningful_value(value):
    if value is None:
        return False
    if isinstance(value, str):
        stripped = value.strip()
        return bool(stripped) and stripped not in {"-", "#"}
    return True


def clean_string(value):
    if not meaningful_value(value):
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    return str(value).strip()


def clean_project_code(value):
    text = clean_string(value)
    if not text:
        return None
    return text


def date_value(value):
    if not meaningful_value(value):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip().replace(".", "-")
    for fmt in ("%Y-%m-%d", "%Y-%m-%d-", "%Y %m %d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    return None


def number_value(value):
    if not meaningful_value(value):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    text = text.replace("\xa0", "").replace(" ", "").replace(",", "")
    try:
        return float(text)
    except ValueError:
        return None


def bool_value(value):
    if not meaningful_value(value):
        return None
    text = normalize_key(value)
    if text in {"igen", "yes", "true", "1", "ok"}:
        return True
    if text in {"nem", "no", "false", "0", "nok"}:
        return False
    return None


def infer_device_type(sheet_name, type_value, product_name):
    sheet_key = normalize_key(sheet_name)
    if sheet_key in {"tolto", "toltok", "bmw tolto"}:
        return "EV charger"
    if sheet_key == "matricak":
        return "Sticker"
    if sheet_key == "kamera":
        return "Camera"
    if sheet_key == "kioszk":
        return "Kiosk"
    if sheet_key == "nyito":
        return "Opener"
    if sheet_key == "egyeb":
        return "Other"
    value = clean_string(type_value or product_name)
    return value or "Other"


def preview_row(kind, row):
    if kind == "Gazdátlan számlasor":
        return {
            "tipus": kind,
            "munkalap": row["source_sheet"],
            "sor": row["source_row_number"],
            "azonosito": row["invoice_number"],
            "nev": row["description"],
            "partner": row["partner"],
            "ertek": row["line_gross_amount_huf"] or row["gross_amount_huf"],
        }
    return {
        "tipus": kind,
        "munkalap": row["source_sheet"],
        "sor": row["source_row_number"],
        "azonosito": row["asset_tag"],
        "nev": row["product_name"],
        "partner": row["supplier_manufacturer"],
        "ertek": row["huf_value"],
    }


def create_movement(
    device,
    movement_type,
    user_id,
    from_location_id=None,
    to_location_id=None,
    project_id=None,
    quantity=None,
    unit=None,
    unit_id=None,
    from_project_id=None,
    to_project_id=None,
    reversal_of_movement_id=None,
    source_balance=None,
    notes="",
):
    from models import DeviceUnit, StockMovement

    if unit is None and unit_id is not None:
        unit = db.session.get(DeviceUnit, unit_id)
    if unit is not None and unit.device_id != device.id:
        raise ValueError("Az eszközpéldány nem ehhez a terméktételhez tartozik.")

    if (
        device.tracking_mode == "bulk"
        and source_balance is None
        and movement_type != "INBOUND"
    ):
        source_balance = infer_bulk_source_balance(device, movement_type, quantity)
    subject = unit or source_balance or device
    if unit is not None:
        quantity = 1
    elif quantity is None:
        quantity = device.quantity

    external_bulk_inbound = (
        device.tracking_mode == "bulk"
        and unit is None
        and source_balance is None
        and movement_type == "INBOUND"
    )
    if source_balance is not None:
        from_location_id = source_balance.location_id
        from_project_id = source_balance.project_id
    elif external_bulk_inbound:
        from_location_id = None
        from_project_id = None
    elif from_location_id is None and movement_type != "INBOUND":
        from_location_id = subject.location_id
    if from_project_id is None:
        from_project_id = subject.project_id
    from_status = subject.status if movement_type != "INBOUND" or source_balance else None
    to_status = movement_target_status(from_status, movement_type)
    effective_location_id, effective_project_id = movement_target_dimensions(
        movement_type,
        from_location_id,
        from_project_id,
        to_location_id,
        project_id if to_project_id is None else to_project_id,
    )
    to_location_id = effective_location_id
    to_project_id = effective_project_id

    movement = StockMovement(
        device=device,
        unit=unit,
        movement_type=movement_type,
        quantity=quantity,
        from_location_id=from_location_id,
        to_location_id=to_location_id,
        project_id=to_project_id,
        from_project_id=from_project_id,
        to_project_id=to_project_id,
        reversal_of_movement_id=reversal_of_movement_id,
        from_status=from_status,
        to_status=to_status,
        notes=notes,
        created_by_id=user_id,
    )
    db.session.add(movement)
    return movement


def validate_movement(
    device,
    movement_type,
    to_location_id=None,
    project_id=None,
    quantity=None,
    unit=None,
    reversal_of_movement_id=None,
    from_location_id=None,
    source_balance=None,
):
    from models import BulkStockBalance, Location, MOVEMENT_TYPES, Project, StockMovement

    if movement_type not in MOVEMENT_TYPES:
        return "Érvénytelen mozgástípus."

    if device.tracking_mode == "unit":
        if unit is None:
            return "Egyedi követésű tételnél konkrét eszközpéldány kiválasztása kötelező."
        if unit.device_id != device.id:
            return "A kiválasztott eszközpéldány nem ehhez a terméktételhez tartozik."
        if quantity not in (None, 1, 1.0):
            return "Egyedi követésű eszközpéldány mozgási mennyisége csak 1 lehet."
    else:
        if unit is not None:
            return "Mennyiségi követésű tételhez nem választható egyedi eszközpéldány."
        if quantity is None or quantity <= 0:
            return "Mennyiségi követésű tételnél pozitív mozgási mennyiség megadása kötelező."
        if source_balance is not None:
            if not isinstance(source_balance, BulkStockBalance):
                return "Érvénytelen forrás készletegyenleg."
            if source_balance.device_id != device.id:
                return "A forrás készletegyenleg nem ehhez a tételhez tartozik."
            if source_balance.quantity <= 0:
                return "A kiválasztott készletegyenlegből nincs elérhető mennyiség."
            if quantity > source_balance.quantity + 1e-9:
                return (
                    f"Legfeljebb {format_number(source_balance.quantity)} mozgatható "
                    "a kiválasztott készletegyenlegből."
                )
        elif movement_type != "INBOUND":
            return "Bulk mozgáshoz forrás készletegyenleg kiválasztása kötelező."

    if reversal_of_movement_id is not None:
        original = db.session.get(StockMovement, reversal_of_movement_id)
        if original is None:
            return "A visszafordítandó készletmozgás nem található."
        if original.device_id != device.id or original.unit_id != (unit.id if unit else None):
            return "Csak ugyanahhoz a tételhez vagy példányhoz tartozó mozgás fordítható vissza."
        if StockMovement.query.filter_by(reversal_of_movement_id=original.id).first():
            return "Ezt a készletmozgást már visszafordították."

    subject = unit or source_balance or device
    if (
        source_balance is None
        and from_location_id is not None
        and movement_type != "INBOUND"
        and subject.location_id != from_location_id
    ):
        return "A megadott forrás készlethely nem egyezik a tétel aktuális készlethelyével."
    target_location = (
        db.session.get(Location, to_location_id) if to_location_id is not None else None
    )
    if to_location_id is not None and target_location is None:
        return "A megadott cél készlethely nem található."
    if reversal_of_movement_id is None and target_location is not None and (
        target_location.archived_at is not None
        or target_location.location_type not in LOGISTIC_LOCATION_TYPES
    ):
        return "Készletmozgás célja csak aktív logisztikai készlethely lehet."
    if project_id is not None and db.session.get(Project, project_id) is None:
        return "A megadott projekt nem található."
    if source_balance is None and device.tracking_mode == "bulk" and movement_type == "INBOUND":
        current_status = None
    else:
        current_status = subject.status
    if current_status == "SCRAPPED":
        return "Selejtezett eszköz nem mozgatható tovább."

    state_error = inventory_state_error(
        current_status,
        subject.location_id,
        subject.project_id,
    )
    if state_error:
        return f"A jelenlegi készletállapot inkonzisztens: {state_error}"

    allowed_statuses = movement_allowed_statuses()
    allowed = allowed_statuses[movement_type]
    if (
        not (movement_type == "INBOUND" and current_status is None)
        and allowed is not None
        and current_status not in allowed
    ):
        readable = ", ".join(status_label(status) for status in sorted(allowed))
        return (
            f"A(z) {movement_type_label(movement_type)} nem engedélyezett "
            f"a(z) {device.asset_tag} eszköznél, mert jelenlegi státusza: "
            f"{status_label(current_status)}. Engedélyezett aktuális státusz: "
            f"{readable}."
        )
    if movement_type in {"RETURN", "INBOUND", "TRANSFER", "SERVICE"} and not to_location_id:
        return f"A(z) {movement_type_label(movement_type)} művelethez cél készlethely megadása kötelező."
    if movement_type == "RESERVE":
        if not project_id:
            return "Előfoglaláshoz projekt megadása kötelező."
        if not subject.location_id:
            return "Csak készlethelyen lévő eszköz vagy mennyiség foglalható elő."
    if movement_type == "ISSUE" and not (project_id or subject.project_id):
        return "Kiadáshoz projekt megadása kötelező."
    if (
        movement_type == "ISSUE"
        and current_status == "RESERVED"
        and project_id is not None
        and project_id != subject.project_id
    ):
        return (
            "Az előfoglalás másik projekthez tartozik. Előbb oldd fel a foglalást, "
            "vagy add ki ugyanarra a projektre."
        )
    if movement_type == "INSTALL":
        if not (project_id or subject.project_id):
            return "Telepítéshez projekt megadása kötelező."
        if (
            subject.project_id is not None
            and project_id is not None
            and project_id != subject.project_id
        ):
            return "Telepítéskor az eszköz nem vihető át másik projektre."
    if movement_type == "SERVICE" and target_location.location_type not in {
        "warehouse",
        "service",
    }:
        return "Szervizbe küldés célja csak raktár vagy szerviz típusú készlethely lehet."
    if movement_type == "RELEASE" and project_id not in {None, subject.project_id}:
        return "Foglalás feloldásakor nem választható másik projekt."
    if (
        movement_type == "TRANSFER"
        and to_location_id == subject.location_id
    ):
        return "Áthelyezéshez az aktuálistól eltérő cél készlethely szükséges."
    target_status = movement_target_status(current_status, movement_type)
    target_location_id, target_project_id = movement_target_dimensions(
        movement_type,
        subject.location_id,
        subject.project_id,
        to_location_id,
        project_id,
    )
    target_error = inventory_state_error(
        target_status,
        target_location_id,
        target_project_id,
    )
    if target_error:
        return f"A művelet érvénytelen célállapotot hozna létre: {target_error}"
    return None


def apply_device_state(
    device,
    movement_type,
    to_location_id=None,
    project_id=None,
    unit=None,
    quantity=None,
    source_balance=None,
):
    if device.tracking_mode == "bulk" and unit is None:
        if source_balance is None and movement_type != "INBOUND":
            source_balance = infer_bulk_source_balance(
                device, movement_type, quantity
            )
        if quantity is None:
            quantity = device.quantity
        apply_bulk_movement(
            device,
            movement_type,
            quantity,
            source_balance,
            to_location_id,
            project_id,
        )
        return

    subject = unit or device
    subject.updated_at = datetime.now(timezone.utc)

    target_status = movement_target_status(subject.status, movement_type)
    target_location_id, target_project_id = movement_target_dimensions(
        movement_type,
        subject.location_id,
        subject.project_id,
        to_location_id,
        project_id,
    )
    subject.status = target_status
    subject.location_id = target_location_id
    subject.project_id = target_project_id


def movement_target_status(from_status, movement_type):
    status_by_movement = {
        "INBOUND": "IN_STOCK",
        "RESERVE": "RESERVED",
        "ISSUE": "ISSUED",
        "INSTALL": "INSTALLED",
        "RETURN": "RETURNED",
        "SERVICE": "IN_SERVICE",
        "SCRAP": "SCRAPPED",
        "RELEASE": "IN_STOCK",
    }
    return status_by_movement.get(movement_type, from_status)


def movement_target_dimensions(
    movement_type,
    from_location_id,
    from_project_id,
    to_location_id,
    project_id,
):
    if movement_type in {"INBOUND", "RETURN", "TRANSFER", "SERVICE"}:
        return to_location_id, None
    if movement_type == "RESERVE":
        return from_location_id, project_id
    if movement_type == "RELEASE":
        return from_location_id, None
    if movement_type in {"ISSUE", "INSTALL"}:
        return None, project_id if project_id is not None else from_project_id
    if movement_type == "SCRAP":
        return None, None
    return from_location_id, from_project_id


def inventory_state_error(status, location_id, project_id):
    if status == "IN_STOCK":
        if location_id is None:
            return "raktáron lévő tételhez készlethely szükséges"
        if project_id is not None:
            return "raktáron lévő tételhez nem tartozhat aktív projekt"
    elif status == "RESERVED":
        if location_id is None or project_id is None:
            return "előfoglaláshoz készlethely és projekt is szükséges"
    elif status in {"ISSUED", "INSTALLED"}:
        if project_id is None:
            return "kiadott vagy telepített tételhez projekt szükséges"
        if location_id is not None:
            return "kiadott vagy telepített tétel nem maradhat készlethelyen"
    elif status in {"RETURNED", "IN_SERVICE"}:
        if location_id is None:
            return "visszavett vagy szervizben lévő tételhez készlethely szükséges"
        if project_id is not None:
            return "visszavett vagy szervizben lévő tételhez nem tartozhat aktív projekt"
    elif status == "SCRAPPED":
        if location_id is not None or project_id is not None:
            return "selejtezett tételnek nem lehet aktív készlethelye vagy projektje"
    return None


def active_bulk_balances(device):
    if device.tracking_mode != "bulk":
        return []
    return [
        balance
        for balance in device.bulk_balances
        if balance.quantity is not None and balance.quantity > 1e-9
    ]


def infer_bulk_source_balance(device, movement_type, quantity=None):
    allowed_statuses = movement_allowed_statuses().get(movement_type)
    candidates = [
        balance
        for balance in active_bulk_balances(device)
        if balance.status != "SCRAPPED"
        and (allowed_statuses is None or balance.status in allowed_statuses)
        and (quantity is None or balance.quantity + 1e-9 >= quantity)
    ]
    if len(candidates) == 1:
        return candidates[0]
    if quantity is None and candidates:
        return max(candidates, key=lambda balance: balance.quantity)
    return None


def find_or_create_bulk_balance(device, status, location_id, project_id):
    from models import BulkStockBalance

    balance = BulkStockBalance.query.filter_by(
        device_id=device.id,
        status=status,
        location_id=location_id,
        project_id=project_id,
    ).first()
    if balance is None:
        balance = BulkStockBalance(
            device=device,
            status=status,
            quantity=0,
            location_id=location_id,
            project_id=project_id,
        )
        db.session.add(balance)
    return balance


def apply_bulk_movement(
    device,
    movement_type,
    quantity,
    source_balance,
    to_location_id,
    project_id,
):
    from models import BulkStockBalance

    if quantity is None or quantity <= 0:
        raise ValueError("Pozitív bulk mozgási mennyiség szükséges.")

    if source_balance is not None and source_balance.id is not None:
        source_balance = (
            BulkStockBalance.query.filter_by(id=source_balance.id)
            .with_for_update()
            .one()
        )
        if quantity > source_balance.quantity + 1e-9:
            raise ValueError(
                f"Legfeljebb {format_number(source_balance.quantity)} mozgatható "
                "a kiválasztott készletegyenlegből."
            )

    source_status = source_balance.status if source_balance else None
    target_status = movement_target_status(source_status, movement_type)
    source_location_id = source_balance.location_id if source_balance else None
    source_project_id = source_balance.project_id if source_balance else None

    if source_balance is not None:
        source_balance.quantity -= quantity
        if source_balance.quantity < -1e-9:
            raise ValueError("A bulk készletegyenleg nem mehet negatívba.")
        if abs(source_balance.quantity) <= 1e-9:
            source_balance.quantity = 0
    else:
        existing_quantity = sum(
            balance.quantity
            for balance in device.bulk_balances
            if balance.quantity is not None and balance.quantity > 1e-9
        )
        if existing_quantity > 1e-9:
            device.quantity = (device.quantity or existing_quantity) + quantity

    target_location_id, target_project_id = movement_target_dimensions(
        movement_type,
        source_location_id,
        source_project_id,
        to_location_id,
        project_id,
    )
    target = find_or_create_bulk_balance(
        device,
        target_status,
        target_location_id,
        target_project_id,
    )
    target.quantity += quantity
    sync_bulk_device_legacy_state(device)


def sync_bulk_device_legacy_state(device):
    balances = sorted(
        active_bulk_balances(device),
        key=lambda balance: balance.quantity,
        reverse=True,
    )
    if not balances:
        return
    representative = balances[0]
    device.status = representative.status
    device.location_id = representative.location_id
    device.project_id = representative.project_id
    device.updated_at = datetime.now(timezone.utc)


app = create_app()


if __name__ == "__main__":
    app.run(debug=True)
